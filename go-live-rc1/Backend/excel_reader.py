"""
SOBHA COLLECTIONS INTELLIGENCE PLATFORM
excel_reader.py — xlsx sheet reader utility

This module encapsulates the logic for reading Excel worksheets from the
R‑series data pipeline. It provides a single function, ``read_sheet``,
that accepts a ``Path`` and a sheet identifier (either an integer
index or a string name) and returns a list of row dictionaries or
``None`` if the workbook cannot be read.  Reading spreadsheets
directly in ``data_loader.py`` made the module more complex and
harder to test.  Splitting this logic into its own file clarifies
responsibilities and enables independent unit testing.

Responsibilities:
  - Open an ``.xlsx`` file in read‑only mode using ``openpyxl``.
  - Resolve sheet identifiers (index or name) gracefully.
  - Convert the header row into normalised column names.
  - Produce a list of dictionaries mapping header names to cell values.
  - Skip blank rows and handle missing files gracefully.

This function is intentionally stateless and will never mutate global
state.  On failure, it returns ``None`` instead of raising so that
callers can decide on graceful degradation policies.
"""

from __future__ import annotations

import logging
from pathlib import Path
from typing import Any, Optional, List, Dict

import openpyxl

logger = logging.getLogger(__name__)


def read_sheet(file_path: Path, sheet_name: Any) -> Optional[List[Dict[str, Any]]]:
    """Read a single xlsx sheet into a list of row dicts.

    Uses openpyxl in read‑only mode for performance.  Accepts either an
    integer index or a string sheet name; if the supplied name is not
    found the workbook's active sheet will be used.  Returns ``None``
    on failure to open the file or locate the sheet.  Empty sheets
    yield an empty list rather than ``None``.

    Args:
        file_path: Path to the ``.xlsx`` file on disk.
        sheet_name: Either a zero‑based index or a string corresponding
            to the worksheet name.

    Returns:
        A list of dictionaries mapping header names to row values, or
        ``None`` if the workbook cannot be read.
    """
    try:
        wb = openpyxl.load_workbook(str(file_path), read_only=True, data_only=True)

        # Resolve sheet name or index to a worksheet
        if isinstance(sheet_name, int):
            ws = wb.worksheets[sheet_name]
        elif isinstance(sheet_name, str):
            if sheet_name not in wb.sheetnames:
                logger.warning(
                    "Sheet '%s' not found in %s. Available: %s", sheet_name, file_path.name, wb.sheetnames
                )
                wb.close()
                return None
            ws = wb[sheet_name]
        else:
            # Fallback to the active worksheet
            ws = wb.active

        rows = list(ws.iter_rows(values_only=True))
        wb.close()

        if not rows:
            return []

        # Normalise header names: strip whitespace and coerce None
        headers = [str(h).strip() if h is not None else f"col_{i}" for i, h in enumerate(rows[0])]
        result: List[Dict[str, Any]] = []
        for row in rows[1:]:
            # Skip completely blank rows
            if all(v is None for v in row):
                continue
            result.append(dict(zip(headers, row)))
        return result

    except FileNotFoundError:
        logger.info("File not found (graceful skip): %s", file_path.name)
        return None
    except Exception as exc:
        logger.warning("Failed to read %s sheet=%s: %s", file_path.name, sheet_name, exc)
        return None
