"""
SCIP Batch 18 semantic fact builder.

Builds lineaged fact records from existing Batch 13 adapters and from report-specific
summary scans for rich action/risk reports. It does not publish metrics directly.
"""
from __future__ import annotations

from collections import Counter, defaultdict
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional

from openpyxl import load_workbook

from analytics_models import FactRecord
from analytics_utils import date_from_filename, make_hash, norm, norm_key, safe_div, to_float, utc_now


class SemanticFactBuilder:
    ADAPTER_REPORTS = {"R02", "R04", "R08", "R18", "R36"}

    def build_facts(self, report_code: str, path: str | Path) -> List[FactRecord]:
        code = report_code.upper()
        if code in self.ADAPTER_REPORTS:
            return self._build_from_batch13_adapter(code, path)
        dispatch = {
            "R01": self._build_r01,
            "R09": self._build_r09,
            "R10": self._build_r10,
            "R16": self._build_r16,
            "R17": self._build_r17,
            "R20": self._build_r20,
            "R30": self._build_r30,
            "R31": self._build_r31,
            "R32": self._build_r32,
            "R34": self._build_r34,
            "R38": self._build_r38,
        }.get(code)
        if dispatch:
            return dispatch(Path(path))
        return []

    def _fact(self, *, fact_type: str, source_code: str, source_file: str, sheet: str, grain: str,
              measure_name: str, value: Any, unit: str = "AED", reporting_basis: str = "",
              validation_status: str = "passed", confidence_state: str = "live_validated",
              cell_or_range: Optional[str] = None, extraction_rule: Optional[str] = None,
              snapshot_date: Optional[str] = None, entity_key: Optional[str] = None,
              project_key: Optional[str] = None, customer_key: Optional[str] = None,
              unit_key: Optional[str] = None, collector_key: Optional[str] = None,
              dimensions: Optional[Dict[str, Any]] = None, metadata: Optional[Dict[str, Any]] = None) -> FactRecord:
        payload = {
            "fact_type": fact_type, "source_code": source_code, "source_file": source_file,
            "sheet": sheet, "grain": grain, "measure_name": measure_name, "value": value,
            "entity_key": entity_key, "project_key": project_key, "customer_key": customer_key,
            "unit_key": unit_key, "collector_key": collector_key, "snapshot_date": snapshot_date,
            "cell_or_range": cell_or_range, "extraction_rule": extraction_rule,
        }
        lineage_hash = make_hash(payload)
        return FactRecord(
            fact_id=f"{source_code}:{fact_type}:{measure_name}:{lineage_hash}",
            fact_type=fact_type,
            source_code=source_code,
            source_file=source_file,
            sheet=sheet,
            grain=grain,
            measure_name=measure_name,
            value=value,
            unit=unit,
            reporting_basis=reporting_basis,
            validation_status=validation_status,
            confidence_state=confidence_state,
            lineage_hash=lineage_hash,
            loaded_at=utc_now(),
            cell_or_range=cell_or_range,
            extraction_rule=extraction_rule,
            snapshot_date=snapshot_date,
            entity_key=entity_key,
            project_key=project_key,
            customer_key=customer_key,
            unit_key=unit_key,
            collector_key=collector_key,
            lineage_refs=[lineage_hash],
            dimensions=dimensions or {},
            metadata=metadata or {},
        )

    def _build_from_batch13_adapter(self, report_code: str, path: str | Path) -> List[FactRecord]:
        try:
            import source_adapters  # type: ignore
        except Exception:
            return []
        p = Path(path)
        try:
            result = source_adapters.run_adapter(report_code, p).to_dict()
        except Exception:
            return []
        metrics = result.get("metrics", {}) or {}
        lineage = result.get("lineage", {}) or {}
        snapshot = metrics.get("SNAPSHOT_DATE") or date_from_filename(p)
        facts: List[FactRecord] = []
        for key, value in metrics.items():
            if key.startswith("_") or isinstance(value, (dict, list)):
                continue
            if value is None:
                continue
            fact_type, grain, unit, basis = self._adapter_metric_fact_metadata(report_code, key)
            lin = lineage.get(key, {}) or {}
            facts.append(self._fact(
                fact_type=fact_type,
                source_code=report_code,
                source_file=p.name,
                sheet=lin.get("sheet", "adapter"),
                grain=grain,
                measure_name=key,
                value=value,
                unit=unit,
                reporting_basis=lin.get("reporting_basis") or basis,
                validation_status=lin.get("validation_status", "passed"),
                confidence_state=lin.get("confidence_state", "live_validated"),
                cell_or_range=lin.get("cell_or_range"),
                extraction_rule=lin.get("extraction_method", "batch13_adapter"),
                snapshot_date=lin.get("snapshot_date") or snapshot,
                entity_key=lin.get("entity_scope"),
                metadata={"adapter_status": result.get("status")},
            ))
        # Also carry row-level adapter output as facts where useful.
        for idx, row in enumerate(result.get("rows", []) or []):
            if not isinstance(row, dict):
                continue
            for measure, value in row.items():
                if measure.endswith("_aed") and to_float(value) is not None:
                    facts.append(self._fact(
                        fact_type=self._row_fact_type(report_code), source_code=report_code, source_file=p.name,
                        sheet=row.get("source_sheet") or row.get("sheet") or "adapter_rows", grain="row_detail",
                        measure_name=measure, value=to_float(value), reporting_basis=f"{report_code} adapter row detail",
                        snapshot_date=snapshot, entity_key=row.get("entity_key"), project_key=row.get("project_key") or row.get("project"),
                        collector_key=row.get("collector"), unit_key=row.get("unit"),
                        cell_or_range=str(row.get("source_cells") or row.get("source_row") or idx),
                        extraction_rule="batch13_adapter_rows", metadata=row,
                    ))
        return facts

    @staticmethod
    def _adapter_metric_fact_metadata(report_code: str, key: str) -> tuple[str, str, str, str]:
        if report_code == "R18":
            if key.startswith("ageing_"):
                return "fact_overdue_ageing", "entity_snapshot", "AED", "R18 ageing buckets"
            return "fact_overdue_snapshot", "entity_snapshot", "AED" if key != "SNAPSHOT_DATE" else "date", "R18 overdue source"
        if report_code == "R04":
            if key.startswith("DAILY_"):
                return "fact_collections_daily", "date_entity", "AED", "R04 Finance daily actuals"
            return "fact_finance_mtd_actuals", "month_entity", "AED", "R04 Finance daily actuals"
        if report_code == "R02":
            if "actual" in key.lower():
                return "fact_mdo_actuals", "month_entity", "AED", "R02 MDO actual basis"
            return "fact_targets_monthly", "month_entity", "AED", "R02 MDO target basis"
        if report_code == "R08":
            return "fact_advance_cyfy", "snapshot_entity", "AED", "R08 Advance CY/FY basis"
        if report_code == "R36":
            return "fact_forward_collectible_calendar", "entity_collection_year", "AED", "R36 milestone cohort calendar"
        return "fact_source_metric", "source_metric", "AED", report_code

    @staticmethod
    def _row_fact_type(report_code: str) -> str:
        return {
            "R36": "fact_milestone_cohort",
            "R18": "fact_entity_overdue_rollup",
            "R04": "fact_collections_daily",
        }.get(report_code, "fact_source_row")

    def _open(self, path: Path):
        return load_workbook(path, read_only=True, data_only=True)

    def _build_r10(self, path: Path) -> List[FactRecord]:
        facts: List[FactRecord] = []
        snapshot = date_from_filename(path)
        wb = self._open(path)
        try:
            if "Allocation" in wb.sheetnames:
                ws = wb["Allocation"]
                sums = defaultdict(float); rows = 0; high_no_touch = 0; high_no_touch_amt = 0.0
                for ridx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                    if not row or not row[4]:
                        continue
                    rows += 1
                    od_today = to_float(row[6], 0.0) or 0.0
                    overdue_eom = to_float(row[7], 0.0) or 0.0
                    comm = to_float(row[16], 0.0) or 0.0
                    sums["od_today_aed"] += od_today
                    sums["overdue_eom_aed"] += overdue_eom
                    sums["amount_in_tracker_aed"] += to_float(row[20], 0.0) or 0.0
                    sums["receipt_generated_aed"] += to_float(row[21], 0.0) or 0.0
                    sums["receipt_not_part_tracker_aed"] += to_float(row[22], 0.0) or 0.0
                    sums["overdue_eom_plus_ms_aed"] += to_float(row[23], 0.0) or 0.0
                    sums["communication_count"] += comm
                    if overdue_eom >= 30000 and comm <= 0:
                        high_no_touch += 1
                        high_no_touch_amt += overdue_eom
                for name, value in sums.items():
                    facts.append(self._fact(fact_type="fact_account_allocation", source_code="R10", source_file=path.name,
                        sheet="Allocation", grain="account_unit_snapshot", measure_name=name, value=value,
                        unit="count" if name.endswith("count") else "AED", reporting_basis="R10 Allocation",
                        snapshot_date=snapshot, cell_or_range="Allocation aggregate", extraction_rule="allocation_summary_scan"))
                facts.append(self._fact(fact_type="fact_account_communication", source_code="R10", source_file=path.name,
                    sheet="Allocation", grain="account_unit_snapshot", measure_name="high_value_no_touch_count", value=high_no_touch,
                    unit="count", reporting_basis="R10 Allocation", snapshot_date=snapshot, extraction_rule="overdue_eom>=30000 and communication_count=0"))
                facts.append(self._fact(fact_type="fact_account_communication", source_code="R10", source_file=path.name,
                    sheet="Allocation", grain="account_unit_snapshot", measure_name="high_value_no_touch_amount_aed", value=high_no_touch_amt,
                    reporting_basis="R10 Allocation", snapshot_date=snapshot, extraction_rule="overdue_eom>=30000 and communication_count=0"))
                facts.append(self._fact(fact_type="fact_account_allocation", source_code="R10", source_file=path.name,
                    sheet="Allocation", grain="account_unit_snapshot", measure_name="allocation_rows", value=rows, unit="count",
                    reporting_basis="R10 Allocation", snapshot_date=snapshot, extraction_rule="row_count"))
            if "PTP Data" in wb.sheetnames:
                ws = wb["PTP Data"]
                count = 0; amount = 0.0; by_channel = Counter()
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if not row or not row[0]:
                        continue
                    count += 1
                    # PTP amount may move across exports; scan last numeric cells as fallback.
                    nums = [to_float(v) for v in row if to_float(v) is not None]
                    if nums:
                        amount += max(nums[-3:]) if len(nums) >= 3 else nums[-1]
                    by_channel[norm(row[0]) or "unknown"] += 1
                facts.append(self._fact(fact_type="fact_ptp_promise", source_code="R10", source_file=path.name,
                    sheet="PTP Data", grain="promise_snapshot", measure_name="ptp_count", value=count, unit="count",
                    reporting_basis="R10 PTP Data", snapshot_date=snapshot, extraction_rule="ptp_data_row_count"))
                facts.append(self._fact(fact_type="fact_ptp_promise", source_code="R10", source_file=path.name,
                    sheet="PTP Data", grain="promise_snapshot", measure_name="ptp_amount_estimated_aed", value=amount,
                    reporting_basis="R10 PTP Data", snapshot_date=snapshot, extraction_rule="ptp_numeric_amount_scan", metadata={"caveat": "amount column may vary; verify before board use"}))
                for ch, cnt in by_channel.items():
                    facts.append(self._fact(fact_type="fact_ptp_promise", source_code="R10", source_file=path.name,
                        sheet="PTP Data", grain="promise_channel", measure_name="ptp_channel_count", value=cnt, unit="count",
                        reporting_basis="R10 PTP Data", snapshot_date=snapshot, dimensions={"channel": ch}))
            if "Agent - Coverage" in wb.sheetnames:
                facts.extend(self._extract_simple_total_row(wb["Agent - Coverage"], path, "R10", "fact_collector_coverage", "R10 Agent Coverage", 3, [2,3,4,5,6,8,9]))
            if "Agent - Productivity" in wb.sheetnames:
                facts.extend(self._extract_simple_total_row(wb["Agent - Productivity"], path, "R10", "fact_collector_productivity", "R10 Agent Productivity", 3, [2,3,4,5,6,7,8,9,10,11]))
        finally:
            wb.close()
        return facts

    def _extract_simple_total_row(self, ws, path: Path, code: str, fact_type: str, basis: str, header_row: int, value_cols: List[int]) -> List[FactRecord]:
        headers = [norm(v) or f"col_{i}" for i, v in enumerate(next(ws.iter_rows(min_row=header_row, max_row=header_row, values_only=True)))]
        total = None
        for row in ws.iter_rows(values_only=True):
            if any(str(v).strip().lower() in {"grand total", "total"} for v in row if v is not None):
                total = row
        if total is None:
            return []
        facts: List[FactRecord] = []
        for idx in value_cols:
            if idx < len(total) and to_float(total[idx]) is not None:
                name = norm_key(headers[idx]) or f"col_{idx}"
                facts.append(self._fact(fact_type=fact_type, source_code=code, source_file=path.name, sheet=ws.title,
                    grain="collector_summary", measure_name=name, value=to_float(total[idx]), unit="pct" if "%" in headers[idx] else "AED",
                    reporting_basis=basis, snapshot_date=date_from_filename(path), cell_or_range="total row", extraction_rule="total_row_scan"))
        return facts

    def _build_r30(self, path: Path) -> List[FactRecord]:
        wb = self._open(path); facts: List[FactRecord] = []; snapshot = date_from_filename(path)
        try:
            ws = wb["Summary"]
            totals = defaultdict(float); collector_count = 0
            for row in ws.iter_rows(min_row=4, values_only=True):
                collector = norm(row[1])
                if not collector or "total" in collector.lower():
                    continue
                collector_count += 1
                names = ["allocated_units", "allocated_od_aed", "target_aed", "attempts", "contacts", "attempt_intensity", "contact_intensity", "units_paid", "total_payment_aed"]
                vals = [row[2], row[3], row[4], row[5], row[6], row[7], row[8], row[9], row[16]]
                for n, v in zip(names, vals):
                    totals[n] += to_float(v, 0.0) or 0.0
                for ch_name, v in [("self_cured_payment_aed", row[17]), ("calls_payment_aed", row[18]), ("emails_payment_aed", row[19]), ("walkin_payment_aed", row[20]), ("whatsapp_payment_aed", row[22])]:
                    totals[ch_name] += to_float(v, 0.0) or 0.0
                # row level collector facts for coaching surfaces
                facts.append(self._fact(fact_type="fact_collector_feedback", source_code="R30", source_file=path.name,
                    sheet="Summary", grain="collector_snapshot", measure_name="total_payment_aed", value=to_float(row[16], 0.0),
                    reporting_basis="R30 Monthly Feedback", snapshot_date=snapshot, collector_key=collector,
                    cell_or_range=f"collector row {collector}", extraction_rule="summary_row_scan",
                    metadata={"allocated_od_aed": to_float(row[3], 0.0), "target_aed": to_float(row[4], 0.0), "contacts": to_float(row[6], 0.0)}))
            totals["collector_count"] = collector_count
            for name, value in totals.items():
                facts.append(self._fact(fact_type="fact_collector_feedback", source_code="R30", source_file=path.name,
                    sheet="Summary", grain="collector_summary", measure_name=name, value=value,
                    unit="count" if name in {"collector_count", "allocated_units", "attempts", "contacts", "units_paid"} else "AED",
                    reporting_basis="R30 Monthly Feedback", snapshot_date=snapshot, extraction_rule="summary_aggregate_scan"))
        finally:
            wb.close()
        return facts

    def _build_r20(self, path: Path) -> List[FactRecord]:
        wb = self._open(path); facts: List[FactRecord] = []; snapshot = date_from_filename(path)
        try:
            if "Summary 1" in wb.sheetnames:
                ws = wb["Summary 1"]
                for row in ws.iter_rows(min_row=5, max_row=9, values_only=True):
                    bucket = norm(row[1])
                    if not bucket:
                        continue
                    for year, idx in [("2024",2),("2025",3),("2026",4),("Grand Total",5)]:
                        val = to_float(row[idx])
                        if val is not None:
                            facts.append(self._fact(fact_type="fact_pr_soa_tat", source_code="R20", source_file=path.name,
                                sheet="Summary 1", grain="tat_bucket_year", measure_name="tat_percentage", value=val, unit="pct",
                                reporting_basis="R20 PR to SOA TAT", snapshot_date=snapshot, dimensions={"bucket": bucket, "year": year},
                                cell_or_range=f"bucket {bucket} year {year}", extraction_rule="tat_percentage_matrix"))
            if "Summary 2" in wb.sheetnames:
                ws = wb["Summary 2"]
                for row in ws.iter_rows(min_row=4, max_row=8, values_only=True):
                    bucket = norm(row[1])
                    if not bucket:
                        continue
                    total = 0.0
                    for idx in range(2, 14):
                        total += to_float(row[idx], 0.0) or 0.0
                    facts.append(self._fact(fact_type="fact_pr_soa_tat", source_code="R20", source_file=path.name,
                        sheet="Summary 2", grain="tat_bucket_year", measure_name="tat_count_2024", value=total, unit="count",
                        reporting_basis="R20 PR to SOA TAT", snapshot_date=snapshot, dimensions={"bucket": bucket, "year": "2024"}, extraction_rule="tat_count_matrix"))
        finally:
            wb.close()
        return facts

    def _build_r09(self, path: Path) -> List[FactRecord]:
        wb = self._open(path); facts: List[FactRecord] = []; snapshot = date_from_filename(path)
        try:
            for sheet in [s for s in wb.sheetnames if s.startswith("Daily Collections")]:
                ws = wb[sheet]
                sums = defaultdict(float); rows = 0
                for row in ws.iter_rows(min_row=8, values_only=True):
                    project = norm(row[3] if len(row) > 3 else "")
                    if not project or "total" in project.lower():
                        continue
                    rows += 1
                    if sheet.endswith("-0"):
                        cols = {"sale_value_aed":7, "milestone_due_aed":8, "collected_to_date_aed":9, "overdue_aed":11, "pro_rata_target_mtd_aed":13, "collected_dues_mtd_aed":14, "total_collected_mtd_aed":17, "future_collectibles_aed":21}
                    else:
                        cols = {"overdue_aed":4, "milestone_due_month_aed":5, "target_net_constraints_aed":8, "pro_rata_target_aed":9, "collected_dues_mtd_aed":10, "pdc_available_aed":11, "total_collected_dues_new_sales_aed":12}
                    for name, idx in cols.items():
                        sums[name] += to_float(row[idx] if len(row) > idx else None, 0.0) or 0.0
                for name, value in sums.items():
                    facts.append(self._fact(fact_type="fact_project_collections", source_code="R09", source_file=path.name,
                        sheet=sheet, grain="project_snapshot", measure_name=name, value=value,
                        reporting_basis="R09 Collections project detail", snapshot_date=snapshot, extraction_rule="project_sheet_aggregate"))
                facts.append(self._fact(fact_type="fact_project_collections", source_code="R09", source_file=path.name,
                    sheet=sheet, grain="project_snapshot", measure_name="project_rows", value=rows, unit="count",
                    reporting_basis="R09 Collections project detail", snapshot_date=snapshot, extraction_rule="project_row_count"))
        finally:
            wb.close()
        return facts

    def _build_r17(self, path: Path) -> List[FactRecord]:
        wb = self._open(path); facts: List[FactRecord] = []; snapshot = date_from_filename(path)
        try:
            ws = wb["Summary"]
            sums = defaultdict(float); rows = 0
            cols = {"overdue_aed":3, "booking_amount_aed":4, "new_sale_overdue_aed":5, "collection_overdue_aed":6,
                    "spa_not_sent_aed":7, "spa_sent_not_signed_aed":8, "spa_signed_not_executed_aed":11,
                    "spa_executed_not_preregistered_aed":12, "prereg_completed_aed":13}
            for row in ws.iter_rows(min_row=4, values_only=True):
                project = norm(row[2] if len(row) > 2 else "")
                if not project or "total" in project.lower():
                    continue
                rows += 1
                for name, idx in cols.items():
                    sums[name] += to_float(row[idx] if len(row) > idx else None, 0.0) or 0.0
            for name, value in sums.items():
                facts.append(self._fact(fact_type="fact_spa_status_exposure", source_code="R17", source_file=path.name,
                    sheet="Summary", grain="project_snapshot", measure_name=name, value=value,
                    reporting_basis="R17 SPA / preregistration overdue bifurcation", snapshot_date=snapshot, extraction_rule="summary_aggregate"))
            facts.append(self._fact(fact_type="fact_spa_status_exposure", source_code="R17", source_file=path.name,
                sheet="Summary", grain="project_snapshot", measure_name="project_rows", value=rows, unit="count",
                reporting_basis="R17 SPA / preregistration overdue bifurcation", snapshot_date=snapshot, extraction_rule="row_count"))
        finally:
            wb.close()
        return facts

    def _build_r31(self, path: Path) -> List[FactRecord]:
        wb = self._open(path); facts: List[FactRecord] = []; snapshot = date_from_filename(path)
        try:
            if "PR Data" in wb.sheetnames:
                ws = wb["PR Data"]
                status_amt = defaultdict(float); receipt_amt = defaultdict(float); rows = 0; total_amt = 0.0
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if not row or not row[0]:
                        continue
                    rows += 1
                    amt = to_float(row[2], None)
                    if amt is None:
                        amt = to_float(row[15], 0.0) or 0.0
                    total_amt += amt
                    status_amt[norm(row[5]) or "blank"] += amt
                    receipt_amt[norm(row[22]) or "blank"] += amt
                facts.append(self._fact(fact_type="fact_pr_request", source_code="R31", source_file=path.name,
                    sheet="PR Data", grain="pr_snapshot", measure_name="pr_rows", value=rows, unit="count",
                    reporting_basis="R31 PR Unit Update", snapshot_date=snapshot, extraction_rule="pr_data_row_count"))
                facts.append(self._fact(fact_type="fact_pr_request", source_code="R31", source_file=path.name,
                    sheet="PR Data", grain="pr_snapshot", measure_name="pr_total_amount_aed", value=total_amt,
                    reporting_basis="R31 PR Unit Update", snapshot_date=snapshot, extraction_rule="pr_amount_sum"))
                for st, amt in status_amt.items():
                    facts.append(self._fact(fact_type="fact_pr_status", source_code="R31", source_file=path.name,
                        sheet="PR Data", grain="pr_status", measure_name="pr_status_amount_aed", value=amt,
                        reporting_basis="R31 PR Unit Update", snapshot_date=snapshot, dimensions={"status": st}, extraction_rule="status_amount_groupby"))
                for st, amt in receipt_amt.items():
                    facts.append(self._fact(fact_type="fact_receipt_status", source_code="R31", source_file=path.name,
                        sheet="PR Data", grain="receipt_status", measure_name="receipt_status_amount_aed", value=amt,
                        reporting_basis="R31 PR Unit Update", snapshot_date=snapshot, dimensions={"receipt_status": st}, extraction_rule="receipt_status_amount_groupby"))
            if "Daily Collection" in wb.sheetnames:
                ws = wb["Daily Collection"]
                dep = 0.0; rows = 0
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if not any(row):
                        continue
                    rows += 1
                    dep += to_float(row[4] if len(row) > 4 else None, 0.0) or 0.0
                facts.append(self._fact(fact_type="fact_daily_collection_receipt", source_code="R31", source_file=path.name,
                    sheet="Daily Collection", grain="date_entity", measure_name="daily_collection_deposit_aed", value=dep,
                    reporting_basis="R31 Daily Collection receipt bridge", snapshot_date=snapshot, extraction_rule="deposit_sum"))
        finally:
            wb.close()
        return facts

    def _build_r34(self, path: Path) -> List[FactRecord]:
        """Build termination facts from R34 summary/funnel sheets.

        The account-level Data sheet can be large and slow in read-only Excel mode.
        Batch 18 runtime uses trusted workbook summary/funnel sheets by default. A
        future account-risk adapter can add full Data-row scanning behind a worker.
        """
        wb = self._open(path); facts: List[FactRecord] = []; snapshot = date_from_filename(path)
        try:
            if "Summary" in wb.sheetnames:
                ws = wb["Summary"]
                for row in ws.iter_rows(min_row=1, max_row=25, values_only=True):
                    label = norm(row[1] if len(row) > 1 else "")
                    amount = to_float(row[2] if len(row) > 2 else None)
                    units = to_float(row[3] if len(row) > 3 else None)
                    if not label or amount is None:
                        continue
                    key = norm_key(label)
                    facts.append(self._fact(fact_type="fact_termination_funnel", source_code="R34", source_file=path.name,
                        sheet="Summary", grain="termination_summary", measure_name=key + "_amount_aed", value=amount,
                        reporting_basis="R34 Termination Status", snapshot_date=snapshot, dimensions={"summary_label": label}, extraction_rule="summary_table_scan"))
                    facts.append(self._fact(fact_type="fact_termination_funnel", source_code="R34", source_file=path.name,
                        sheet="Summary", grain="termination_summary", measure_name=key + "_units", value=units or 0.0, unit="count",
                        reporting_basis="R34 Termination Status", snapshot_date=snapshot, dimensions={"summary_label": label}, extraction_rule="summary_table_scan"))
                    if "eligible" in key:
                        facts.append(self._fact(fact_type="fact_termination_eligibility", source_code="R34", source_file=path.name,
                            sheet="Summary", grain="termination_summary", measure_name="eligible_amount_aed", value=amount,
                            reporting_basis="R34 Termination Status", snapshot_date=snapshot, dimensions={"eligible": "Eligible"}, extraction_rule="summary_eligible_row"))
            if "Funnel" in wb.sheetnames:
                ws = wb["Funnel"]
                for row in ws.iter_rows(min_row=3, max_row=80, values_only=True):
                    status = norm(row[3] if len(row) > 3 else "")
                    if not status:
                        continue
                    units = to_float(row[6] if len(row) > 6 else None, 0.0) or 0.0
                    amount = to_float(row[7] if len(row) > 7 else None, 0.0) or 0.0
                    if units == 0 and amount == 0:
                        continue
                    facts.append(self._fact(fact_type="fact_termination_system_status", source_code="R34", source_file=path.name,
                        sheet="Funnel", grain="termination_status", measure_name="system_status_amount_aed", value=amount,
                        reporting_basis="R34 Termination Status", snapshot_date=snapshot, dimensions={"system_status": status}, extraction_rule="funnel_status_scan"))
                    facts.append(self._fact(fact_type="fact_termination_system_status", source_code="R34", source_file=path.name,
                        sheet="Funnel", grain="termination_status", measure_name="system_status_units", value=units, unit="count",
                        reporting_basis="R34 Termination Status", snapshot_date=snapshot, dimensions={"system_status": status}, extraction_rule="funnel_status_scan"))
                    if "pdn" in status.lower() or "termination" in status.lower():
                        facts.append(self._fact(fact_type="fact_pdn_exposure", source_code="R34", source_file=path.name,
                            sheet="Funnel", grain="termination_status", measure_name="pdn_or_termination_status_amount_aed", value=amount,
                            reporting_basis="R34 Termination Status", snapshot_date=snapshot, dimensions={"status": status}, extraction_rule="funnel_status_scan"))
        finally:
            wb.close()
        return facts

    def _build_r32(self, path: Path) -> List[FactRecord]:
        wb = self._open(path); facts: List[FactRecord] = []; snapshot = date_from_filename(path)
        try:
            if "RM Collector" in wb.sheetnames:
                ws = wb["RM Collector"]
                sums = defaultdict(float); rows = 0
                for row in ws.iter_rows(min_row=6, values_only=True):
                    exec_name = norm(row[2] if len(row) > 2 else "")
                    if not exec_name or "total" in exec_name.lower():
                        continue
                    rows += 1
                    sums["termination_excluded_collection_aed"] += to_float(row[3], 0.0) or 0.0
                    sums["termination_excluded_dues_mtd_aed"] += to_float(row[4], 0.0) or 0.0
                    sums["termination_excluded_target_aed"] += to_float(row[6], 0.0) or 0.0
                    sums["termination_excluded_receipt_dues_aed"] += to_float(row[7], 0.0) or 0.0
                    sums["termination_excluded_receipt_advance_aed"] += to_float(row[8], 0.0) or 0.0
                    sums["all_collection_aed"] += to_float(row[11], 0.0) or 0.0
                    sums["all_dues_mtd_aed"] += to_float(row[12], 0.0) or 0.0
                    sums["all_target_aed"] += to_float(row[14], 0.0) or 0.0
                    sums["all_receipt_dues_aed"] += to_float(row[15], 0.0) or 0.0
                    sums["all_receipt_advance_aed"] += to_float(row[16], 0.0) or 0.0
                for name, value in sums.items():
                    facts.append(self._fact(fact_type="fact_rm_collector_receipts", source_code="R32", source_file=path.name,
                        sheet="RM Collector", grain="collector_summary", measure_name=name, value=value,
                        reporting_basis="R32 RM Collectors and PR Receipts", snapshot_date=snapshot, extraction_rule="rm_collector_summary"))
                facts.append(self._fact(fact_type="fact_rm_collector_receipts", source_code="R32", source_file=path.name,
                    sheet="RM Collector", grain="collector_summary", measure_name="collector_rows", value=rows, unit="count",
                    reporting_basis="R32 RM Collectors and PR Receipts", snapshot_date=snapshot, extraction_rule="collector_row_count"))
        finally:
            wb.close()
        return facts

    def _build_r38(self, path: Path) -> List[FactRecord]:
        wb = self._open(path); facts: List[FactRecord] = []; snapshot = date_from_filename(path)
        try:
            for sheet in wb.sheetnames:
                if sheet.lower() == "index":
                    continue
                ws = wb[sheet]
                sums = defaultdict(float); rows = 0
                # Most risk sheets have useful values in numeric grids; preserve conservative totals.
                for row in ws.iter_rows(min_row=2, values_only=True):
                    label = norm(row[0] if row else "")
                    if not any(row) or "total" in label.lower():
                        continue
                    nums = [to_float(v) for v in row if to_float(v) is not None]
                    if not nums:
                        continue
                    rows += 1
                    sums["numeric_exposure_total"] += sum(nums)
                    sums["row_count"] += 1
                for name, value in sums.items():
                    facts.append(self._fact(fact_type="fact_risk_analysis", source_code="R38", source_file=path.name,
                        sheet=sheet, grain="risk_bucket_snapshot", measure_name=name, value=value,
                        unit="count" if name == "row_count" else "AED", reporting_basis="R38 Risk Analysis",
                        snapshot_date=snapshot, extraction_rule="risk_numeric_grid_scan"))
        finally:
            wb.close()
        return facts

    def _build_r16(self, path: Path) -> List[FactRecord]:
        wb = self._open(path); facts: List[FactRecord] = []; snapshot = date_from_filename(path)
        try:
            if "Customer" in wb.sheetnames:
                ws = wb["Customer"]
                sums = defaultdict(float); rows = 0
                for row in ws.iter_rows(min_row=5, values_only=True):
                    if not row or not row[0]:
                        continue
                    rows += 1
                    sums["purchase_price_aed"] += to_float(row[4], 0.0) or 0.0
                    sums["ms_due_aed"] += to_float(row[5], 0.0) or 0.0
                    sums["paid_till_date_aed"] += to_float(row[6], 0.0) or 0.0
                    sums["overdues_aed"] += to_float(row[8], 0.0) or 0.0
                    sums["advance_aed"] += to_float(row[9], 0.0) or 0.0
                    sums["future_collectibles_aed"] += to_float(row[10], 0.0) or 0.0
                for name, value in sums.items():
                    facts.append(self._fact(fact_type="fact_customer_portfolio_snapshot", source_code="R16", source_file=path.name,
                        sheet="Customer", grain="customer_snapshot", measure_name=name, value=value,
                        reporting_basis="R16 Customer and Nationality Portfolio", snapshot_date=snapshot, extraction_rule="customer_sheet_aggregate"))
                facts.append(self._fact(fact_type="fact_customer_portfolio_snapshot", source_code="R16", source_file=path.name,
                    sheet="Customer", grain="customer_snapshot", measure_name="customer_rows", value=rows, unit="count",
                    reporting_basis="R16 Customer and Nationality Portfolio", snapshot_date=snapshot, extraction_rule="customer_row_count"))
        finally:
            wb.close()
        return facts

    def _build_r01(self, path: Path) -> List[FactRecord]:
        wb = self._open(path); facts: List[FactRecord] = []; snapshot = date_from_filename(path)
        try:
            for sheet in wb.sheetnames:
                ws = wb[sheet]
                for row in ws.iter_rows(min_row=1, max_row=10, values_only=True):
                    label = norm(row[0] if row else "")
                    if label.lower() in {"dues", "advance", "current year advance", "future years advance", "dues+ advance", "new sales"}:
                        facts.append(self._fact(fact_type="fact_year_end_snapshot", source_code="R01", source_file=path.name,
                            sheet=sheet, grain="year_entity", measure_name=norm_key(label) + "_total_aed", value=to_float(row[1], 0.0),
                            reporting_basis="R01 Year-end snapshot", snapshot_date=snapshot, dimensions={"sheet_year": sheet}, extraction_rule="year_end_top_table"))
        finally:
            wb.close()
        return facts
