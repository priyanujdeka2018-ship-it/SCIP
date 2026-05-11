"""
SCIP Batch 6 - Account/action queue foundation.

Purpose
-------
Build lineaged action queues and collector drilldown contracts on top of the
trusted Batch 1-5 backend contracts without moving calculations into the UI.

Important data-grain disclosure
-------------------------------
The current attached R18 sample exposes project/category/ageing overdue facts,
but does not expose true account IDs, customer IDs, collector/RM owner, promise-
to-pay fields, or escalation status. This module therefore:

1) Emits lineaged project-ageing facts from R18 Overdue-1.
2) Emits project risk cohorts for management / finance / MIS evidence views.
3) Blocks true account-level Collector/RM actions until an account/collector
   source is onboarded with account_id + owner lineage.

This preserves the no-silent-fallback rule and prevents assigning actions to an
untrusted or missing owner.
"""

from __future__ import annotations

from dataclasses import dataclass, asdict
from datetime import datetime, date
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional

try:
    from fastapi import APIRouter, Query
    from fastapi.responses import JSONResponse
except Exception:  # pragma: no cover - allows smoke tests without FastAPI import.
    APIRouter = None
    Query = None
    JSONResponse = None

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

TOLERANCE_PCT = 0.0005  # User-approved 0.05% general rule.

ENTITY_BY_CATEGORY = {
    "A": "sobha_dubai",
    "B": "sobha_dubai",
    "C": "sobha_dubai",
    "D": "sobha_dubai",
    "E": "sobha_dubai",
    "F": "siniya",
    "G": "downtown_uaq",
    "H": "sobha_auh",
}

ENTITY_DISPLAY = {
    "group": "Group",
    "sobha": "Sobha",
    "sobha_dubai": "Sobha Dubai",
    "sobha_auh": "Sobha AUH",
    "uaq": "UAQ",
    "siniya": "Siniya",
    "downtown_uaq": "Downtown UAQ",
    "unknown": "Unknown",
}

PARENT_ENTITY = {
    "sobha_dubai": "sobha",
    "sobha_auh": "sobha",
    "siniya": "uaq",
    "downtown_uaq": "uaq",
    "sobha": "group",
    "uaq": "group",
}

AGEING_COLUMNS = {
    "0-30": 9,
    "31-60": 10,
    "61-90": 11,
    "91-120": 12,
    "121-180": 13,
    "180+": 14,
}

AGEING_PRIORITY = {
    "180+": 6,
    "121-180": 5,
    "91-120": 4,
    "61-90": 3,
    "31-60": 2,
    "0-30": 1,
}

SOURCE_CODE = "R18"
SOURCE_SHEET = "Overdue-1"


@dataclass
class ProjectAgeingFact:
    fact_id: str
    grain: str
    source_code: str
    source_file: str
    sheet: str
    row_number: int
    amount_cell: str
    snapshot_date: Optional[str]
    project: str
    project_category_code: Optional[str]
    entity_code: str
    entity_display: str
    parent_entity_code: Optional[str]
    parent_entity_display: Optional[str]
    ageing_bucket: str
    amount_aed: float
    sale_value_aed: Optional[float]
    milestone_due_aed: Optional[float]
    collected_against_due_aed: Optional[float]
    collected_against_advance_aed: Optional[float]
    row_total_aed: Optional[float]
    account_id: Optional[str]
    customer_id: Optional[str]
    collector_rm_owner: Optional[str]
    promised_amount_aed: Optional[float]
    promised_date: Optional[str]
    escalation_status: Optional[str]
    owner_mapping_status: str
    validation_status: str
    confidence_state: str
    lineage: Dict[str, Any]


class R18ProjectAgeingFactAdapter:
    """Extract project-ageing facts from R18 Overdue-1."""

    def __init__(self, path: Path | str):
        self.path = Path(path)
        self.loaded_at = datetime.utcnow().isoformat() + "Z"

    @staticmethod
    def _to_float(value: Any) -> Optional[float]:
        if value is None or value == "":
            return None
        try:
            return float(value)
        except (TypeError, ValueError):
            return None

    @staticmethod
    def _norm(value: Any) -> str:
        return str(value or "").strip()

    @staticmethod
    def _cell(row: int, col: int) -> str:
        return f"{get_column_letter(col)}{row}"

    @staticmethod
    def _date_to_iso(value: Any) -> Optional[str]:
        if isinstance(value, datetime):
            return value.date().isoformat()
        if isinstance(value, date):
            return value.isoformat()
        if value is None:
            return None
        text = str(value).strip()
        for fmt in ("%Y-%m-%d", "%d-%m-%Y", "%d-%m-%y", "%d/%m/%Y", "%d/%m/%y"):
            try:
                return datetime.strptime(text, fmt).date().isoformat()
            except ValueError:
                continue
        return None

    def extract(self) -> Dict[str, Any]:
        if not self.path.exists():
            return {
                "status": "missing_source",
                "source_code": SOURCE_CODE,
                "facts": [],
                "validations": {},
                "warnings": [],
                "errors": [f"R18 file not found: {self.path}"],
            }

        wb = load_workbook(self.path, data_only=True, read_only=False)
        if SOURCE_SHEET not in wb.sheetnames:
            return {
                "status": "missing_sheet",
                "source_code": SOURCE_CODE,
                "facts": [],
                "validations": {},
                "warnings": [],
                "errors": [f"Missing sheet: {SOURCE_SHEET}"],
            }

        ws = wb[SOURCE_SHEET]
        snapshot_date = self._date_to_iso(ws["B3"].value)
        current_category: Optional[str] = None
        facts: List[ProjectAgeingFact] = []
        row_validations: List[Dict[str, Any]] = []
        grand_total = 0.0

        for row in range(7, ws.max_row + 1):
            category_marker = self._norm(ws.cell(row, 3).value).upper()
            project = self._norm(ws.cell(row, 4).value)

            if category_marker in ENTITY_BY_CATEGORY and not project:
                current_category = category_marker
                continue

            if not project:
                continue

            is_subtotal = project.lower().startswith("sub total") or project.lower() == "grand total"
            if project.lower() == "grand total":
                grand_total = self._to_float(ws.cell(row, 15).value) or 0.0

            if is_subtotal:
                continue

            row_total = self._to_float(ws.cell(row, 15).value)
            bucket_sum = 0.0
            for bucket, col in AGEING_COLUMNS.items():
                amount = self._to_float(ws.cell(row, col).value) or 0.0
                bucket_sum += amount
                if amount <= 0:
                    continue

                entity_code = ENTITY_BY_CATEGORY.get(current_category or "", "unknown")
                parent = PARENT_ENTITY.get(entity_code)
                cell_ref = self._cell(row, col)
                fact_id = f"R18-{SOURCE_SHEET}-R{row}-{bucket.replace('+','plus').replace('-','_')}"
                lineage = {
                    "source_code": SOURCE_CODE,
                    "source_file": self.path.name,
                    "sheet": SOURCE_SHEET,
                    "cell_or_range": cell_ref,
                    "row_number": row,
                    "snapshot_date": snapshot_date,
                    "extraction_method": "R18ProjectAgeingFactAdapter.project_row_ageing_bucket",
                    "business_definition": "Project-level overdue amount by ageing bucket from R18 Overdue-1. Not account-level until an account/collector source is onboarded.",
                    "validation_status": "passed",
                    "confidence_state": "live_validated_project_grain_owner_missing",
                    "last_loaded_at": self.loaded_at,
                }
                facts.append(ProjectAgeingFact(
                    fact_id=fact_id,
                    grain="project_ageing_bucket",
                    source_code=SOURCE_CODE,
                    source_file=self.path.name,
                    sheet=SOURCE_SHEET,
                    row_number=row,
                    amount_cell=cell_ref,
                    snapshot_date=snapshot_date,
                    project=project,
                    project_category_code=current_category,
                    entity_code=entity_code,
                    entity_display=ENTITY_DISPLAY.get(entity_code, "Unknown"),
                    parent_entity_code=parent,
                    parent_entity_display=ENTITY_DISPLAY.get(parent) if parent else None,
                    ageing_bucket=bucket,
                    amount_aed=round(amount, 2),
                    sale_value_aed=self._to_float(ws.cell(row, 5).value),
                    milestone_due_aed=self._to_float(ws.cell(row, 6).value),
                    collected_against_due_aed=self._to_float(ws.cell(row, 7).value),
                    collected_against_advance_aed=self._to_float(ws.cell(row, 8).value),
                    row_total_aed=row_total,
                    account_id=None,
                    customer_id=None,
                    collector_rm_owner=None,
                    promised_amount_aed=None,
                    promised_date=None,
                    escalation_status=None,
                    owner_mapping_status="missing_account_collector_source",
                    validation_status="passed",
                    confidence_state="live_validated_project_grain_owner_missing",
                    lineage=lineage,
                ))

            if row_total and row_total > 0:
                tolerance = abs(row_total) * TOLERANCE_PCT
                row_validations.append({
                    "row": row,
                    "project": project,
                    "row_total_aed": row_total,
                    "bucket_sum_aed": bucket_sum,
                    "diff_aed": abs(row_total - bucket_sum),
                    "tolerance_aed": tolerance,
                    "passed": abs(row_total - bucket_sum) <= tolerance,
                })

        total_from_facts = round(sum(f.amount_aed for f in facts), 2)
        grand_total_check = None
        if grand_total:
            tolerance = abs(grand_total) * TOLERANCE_PCT
            grand_total_check = {
                "passed": abs(total_from_facts - grand_total) <= tolerance,
                "fact_sum_aed": total_from_facts,
                "grand_total_aed": round(grand_total, 2),
                "diff_aed": round(abs(total_from_facts - grand_total), 2),
                "tolerance_aed": tolerance,
                "tolerance_pct": TOLERANCE_PCT,
            }

        failed_rows = [v for v in row_validations if not v["passed"]]
        validations = {
            "row_bucket_sum_checks": {
                "passed": not failed_rows,
                "checked_rows": len(row_validations),
                "failed_rows": failed_rows[:20],
                "tolerance_pct": TOLERANCE_PCT,
            },
            "grand_total_reconciliation": grand_total_check,
            "lineage_required_fields": {
                "passed": all(_has_fact_lineage(asdict(f)) for f in facts),
                "required": ["source_code", "source_file", "sheet", "cell_or_range", "snapshot_date", "validation_status", "confidence_state"],
            },
        }
        status = "ok_project_grain" if facts else "empty"
        return {
            "status": status,
            "source_code": SOURCE_CODE,
            "source_file": self.path.name,
            "sheet": SOURCE_SHEET,
            "snapshot_date": snapshot_date,
            "facts": [asdict(f) for f in facts],
            "validations": validations,
            "warnings": [
                "R18 Overdue-1 provides project/category/ageing facts, not account IDs or collector/RM owner fields.",
                "Collector/RM action queues are blocked until account/collector source reports are onboarded.",
            ],
            "errors": [],
        }


def _has_fact_lineage(fact: Dict[str, Any]) -> bool:
    lin = fact.get("lineage") or {}
    required = ["source_code", "source_file", "sheet", "cell_or_range", "snapshot_date", "validation_status", "confidence_state"]
    return all(bool(lin.get(k)) for k in required)


def _find_r18_file(data_dir: Path) -> Optional[Path]:
    candidates = sorted(data_dir.glob("R18*.xlsx")) + sorted(data_dir.glob("*R18*.xlsx"))
    return candidates[0] if candidates else None


def _format_aed(value: Optional[float]) -> str:
    if value is None:
        return "Unavailable"
    abs_value = abs(value)
    if abs_value >= 1_000_000_000:
        return f"AED {value / 1_000_000_000:.3f}B"
    if abs_value >= 1_000_000:
        return f"AED {value / 1_000_000:.1f}M"
    return f"AED {value:,.0f}"


def _action_severity(bucket: str, amount: float) -> str:
    if bucket == "180+" or amount >= 50_000_000:
        return "risk"
    if bucket in {"121-180", "91-120"} or amount >= 10_000_000:
        return "attention"
    return "watch"


def _lineage_ref_from_fact(fact: Dict[str, Any]) -> Dict[str, Any]:
    lin = fact.get("lineage") or {}
    return {
        "metric_key": fact.get("fact_id"),
        "metric_label": f"{fact.get('project')} {fact.get('ageing_bucket')} overdue",
        "source_code": lin.get("source_code"),
        "source_file": lin.get("source_file"),
        "sheet": lin.get("sheet"),
        "cell_or_range": lin.get("cell_or_range"),
        "snapshot_date": lin.get("snapshot_date"),
        "validation_status": lin.get("validation_status"),
        "confidence_state": lin.get("confidence_state"),
        "reporting_basis": "R18 project ageing evidence",
    }


def _make_project_action(fact: Dict[str, Any], role: str) -> Dict[str, Any]:
    amount = float(fact.get("amount_aed") or 0.0)
    severity = _action_severity(str(fact.get("ageing_bucket")), amount)
    action_map = {
        "cco_gm_agm": "Review project-level overdue concentration and request account-owner drilldown.",
        "finance": "Check finance classification and reconcile overdue amount before escalation.",
        "mis_qcg_admin": "Onboard missing account/collector source and validate owner mapping.",
        "collector_rm": "Blocked: collector/RM owner is not present in current source lineage.",
    }
    return {
        "action_id": f"{role}-{fact.get('fact_id')}",
        "action_type": "project_risk_cohort",
        "grain": fact.get("grain"),
        "status": "available_project_grain" if role != "collector_rm" else "blocked_missing_owner_lineage",
        "title": f"{fact.get('project')} · {fact.get('ageing_bucket')} ageing",
        "summary": f"{_format_aed(amount)} overdue in {fact.get('entity_display')} from R18 project-ageing evidence.",
        "recommended_action": action_map.get(role, "Review source-lineaged risk cohort."),
        "severity": severity,
        "amount_aed": amount,
        "display_amount": _format_aed(amount),
        "ageing_bucket": fact.get("ageing_bucket"),
        "project": fact.get("project"),
        "entity_code": fact.get("entity_code"),
        "entity_display": fact.get("entity_display"),
        "parent_entity_code": fact.get("parent_entity_code"),
        "parent_entity_display": fact.get("parent_entity_display"),
        "account_id": fact.get("account_id"),
        "customer_id": fact.get("customer_id"),
        "collector_rm_owner": fact.get("collector_rm_owner"),
        "owner_mapping_status": fact.get("owner_mapping_status"),
        "blocked_reason": None if role != "collector_rm" else "Current attached source lacks collector/RM owner and account ID fields.",
        "lineage_refs": [_lineage_ref_from_fact(fact)],
        "reporting_basis": "R18 project ageing evidence",
        "confidence_state": fact.get("confidence_state"),
    }


def _top_facts(facts: Iterable[Dict[str, Any]], limit: int = 12) -> List[Dict[str, Any]]:
    return sorted(
        [f for f in facts if float(f.get("amount_aed") or 0) > 0],
        key=lambda f: (AGEING_PRIORITY.get(str(f.get("ageing_bucket")), 0), float(f.get("amount_aed") or 0)),
        reverse=True,
    )[:limit]


def build_action_queue_payload(data_dir: Optional[Path | str] = None, role: Optional[str] = None) -> Dict[str, Any]:
    base = Path(data_dir or Path(__file__).resolve().parent)
    # In deployed backend, DATA_DIR may be ../data or supplied through env; in this patch pack, /mnt/data works for smoke.
    if not any(base.glob("R18*.xlsx")) and Path("/mnt/data").exists():
        base = Path("/mnt/data")
    r18_file = _find_r18_file(base)

    if not r18_file:
        return {
            "contract_version": "action_queues.v1.batch6",
            "status": "blocked_missing_r18",
            "generated_at": datetime.utcnow().isoformat() + "Z",
            "guardrails": _guardrails(),
            "source_availability": {"R18": "missing"},
            "roles": {},
            "validations": {"passed": False, "reason": "R18 source missing"},
        }

    extracted = R18ProjectAgeingFactAdapter(r18_file).extract()
    facts = extracted.get("facts", [])
    top = _top_facts(facts, limit=15)

    eligible_account_actions = [
        f for f in facts
        if f.get("account_id") and f.get("collector_rm_owner") and f.get("entity_code") and _has_fact_lineage(f)
    ]
    account_source_ready = bool(eligible_account_actions)

    roles = {}
    for role_key in ["collector_rm", "cco_gm_agm", "finance", "mis_qcg_admin"]:
        project_actions = [_make_project_action(f, role_key) for f in top]
        account_actions = []
        if account_source_ready:
            # Future extension point: true account actions once account/collector reports are onboarded.
            account_actions = project_actions[:0]
        roles[role_key] = {
            "role": role_key,
            "status": "ready" if account_source_ready else "account_source_required",
            "headline": _role_headline(role_key, account_source_ready),
            "disclosure": "No account-level action is shown without account ID, owner mapping, entity mapping, amount/ageing validation, and source lineage.",
            "account_actions": account_actions,
            "project_risk_cohorts": project_actions if role_key != "collector_rm" else [],
            "blocked_actions": project_actions if role_key == "collector_rm" else [],
            "required_source_fields": ["account_id", "customer_id", "collector_rm_owner", "project", "entity", "ageing_bucket", "overdue_amount", "source_lineage"],
            "reporting_basis": "R18 project ageing evidence; true account queues blocked until account/collector source onboarded",
        }

    validations = _validate_payload(extracted, roles)
    return {
        "contract_version": "action_queues.v1.batch6",
        "status": "partial_project_grain_account_source_required" if not account_source_ready else "ready",
        "generated_at": datetime.utcnow().isoformat() + "Z",
        "snapshot_date": extracted.get("snapshot_date"),
        "guardrails": _guardrails(),
        "source_availability": {
            "R18": "loaded_project_ageing_grain",
            "account_collector_report": "missing_not_attached",
        },
        "data_grain_disclosure": {
            "current_grain": "project_ageing_bucket",
            "true_account_level_available": account_source_ready,
            "reason": "Current attached R18 sample lacks account ID, customer ID, collector/RM owner, promise-to-pay, and escalation fields.",
        },
        "facts_summary": {
            "project_ageing_facts": len(facts),
            "top_project_risk_cohorts": len(top),
            "eligible_account_actions": len(eligible_account_actions),
            "project_ageing_observation_amount_aed_non_additive": round(sum(float(f.get("amount_aed") or 0) for f in facts), 2),
            "amount_sum_disclosure": "Non-additive: R18 Overdue-1 contains parent project rows and child phase/tower rows. Use line-level actions, not this observation sum, for financial reconciliation.",
        },
        "roles": roles,
        "lineage_sample": [_lineage_ref_from_fact(f) for f in top[:5]],
        "source_validations": extracted.get("validations"),
        "validations": validations,
        "warnings": extracted.get("warnings", []),
        "errors": extracted.get("errors", []),
    }


def _role_headline(role_key: str, account_source_ready: bool) -> str:
    if account_source_ready:
        return "Lineaged account actions are ready."
    headlines = {
        "collector_rm": "Collector actions are blocked until owner-level account data is onboarded.",
        "cco_gm_agm": "Project risk cohorts are ready; account owner drilldown is awaiting source onboarding.",
        "finance": "Finance can review project-level overdue cohorts while account mapping is pending.",
        "mis_qcg_admin": "MIS/QCG must onboard the account/collector report to unlock owner-level queues.",
    }
    return headlines.get(role_key, "Action queue requires account-source onboarding.")


def _guardrails() -> Dict[str, Any]:
    return {
        "tolerance_pct": 0.05,
        "no_silent_fallback": True,
        "locked_hierarchy": "Group > Sobha(Sobha Dubai, Sobha AUH) + UAQ(Siniya, Downtown UAQ)",
        "account_action_gate": "account actions require account_id + collector_rm_owner + entity + amount + ageing + lineage",
        "liquid_glass_gate": "collector tables and action lists appear only as L4 evidence/action output",
    }


def _validate_payload(extracted: Dict[str, Any], roles: Dict[str, Any]) -> Dict[str, Any]:
    facts = extracted.get("facts", [])
    all_facts_lineaged = all(_has_fact_lineage(f) for f in facts)

    account_actions = []
    project_actions = []
    for role_payload in roles.values():
        account_actions.extend(role_payload.get("account_actions") or [])
        project_actions.extend(role_payload.get("project_risk_cohorts") or [])
        project_actions.extend(role_payload.get("blocked_actions") or [])

    no_unlineaged_account_action = all(
        action.get("account_id")
        and action.get("collector_rm_owner")
        and action.get("entity_code")
        and action.get("ageing_bucket")
        and action.get("amount_aed") is not None
        and action.get("lineage_refs")
        for action in account_actions
    )
    project_actions_lineaged = all(action.get("lineage_refs") for action in project_actions)
    no_collector_visible_actions_without_owner = len(roles.get("collector_rm", {}).get("account_actions") or []) == 0

    checks = {
        "project_facts_have_lineage": all_facts_lineaged,
        "project_actions_have_lineage": project_actions_lineaged,
        "no_account_action_without_owner_and_lineage": no_unlineaged_account_action,
        "collector_rm_queue_blocked_without_owner_source": no_collector_visible_actions_without_owner,
        "source_row_reconciliation_passed": bool((extracted.get("validations") or {}).get("row_bucket_sum_checks", {}).get("passed")),
    }
    return {
        "passed": all(checks.values()),
        "checks": checks,
    }


if APIRouter is not None:
    router = APIRouter(prefix="/action-queues", tags=["action-queues"])

    @router.get("")
    async def get_action_queues(role: Optional[str] = None) -> Any:
        payload = build_action_queue_payload(role=role)
        return JSONResponse(content=payload, status_code=200)

    @router.get("/collector-drilldown")
    async def get_collector_drilldown(role: str = "collector_rm") -> Any:
        payload = build_action_queue_payload(role=role)
        collector = payload.get("roles", {}).get("collector_rm", {})
        return JSONResponse(content={
            "contract_version": "collector_drilldown.v1.batch6",
            "status": collector.get("status", "account_source_required"),
            "headline": collector.get("headline"),
            "disclosure": collector.get("disclosure"),
            "account_actions": collector.get("account_actions", []),
            "blocked_actions": collector.get("blocked_actions", []),
            "required_source_fields": collector.get("required_source_fields", []),
            "validations": payload.get("validations"),
        }, status_code=200)
