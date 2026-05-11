"""
SCIP source_adapters.py — Batch 1 ingestion foundation

Purpose:
- Provide source-specific adapters for complex R-series workbooks.
- Implement the R18 OD adapter using the locked Group/Sobha/UAQ hierarchy.
- Generate metric-level lineage and validation results for every OD metric.

Runtime dependencies: stdlib + openpyxl only.
"""

from __future__ import annotations

from dataclasses import dataclass, field, asdict
from datetime import datetime, date
from pathlib import Path
from typing import Any, Dict, List, Optional

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

# User-approved general reconciliation tolerance: 0.05%.
TOLERANCE_PCT = 0.0005


ENTITY_HIERARCHY = {
    "group": {"display": "Group", "parent": None, "children": ["sobha", "uaq"]},
    "sobha": {"display": "Sobha", "parent": "group", "children": ["sobha_dubai", "sobha_auh"]},
    "sobha_dubai": {"display": "Sobha Dubai", "parent": "sobha", "children": []},
    "sobha_auh": {"display": "Sobha AUH", "parent": "sobha", "children": []},
    "uaq": {"display": "UAQ", "parent": "group", "children": ["siniya", "downtown_uaq"]},
    "siniya": {"display": "Siniya", "parent": "uaq", "children": []},
    "downtown_uaq": {"display": "Downtown UAQ", "parent": "uaq", "children": []},
}


@dataclass
class LineageRecord:
    metric_key: str
    metric_label: str
    value: Optional[float]
    unit: str
    source_code: str
    source_file: str
    sheet: str
    cell_or_range: str
    snapshot_date: Optional[str]
    extraction_method: str
    entity_scope: str
    business_definition: str
    validation_status: str
    confidence_state: str
    last_loaded_at: str


@dataclass
class AdapterResult:
    source_code: str
    status: str
    metrics: Dict[str, Any] = field(default_factory=dict)
    rows: List[Dict[str, Any]] = field(default_factory=list)
    lineage: Dict[str, Dict[str, Any]] = field(default_factory=dict)
    warnings: List[str] = field(default_factory=list)
    errors: List[str] = field(default_factory=list)
    validations: Dict[str, Any] = field(default_factory=dict)

    def to_dict(self) -> Dict[str, Any]:
        return asdict(self)


class BaseAdapter:
    source_code = "BASE"

    def __init__(self, path: Path | str):
        self.path = Path(path)
        self.loaded_at = datetime.utcnow().isoformat() + "Z"

    def extract(self) -> AdapterResult:
        raise NotImplementedError

    @staticmethod
    def _to_float(value: Any) -> Optional[float]:
        if value is None:
            return None
        try:
            return float(value)
        except (TypeError, ValueError):
            return None

    @staticmethod
    def _norm(value: Any) -> str:
        return str(value or "").strip().lower()

    @staticmethod
    def _date_to_iso(value: Any) -> Optional[str]:
        if isinstance(value, datetime):
            return value.date().isoformat()
        if isinstance(value, date):
            return value.isoformat()
        if value is None:
            return None
        text = str(value).strip()
        for fmt in ("%Y-%m-%d", "%d-%m-%Y", "%d-%m-%y", "%d/%m/%Y", "%d/%m/%y"):
            try:
                return datetime.strptime(text, fmt).date().isoformat()
            except ValueError:
                pass
        return None

    @staticmethod
    def _currency_cell(row_idx: int, col_idx_0_based: int) -> str:
        return f"{get_column_letter(col_idx_0_based + 1)}{row_idx}"

    @staticmethod
    def _pct_tolerance(reference_value: Optional[float], pct: float = TOLERANCE_PCT) -> float:
        if reference_value is None:
            return 0.0
        return abs(float(reference_value)) * pct

    @classmethod
    def _within_pct_tolerance(
        cls,
        left: Optional[float],
        right: Optional[float],
        reference_value: Optional[float] = None,
        pct: float = TOLERANCE_PCT,
    ) -> Dict[str, Any]:
        if left is None or right is None:
            return {
                "passed": False,
                "left": left,
                "right": right,
                "diff": None,
                "tolerance_pct": pct,
                "tolerance_aed": None,
            }
        ref = reference_value if reference_value is not None else max(abs(float(left)), abs(float(right)))
        tolerance = cls._pct_tolerance(ref, pct)
        diff = abs(float(left) - float(right))
        return {
            "passed": diff <= tolerance,
            "left": left,
            "right": right,
            "diff": diff,
            "tolerance_pct": pct,
            "tolerance_aed": tolerance,
        }


class R18OverdueAdapter(BaseAdapter):
    """Extract OD and ageing from R18 Overdue-1 using positional row scans."""

    source_code = "R18"
    sheet_name = "Overdue-1"

    # openpyxl row tuple indices are 0-based. Excel total column is O.
    total_col = 14
    ageing_cols = {
        "ageing_0_30": 8,
        "ageing_31_60": 9,
        "ageing_61_90": 10,
        "ageing_91_120": 11,
        "ageing_121_180": 12,
        "ageing_180plus": 13,
    }

    # Locked hierarchy mapping approved by user.
    subtotal_to_entity = {
        "sub total (a)": "sobha_dubai",
        "sub total (b)": "sobha_dubai",
        "sub total (c)": "sobha_dubai",
        "sub total (d)": "sobha_dubai",
        "sub total (e)": "sobha_dubai",
        "sub total (f)": "siniya",
        "sub total (g)": "downtown_uaq",
        "sub total (h)": "sobha_auh",
    }

    metric_labels = {
        "OD_SOBHA_DUBAI": "Overdue - Sobha Dubai",
        "OD_SOBHA_AUH": "Overdue - Sobha AUH",
        "OD_SOBHA": "Overdue - Sobha",
        "OD_SINIYA": "Overdue - Siniya",
        "OD_DOWNTOWN_UAQ": "Overdue - Downtown UAQ",
        "OD_DT": "Overdue - Downtown UAQ alias",
        "OD_UAQ": "Overdue - UAQ",
        "OD_GROUP": "Overdue - Group roll-up",
        "OD_TODAY": "Overdue - R18 Grand Total",
    }

    def extract(self) -> AdapterResult:
        result = AdapterResult(source_code=self.source_code, status="unknown")
        if not self.path.exists():
            result.status = "unavailable"
            result.errors.append(f"File not found: {self.path}")
            return result

        try:
            wb = load_workbook(self.path, data_only=True, read_only=True)
        except Exception as exc:
            result.status = "unavailable"
            result.errors.append(f"Could not open workbook: {exc}")
            return result

        try:
            if self.sheet_name not in wb.sheetnames:
                result.status = "unavailable"
                result.errors.append(f"Missing sheet: {self.sheet_name}")
                return result

            ws = wb[self.sheet_name]
            snapshot_date = self._find_snapshot_date(ws)
            subtotals: Dict[str, float] = {
                "sobha_dubai": 0.0,
                "sobha_auh": 0.0,
                "siniya": 0.0,
                "downtown_uaq": 0.0,
            }
            subtotal_rows_seen: Dict[str, int] = {}
            subtotal_cells_by_entity: Dict[str, List[str]] = {k: [] for k in subtotals}
            subtotal_values_by_label: Dict[str, float] = {}
            grand_total: Optional[float] = None
            grand_total_row: Optional[int] = None
            ageing: Dict[str, Optional[float]] = {}

            for row_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
                row_text = [self._norm(v) for v in row]

                # R18 labels are in column D in the current sample, but scan whole row
                # because historic R18 exports have shifted merged label columns.
                for cell_text in row_text:
                    if cell_text in self.subtotal_to_entity:
                        entity = self.subtotal_to_entity[cell_text]
                        value = self._to_float(row[self.total_col] if len(row) > self.total_col else None) or 0.0
                        subtotals[entity] += value
                        subtotal_rows_seen[cell_text] = row_idx
                        subtotal_values_by_label[cell_text] = value
                        subtotal_cells_by_entity[entity].append(self._currency_cell(row_idx, self.total_col))
                        break

                if "grand total" in row_text:
                    grand_total = self._to_float(row[self.total_col] if len(row) > self.total_col else None)
                    grand_total_row = row_idx
                    for key, idx in self.ageing_cols.items():
                        ageing[key] = self._to_float(row[idx] if len(row) > idx else None)

            od_sobha = subtotals["sobha_dubai"] + subtotals["sobha_auh"]
            od_uaq = subtotals["siniya"] + subtotals["downtown_uaq"]
            od_group = od_sobha + od_uaq

            metrics = {
                "OD_SOBHA_DUBAI": subtotals["sobha_dubai"],
                "OD_SOBHA_AUH": subtotals["sobha_auh"],
                "OD_SOBHA": od_sobha,
                "OD_SINIYA": subtotals["siniya"],
                "OD_DOWNTOWN_UAQ": subtotals["downtown_uaq"],
                "OD_DT": subtotals["downtown_uaq"],  # backward-compatible alias only
                "OD_UAQ": od_uaq,
                "OD_GROUP": od_group,
                "OD_TODAY": grand_total,
                "SNAPSHOT_DATE": snapshot_date,
                **ageing,
            }
            result.metrics = metrics

            result.rows = [
                {
                    "entity_key": "sobha_dubai",
                    "entity_label": "Sobha Dubai",
                    "parent_key": "sobha",
                    "metric_key": "OD_SOBHA_DUBAI",
                    "value_aed": subtotals["sobha_dubai"],
                    "source_cells": subtotal_cells_by_entity["sobha_dubai"],
                },
                {
                    "entity_key": "sobha_auh",
                    "entity_label": "Sobha AUH",
                    "parent_key": "sobha",
                    "metric_key": "OD_SOBHA_AUH",
                    "value_aed": subtotals["sobha_auh"],
                    "source_cells": subtotal_cells_by_entity["sobha_auh"],
                },
                {
                    "entity_key": "siniya",
                    "entity_label": "Siniya",
                    "parent_key": "uaq",
                    "metric_key": "OD_SINIYA",
                    "value_aed": subtotals["siniya"],
                    "source_cells": subtotal_cells_by_entity["siniya"],
                },
                {
                    "entity_key": "downtown_uaq",
                    "entity_label": "Downtown UAQ",
                    "parent_key": "uaq",
                    "metric_key": "OD_DOWNTOWN_UAQ",
                    "value_aed": subtotals["downtown_uaq"],
                    "source_cells": subtotal_cells_by_entity["downtown_uaq"],
                },
            ]

            # Validations use the approved 0.05% tolerance.
            result.validations["R18_REQUIRED_SUBTOTALS"] = self._validate_required_subtotals(subtotal_rows_seen)
            result.validations["R18_OD_SOBHA_PLUS_UAQ_EQUALS_OD_TODAY"] = self._within_pct_tolerance(od_sobha + od_uaq, grand_total, grand_total)
            result.validations["R18_OD_GROUP_EQUALS_GRAND_TOTAL"] = self._within_pct_tolerance(od_group, grand_total, grand_total)
            result.validations["R18_SOBHA_EQUALS_CHILDREN"] = self._within_pct_tolerance(od_sobha, subtotals["sobha_dubai"] + subtotals["sobha_auh"], od_sobha)
            result.validations["R18_UAQ_EQUALS_CHILDREN"] = self._within_pct_tolerance(od_uaq, subtotals["siniya"] + subtotals["downtown_uaq"], od_uaq)
            if grand_total is not None and ageing:
                ageing_sum = sum(v or 0 for v in ageing.values())
                result.validations["R18_AGEING_BUCKETS_EQUAL_TOTAL"] = self._within_pct_tolerance(ageing_sum, grand_total, grand_total)

            failed = [k for k, v in result.validations.items() if not v.get("passed", False)]
            if failed:
                confidence = "live_warning"
                validation_status = "warning"
                result.status = "warning"
                result.warnings.append("Failed validations: " + ", ".join(failed))
            else:
                confidence = "live_validated"
                validation_status = "passed"
                result.status = "ok"

            lineage_specs = {
                "OD_SOBHA_DUBAI": (
                    "sobha_dubai",
                    ",".join(subtotal_cells_by_entity["sobha_dubai"]),
                    "Sobha Dubai OD equals R18 Sub Total A-E total OD.",
                ),
                "OD_SOBHA_AUH": (
                    "sobha_auh",
                    ",".join(subtotal_cells_by_entity["sobha_auh"]),
                    "Sobha AUH OD equals R18 Sub Total H total OD.",
                ),
                "OD_SOBHA": (
                    "sobha",
                    "rollup: OD_SOBHA_DUBAI + OD_SOBHA_AUH",
                    "Sobha parent OD equals Sobha Dubai plus Sobha AUH.",
                ),
                "OD_SINIYA": (
                    "siniya",
                    ",".join(subtotal_cells_by_entity["siniya"]),
                    "Siniya OD equals R18 Sub Total F total OD.",
                ),
                "OD_DOWNTOWN_UAQ": (
                    "downtown_uaq",
                    ",".join(subtotal_cells_by_entity["downtown_uaq"]),
                    "Downtown UAQ OD equals R18 Sub Total G total OD.",
                ),
                "OD_DT": (
                    "downtown_uaq",
                    "alias: OD_DOWNTOWN_UAQ",
                    "Backward-compatible alias for Downtown UAQ OD; do not use as a hierarchy label.",
                ),
                "OD_UAQ": (
                    "uaq",
                    "rollup: OD_SINIYA + OD_DOWNTOWN_UAQ",
                    "UAQ parent OD equals Siniya plus Downtown UAQ.",
                ),
                "OD_GROUP": (
                    "group",
                    "rollup: OD_SOBHA + OD_UAQ",
                    "Group OD equals Sobha plus UAQ.",
                ),
                "OD_TODAY": (
                    "group",
                    self._currency_cell(grand_total_row, self.total_col) if grand_total_row else "Grand Total row not found",
                    "R18 Grand Total OD, used as OD_TODAY.",
                ),
            }
            for metric_key, (entity_scope, cell_or_range, definition) in lineage_specs.items():
                result.lineage[metric_key] = asdict(LineageRecord(
                    metric_key=metric_key,
                    metric_label=self.metric_labels[metric_key],
                    value=metrics.get(metric_key),
                    unit="AED",
                    source_code=self.source_code,
                    source_file=self.path.name,
                    sheet=self.sheet_name,
                    cell_or_range=cell_or_range,
                    snapshot_date=snapshot_date,
                    extraction_method="positional_row_scan",
                    entity_scope=entity_scope,
                    business_definition=definition,
                    validation_status=validation_status,
                    confidence_state=confidence,
                    last_loaded_at=self.loaded_at,
                ))

            # Add ageing lineage too, because ageing is part of the OD trust contract.
            for key, idx in self.ageing_cols.items():
                result.lineage[key] = asdict(LineageRecord(
                    metric_key=key,
                    metric_label=f"OD ageing {key.replace('ageing_', '').replace('_', '-')}",
                    value=metrics.get(key),
                    unit="AED",
                    source_code=self.source_code,
                    source_file=self.path.name,
                    sheet=self.sheet_name,
                    cell_or_range=self._currency_cell(grand_total_row, idx) if grand_total_row else "Grand Total row not found",
                    snapshot_date=snapshot_date,
                    extraction_method="positional_row_scan",
                    entity_scope="group",
                    business_definition="Group OD ageing bucket from R18 Grand Total row.",
                    validation_status=validation_status,
                    confidence_state=confidence,
                    last_loaded_at=self.loaded_at,
                ))

            # Helpful extraction audit for smoke files.
            result.metrics["_subtotal_values_by_label"] = subtotal_values_by_label
            result.metrics["_subtotal_rows_seen"] = subtotal_rows_seen

            return result
        finally:
            try:
                wb.close()
            except Exception:
                pass

    def _find_snapshot_date(self, ws) -> Optional[str]:
        # R18 sample has snapshot date in B3. Scan early visible cells to tolerate exports.
        for row in ws.iter_rows(min_row=1, max_row=6, values_only=True):
            for value in row[:10]:
                iso = self._date_to_iso(value)
                if iso:
                    return iso
        return None

    @staticmethod
    def _validate_required_subtotals(subtotal_rows_seen: Dict[str, int]) -> Dict[str, Any]:
        required = {f"sub total ({letter})" for letter in "abcdefgh"}
        seen = set(subtotal_rows_seen)
        missing = sorted(required - seen)
        return {"passed": not missing, "missing": missing, "seen": sorted(seen)}


class SheetPresenceAdapter(BaseAdapter):
    """Validation-only scaffold used by non-R18 adapters in Batch 1."""

    required_sheets: List[str] = []

    def extract(self) -> AdapterResult:
        result = AdapterResult(source_code=self.source_code, status="unknown")
        if not self.path.exists():
            result.status = "unavailable"
            result.errors.append(f"File not found: {self.path}")
            return result
        try:
            wb = load_workbook(self.path, data_only=True, read_only=True)
            missing = [s for s in self.required_sheets if s not in wb.sheetnames]
            result.metrics["sheetnames"] = wb.sheetnames
            result.validations[f"{self.source_code}_REQUIRED_SHEETS_PRESENT"] = {"passed": not missing, "missing": missing}
            result.status = "ok" if not missing else "warning"
            return result
        except Exception as exc:
            result.status = "unavailable"
            result.errors.append(str(exc))
            return result
        finally:
            try:
                wb.close()
            except Exception:
                pass


class R02MDOAdapter(SheetPresenceAdapter):
    source_code = "R02"
    required_sheets = ["MDO Dynamic", "Monthwise MDO", "daily", "daily (2)", "daily (3)", "daily (4)", "daily (5)"]


class R04FinanceDailyAdapter(SheetPresenceAdapter):
    source_code = "R04"
    required_sheets = ["daily", "month_target"]


class R08AdvanceAdapter(SheetPresenceAdapter):
    source_code = "R08"
    required_sheets = ["Summary", "Rebate Summary", "Advance 2026CYFY", "Siniya CY FY", "DT CY FY"]


class R36MilestoneCohortAdapter(SheetPresenceAdapter):
    source_code = "R36"
    required_sheets = ["Total", "Active", "Sobha", "Siniya", "DT"]


ADAPTERS = {
    "R18": R18OverdueAdapter,
    "R02": R02MDOAdapter,
    "R04": R04FinanceDailyAdapter,
    "R08": R08AdvanceAdapter,
    "R36": R36MilestoneCohortAdapter,
}


def run_adapter(r_code: str, path: str | Path) -> AdapterResult:
    adapter_cls = ADAPTERS.get(r_code.upper())
    if not adapter_cls:
        raise ValueError(f"No adapter registered for {r_code}")
    return adapter_cls(path).extract()
