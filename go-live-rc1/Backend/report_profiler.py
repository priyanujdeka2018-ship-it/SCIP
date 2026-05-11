"""
SCIP Batch 18 report profiler.

Profiles R-series workbooks before semantic extraction. The profiler is intentionally
read-only and tolerant; extraction failures should produce warnings, not silent facts.
"""
from __future__ import annotations

from pathlib import Path
from typing import Any, Dict, List, Optional

from openpyxl import load_workbook

from analytics_models import ReportProfile
from analytics_utils import date_from_filename, norm, norm_key, parse_date, utc_now


class ReportProfiler:
    def profile_report(self, report_code: str, path: str | Path) -> ReportProfile:
        source_path = Path(path)
        warnings: List[str] = []
        sheets: List[Dict[str, Any]] = []
        snapshot_date: Optional[str] = date_from_filename(source_path)
        detected_grains: set[str] = set()
        header_strategy = "unknown"

        if not source_path.exists():
            return ReportProfile(
                report_code=report_code.upper(),
                source_file=source_path.name,
                profiled_at=utc_now(),
                sheets=[],
                snapshot_date=snapshot_date,
                warnings=[f"File not found: {source_path}"],
            )

        try:
            wb = load_workbook(source_path, read_only=True, data_only=True)
        except Exception as exc:
            return ReportProfile(
                report_code=report_code.upper(),
                source_file=source_path.name,
                profiled_at=utc_now(),
                sheets=[],
                snapshot_date=snapshot_date,
                warnings=[f"Could not open workbook: {exc}"],
            )

        try:
            for ws in wb.worksheets:
                non_empty_rows = 0
                first_non_empty_row = None
                candidate_headers: List[Any] = []
                sample_rows: List[List[Any]] = []
                local_dates: List[str] = []
                for r_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
                    values = list(row)
                    if any(v not in (None, "") for v in values):
                        non_empty_rows += 1
                        if first_non_empty_row is None:
                            first_non_empty_row = r_idx
                        if len(sample_rows) < 5:
                            sample_rows.append([norm(v) for v in values[:18]])
                        # choose the first row with at least three non-empty text-like labels as header candidate
                        non_empty_cells = [v for v in values if v not in (None, "")]
                        if not candidate_headers and len(non_empty_cells) >= 3:
                            candidate_headers = values[: min(len(values), 40)]
                        if not snapshot_date or len(local_dates) < 3:
                            for v in values[:12]:
                                iso = parse_date(v)
                                if iso:
                                    local_dates.append(iso)
                                    if not snapshot_date:
                                        snapshot_date = iso
                    if r_idx >= 50:
                        break

                header_keys = [norm_key(h) for h in candidate_headers if norm_key(h)]
                grain = self._detect_grain(report_code.upper(), ws.title, header_keys)
                if grain:
                    detected_grains.add(grain)
                if header_keys and header_strategy == "unknown":
                    header_strategy = "flat_or_multirow_detected"

                sheets.append({
                    "sheet_name": ws.title,
                    "max_row": ws.max_row,
                    "max_column": ws.max_column,
                    "non_empty_rows_sampled": non_empty_rows,
                    "first_non_empty_row": first_non_empty_row,
                    "candidate_headers": [norm(h) for h in candidate_headers[:30]],
                    "header_keys": header_keys[:30],
                    "detected_grain": grain,
                    "sample_rows": sample_rows,
                    "dates_seen": local_dates[:5],
                })
        finally:
            try:
                wb.close()
            except Exception:
                pass

        if not sheets:
            warnings.append("Workbook has no readable sheets")

        return ReportProfile(
            report_code=report_code.upper(),
            source_file=source_path.name,
            profiled_at=utc_now(),
            sheets=sheets,
            snapshot_date=snapshot_date,
            detected_grains=sorted(detected_grains),
            header_strategy=header_strategy,
            warnings=warnings,
        )

    @staticmethod
    def _detect_grain(report_code: str, sheet_name: str, header_keys: List[str]) -> Optional[str]:
        joined = " ".join(header_keys + [norm_key(sheet_name)])
        if report_code in {"R18", "R02", "R04", "R08", "R36", "R01"}:
            if "daily" in joined:
                return "date_entity"
            if "project" in joined:
                return "project_snapshot"
            return "entity_snapshot"
        if any(k in joined for k in ["unit", "booking", "account", "customer"]):
            return "account_unit_snapshot"
        if any(k in joined for k in ["collector", "executive", "agent"]):
            return "collector_snapshot"
        if any(k in joined for k in ["project", "tower"]):
            return "project_snapshot"
        if "tat" in joined or "receipt" in joined or "pr" in joined:
            return "process_snapshot"
        return None
