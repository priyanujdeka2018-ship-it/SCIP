"""
SCIP source_adapters.py — Batch 1 ingestion foundation

Purpose:
- Provide source-specific adapters for complex R-series workbooks.
- Implement the R18 OD adapter using the locked Group/Sobha/UAQ hierarchy.
- Generate metric-level lineage and validation results for every OD metric.

Runtime dependencies: stdlib + openpyxl only.
"""

from __future__ import annotations

from dataclasses import dataclass, field, asdict
from datetime import datetime, date
from pathlib import Path
from typing import Any, Dict, List, Optional

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

# User-approved general reconciliation tolerance: 0.05%.
TOLERANCE_PCT = 0.0005


ENTITY_HIERARCHY = {
    "group": {"display": "Group", "parent": None, "children": ["sobha", "uaq"]},
    "sobha": {"display": "Sobha", "parent": "group", "children": ["sobha_dubai", "sobha_auh"]},
    "sobha_dubai": {"display": "Sobha Dubai", "parent": "sobha", "children": []},
    "sobha_auh": {"display": "Sobha AUH", "parent": "sobha", "children": []},
    "uaq": {"display": "UAQ", "parent": "group", "children": ["siniya", "downtown_uaq"]},
    "siniya": {"display": "Siniya", "parent": "uaq", "children": []},
    "downtown_uaq": {"display": "Downtown UAQ", "parent": "uaq", "children": []},
}


@dataclass
class LineageRecord:
    metric_key: str
    metric_label: str
    value: Optional[float]
    unit: str
    source_code: str
    source_file: str
    sheet: str
    cell_or_range: str
    snapshot_date: Optional[str]
    extraction_method: str
    entity_scope: str
    business_definition: str
    validation_status: str
    confidence_state: str
    last_loaded_at: str


@dataclass
class AdapterResult:
    source_code: str
    status: str
    metrics: Dict[str, Any] = field(default_factory=dict)
    rows: List[Dict[str, Any]] = field(default_factory=list)
    lineage: Dict[str, Dict[str, Any]] = field(default_factory=dict)
    warnings: List[str] = field(default_factory=list)
    errors: List[str] = field(default_factory=list)
    validations: Dict[str, Any] = field(default_factory=dict)

    def to_dict(self) -> Dict[str, Any]:
        return asdict(self)


class BaseAdapter:
    source_code = "BASE"

    def __init__(self, path: Path | str):
        self.path = Path(path)
        self.loaded_at = datetime.utcnow().isoformat() + "Z"

    def extract(self) -> AdapterResult:
        raise NotImplementedError

    @staticmethod
    def _to_float(value: Any) -> Optional[float]:
        if value is None:
            return None
        try:
            return float(value)
        except (TypeError, ValueError):
            return None

    @staticmethod
    def _norm(value: Any) -> str:
        return str(value or "").strip().lower()

    @staticmethod
    def _date_to_iso(value: Any) -> Optional[str]:
        if isinstance(value, datetime):
            return value.date().isoformat()
        if isinstance(value, date):
            return value.isoformat()
        if value is None:
            return None
        text = str(value).strip()
        for fmt in ("%Y-%m-%d", "%d-%m-%Y", "%d-%m-%y", "%d/%m/%Y", "%d/%m/%y"):
            try:
                return datetime.strptime(text, fmt).date().isoformat()
            except ValueError:
                pass
        return None

    @staticmethod
    def _currency_cell(row_idx: int, col_idx_0_based: int) -> str:
        return f"{get_column_letter(col_idx_0_based + 1)}{row_idx}"

    @staticmethod
    def _pct_tolerance(reference_value: Optional[float], pct: float = TOLERANCE_PCT) -> float:
        if reference_value is None:
            return 0.0
        return abs(float(reference_value)) * pct

    @classmethod
    def _within_pct_tolerance(
        cls,
        left: Optional[float],
        right: Optional[float],
        reference_value: Optional[float] = None,
        pct: float = TOLERANCE_PCT,
    ) -> Dict[str, Any]:
        if left is None or right is None:
            return {
                "passed": False,
                "left": left,
                "right": right,
                "diff": None,
                "tolerance_pct": pct,
                "tolerance_aed": None,
            }
        ref = reference_value if reference_value is not None else max(abs(float(left)), abs(float(right)))
        tolerance = cls._pct_tolerance(ref, pct)
        diff = abs(float(left) - float(right))
        return {
            "passed": diff <= tolerance,
            "left": left,
            "right": right,
            "diff": diff,
            "tolerance_pct": pct,
            "tolerance_aed": tolerance,
        }


class R18OverdueAdapter(BaseAdapter):
    """Extract OD and ageing from R18 Overdue-1 using positional row scans."""

    source_code = "R18"
    sheet_name = "Overdue-1"

    # openpyxl row tuple indices are 0-based. Excel total column is O.
    total_col = 14
    ageing_cols = {
        "ageing_0_30": 8,
        "ageing_31_60": 9,
        "ageing_61_90": 10,
        "ageing_91_120": 11,
        "ageing_121_180": 12,
        "ageing_180plus": 13,
    }

    # Locked hierarchy mapping approved by user.
    subtotal_to_entity = {
        "sub total (a)": "sobha_dubai",
        "sub total (b)": "sobha_dubai",
        "sub total (c)": "sobha_dubai",
        "sub total (d)": "sobha_dubai",
        "sub total (e)": "sobha_dubai",
        "sub total (f)": "siniya",
        "sub total (g)": "downtown_uaq",
        "sub total (h)": "sobha_auh",
    }

    metric_labels = {
        "OD_SOBHA_DUBAI": "Overdue - Sobha Dubai",
        "OD_SOBHA_AUH": "Overdue - Sobha AUH",
        "OD_SOBHA": "Overdue - Sobha",
        "OD_SINIYA": "Overdue - Siniya",
        "OD_DOWNTOWN_UAQ": "Overdue - Downtown UAQ",
        "OD_DT": "Overdue - Downtown UAQ alias",
        "OD_UAQ": "Overdue - UAQ",
        "OD_GROUP": "Overdue - Group roll-up",
        "OD_TODAY": "Overdue - R18 Grand Total",
    }

    def extract(self) -> AdapterResult:
        result = AdapterResult(source_code=self.source_code, status="unknown")
        if not self.path.exists():
            result.status = "unavailable"
            result.errors.append(f"File not found: {self.path}")
            return result

        try:
            wb = load_workbook(self.path, data_only=True, read_only=True)
        except Exception as exc:
            result.status = "unavailable"
            result.errors.append(f"Could not open workbook: {exc}")
            return result

        try:
            if self.sheet_name not in wb.sheetnames:
                result.status = "unavailable"
                result.errors.append(f"Missing sheet: {self.sheet_name}")
                return result

            ws = wb[self.sheet_name]
            snapshot_date = self._find_snapshot_date(ws)
            subtotals: Dict[str, float] = {
                "sobha_dubai": 0.0,
                "sobha_auh": 0.0,
                "siniya": 0.0,
                "downtown_uaq": 0.0,
            }
            subtotal_rows_seen: Dict[str, int] = {}
            subtotal_cells_by_entity: Dict[str, List[str]] = {k: [] for k in subtotals}
            subtotal_values_by_label: Dict[str, float] = {}
            grand_total: Optional[float] = None
            grand_total_row: Optional[int] = None
            ageing: Dict[str, Optional[float]] = {}

            for row_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
                row_text = [self._norm(v) for v in row]

                # R18 labels are in column D in the current sample, but scan whole row
                # because historic R18 exports have shifted merged label columns.
                for cell_text in row_text:
                    if cell_text in self.subtotal_to_entity:
                        entity = self.subtotal_to_entity[cell_text]
                        value = self._to_float(row[self.total_col] if len(row) > self.total_col else None) or 0.0
                        subtotals[entity] += value
                        subtotal_rows_seen[cell_text] = row_idx
                        subtotal_values_by_label[cell_text] = value
                        subtotal_cells_by_entity[entity].append(self._currency_cell(row_idx, self.total_col))
                        break

                if "grand total" in row_text:
                    grand_total = self._to_float(row[self.total_col] if len(row) > self.total_col else None)
                    grand_total_row = row_idx
                    for key, idx in self.ageing_cols.items():
                        ageing[key] = self._to_float(row[idx] if len(row) > idx else None)

            od_sobha = subtotals["sobha_dubai"] + subtotals["sobha_auh"]
            od_uaq = subtotals["siniya"] + subtotals["downtown_uaq"]
            od_group = od_sobha + od_uaq

            metrics = {
                "OD_SOBHA_DUBAI": subtotals["sobha_dubai"],
                "OD_SOBHA_AUH": subtotals["sobha_auh"],
                "OD_SOBHA": od_sobha,
                "OD_SINIYA": subtotals["siniya"],
                "OD_DOWNTOWN_UAQ": subtotals["downtown_uaq"],
                "OD_DT": subtotals["downtown_uaq"],  # backward-compatible alias only
                "OD_UAQ": od_uaq,
                "OD_GROUP": od_group,
                "OD_TODAY": grand_total,
                "SNAPSHOT_DATE": snapshot_date,
                **ageing,
            }
            result.metrics = metrics

            result.rows = [
                {
                    "entity_key": "sobha_dubai",
                    "entity_label": "Sobha Dubai",
                    "parent_key": "sobha",
                    "metric_key": "OD_SOBHA_DUBAI",
                    "value_aed": subtotals["sobha_dubai"],
                    "source_cells": subtotal_cells_by_entity["sobha_dubai"],
                },
                {
                    "entity_key": "sobha_auh",
                    "entity_label": "Sobha AUH",
                    "parent_key": "sobha",
                    "metric_key": "OD_SOBHA_AUH",
                    "value_aed": subtotals["sobha_auh"],
                    "source_cells": subtotal_cells_by_entity["sobha_auh"],
                },
                {
                    "entity_key": "siniya",
                    "entity_label": "Siniya",
                    "parent_key": "uaq",
                    "metric_key": "OD_SINIYA",
                    "value_aed": subtotals["siniya"],
                    "source_cells": subtotal_cells_by_entity["siniya"],
                },
                {
                    "entity_key": "downtown_uaq",
                    "entity_label": "Downtown UAQ",
                    "parent_key": "uaq",
                    "metric_key": "OD_DOWNTOWN_UAQ",
                    "value_aed": subtotals["downtown_uaq"],
                    "source_cells": subtotal_cells_by_entity["downtown_uaq"],
                },
            ]

            # Validations use the approved 0.05% tolerance.
            result.validations["R18_REQUIRED_SUBTOTALS"] = self._validate_required_subtotals(subtotal_rows_seen)
            result.validations["R18_OD_SOBHA_PLUS_UAQ_EQUALS_OD_TODAY"] = self._within_pct_tolerance(od_sobha + od_uaq, grand_total, grand_total)
            result.validations["R18_OD_GROUP_EQUALS_GRAND_TOTAL"] = self._within_pct_tolerance(od_group, grand_total, grand_total)
            result.validations["R18_SOBHA_EQUALS_CHILDREN"] = self._within_pct_tolerance(od_sobha, subtotals["sobha_dubai"] + subtotals["sobha_auh"], od_sobha)
            result.validations["R18_UAQ_EQUALS_CHILDREN"] = self._within_pct_tolerance(od_uaq, subtotals["siniya"] + subtotals["downtown_uaq"], od_uaq)
            if grand_total is not None and ageing:
                ageing_sum = sum(v or 0 for v in ageing.values())
                result.validations["R18_AGEING_BUCKETS_EQUAL_TOTAL"] = self._within_pct_tolerance(ageing_sum, grand_total, grand_total)

            failed = [k for k, v in result.validations.items() if not v.get("passed", False)]
            if failed:
                confidence = "live_warning"
                validation_status = "warning"
                result.status = "warning"
                result.warnings.append("Failed validations: " + ", ".join(failed))
            else:
                confidence = "live_validated"
                validation_status = "passed"
                result.status = "ok"

            lineage_specs = {
                "OD_SOBHA_DUBAI": (
                    "sobha_dubai",
                    ",".join(subtotal_cells_by_entity["sobha_dubai"]),
                    "Sobha Dubai OD equals R18 Sub Total A-E total OD.",
                ),
                "OD_SOBHA_AUH": (
                    "sobha_auh",
                    ",".join(subtotal_cells_by_entity["sobha_auh"]),
                    "Sobha AUH OD equals R18 Sub Total H total OD.",
                ),
                "OD_SOBHA": (
                    "sobha",
                    "rollup: OD_SOBHA_DUBAI + OD_SOBHA_AUH",
                    "Sobha parent OD equals Sobha Dubai plus Sobha AUH.",
                ),
                "OD_SINIYA": (
                    "siniya",
                    ",".join(subtotal_cells_by_entity["siniya"]),
                    "Siniya OD equals R18 Sub Total F total OD.",
                ),
                "OD_DOWNTOWN_UAQ": (
                    "downtown_uaq",
                    ",".join(subtotal_cells_by_entity["downtown_uaq"]),
                    "Downtown UAQ OD equals R18 Sub Total G total OD.",
                ),
                "OD_DT": (
                    "downtown_uaq",
                    "alias: OD_DOWNTOWN_UAQ",
                    "Backward-compatible alias for Downtown UAQ OD; do not use as a hierarchy label.",
                ),
                "OD_UAQ": (
                    "uaq",
                    "rollup: OD_SINIYA + OD_DOWNTOWN_UAQ",
                    "UAQ parent OD equals Siniya plus Downtown UAQ.",
                ),
                "OD_GROUP": (
                    "group",
                    "rollup: OD_SOBHA + OD_UAQ",
                    "Group OD equals Sobha plus UAQ.",
                ),
                "OD_TODAY": (
                    "group",
                    self._currency_cell(grand_total_row, self.total_col) if grand_total_row else "Grand Total row not found",
                    "R18 Grand Total OD, used as OD_TODAY.",
                ),
            }
            for metric_key, (entity_scope, cell_or_range, definition) in lineage_specs.items():
                result.lineage[metric_key] = asdict(LineageRecord(
                    metric_key=metric_key,
                    metric_label=self.metric_labels[metric_key],
                    value=metrics.get(metric_key),
                    unit="AED",
                    source_code=self.source_code,
                    source_file=self.path.name,
                    sheet=self.sheet_name,
                    cell_or_range=cell_or_range,
                    snapshot_date=snapshot_date,
                    extraction_method="positional_row_scan",
                    entity_scope=entity_scope,
                    business_definition=definition,
                    validation_status=validation_status,
                    confidence_state=confidence,
                    last_loaded_at=self.loaded_at,
                ))

            # Add ageing lineage too, because ageing is part of the OD trust contract.
            for key, idx in self.ageing_cols.items():
                result.lineage[key] = asdict(LineageRecord(
                    metric_key=key,
                    metric_label=f"OD ageing {key.replace('ageing_', '').replace('_', '-')}",
                    value=metrics.get(key),
                    unit="AED",
                    source_code=self.source_code,
                    source_file=self.path.name,
                    sheet=self.sheet_name,
                    cell_or_range=self._currency_cell(grand_total_row, idx) if grand_total_row else "Grand Total row not found",
                    snapshot_date=snapshot_date,
                    extraction_method="positional_row_scan",
                    entity_scope="group",
                    business_definition="Group OD ageing bucket from R18 Grand Total row.",
                    validation_status=validation_status,
                    confidence_state=confidence,
                    last_loaded_at=self.loaded_at,
                ))

            # Helpful extraction audit for smoke files.
            result.metrics["_subtotal_values_by_label"] = subtotal_values_by_label
            result.metrics["_subtotal_rows_seen"] = subtotal_rows_seen

            return result
        finally:
            try:
                wb.close()
            except Exception:
                pass

    def _find_snapshot_date(self, ws) -> Optional[str]:
        # R18 sample has snapshot date in B3. Scan early visible cells to tolerate exports.
        for row in ws.iter_rows(min_row=1, max_row=6, values_only=True):
            for value in row[:10]:
                iso = self._date_to_iso(value)
                if iso:
                    return iso
        return None

    @staticmethod
    def _validate_required_subtotals(subtotal_rows_seen: Dict[str, int]) -> Dict[str, Any]:
        required = {f"sub total ({letter})" for letter in "abcdefgh"}
        seen = set(subtotal_rows_seen)
        missing = sorted(required - seen)
        return {"passed": not missing, "missing": missing, "seen": sorted(seen)}




class R04FinanceDailyAdapter(BaseAdapter):
    """Extract finance daily collections from R04 daily/month_target sheets.

    R04 is a Finance daily report. In this workbook, column C is labelled
    "Collection Due" and is treated by SCIP as Finance D+A / collection-due.
    It is not a pure MDO dues metric and R04 does not expose a separate
    advance split. Advance split is therefore explicitly marked unavailable
    in R04 lineage and sourced from R02/R08 where available.
    """

    source_code = "R04"
    daily_sheet = "daily"
    month_target_sheet = "month_target"

    def extract(self) -> AdapterResult:
        result = AdapterResult(source_code=self.source_code, status="unknown")
        if not self.path.exists():
            result.status = "unavailable"
            result.errors.append(f"File not found: {self.path}")
            return result

        try:
            wb = load_workbook(self.path, data_only=True, read_only=False)
        except Exception as exc:
            result.status = "unavailable"
            result.errors.append(f"Could not open workbook: {exc}")
            return result

        try:
            missing = [s for s in [self.daily_sheet, self.month_target_sheet] if s not in wb.sheetnames]
            result.validations["R04_REQUIRED_SHEETS_PRESENT"] = {"passed": not missing, "missing": missing}
            if missing:
                result.status = "unavailable"
                result.errors.append("Missing required sheets: " + ", ".join(missing))
                return result

            ws = wb[self.daily_sheet]
            snapshot_date = self._date_to_iso(ws["F3"].value) or self._find_first_date(ws, max_row=5)

            target_row = 5
            prorata_row = 6
            header_row = 8
            total_row = self._find_row_by_first_value(ws, "total mtd")
            month_collection_row = self._find_row_by_first_value(ws, "collection for the month")

            daily_rows: List[Dict[str, Any]] = []
            daily_sums = {
                "collection_due_aed": 0.0,
                "new_sales_aed": 0.0,
                "total_collections_aed": 0.0,
                "dld_oqood_aed": 0.0,
            }
            day_total_failures: List[Dict[str, Any]] = []

            if total_row is None:
                result.errors.append("Could not locate Total MTD row in R04 daily sheet")
                result.status = "unavailable"
                return result

            for row_idx in range(header_row + 1, total_row):
                dt = ws.cell(row_idx, 2).value
                iso_dt = self._date_to_iso(dt)
                if not iso_dt:
                    continue
                # Keep visible MTD actual days only. Future zero rows stay out of DAILY_DAYS.
                if snapshot_date and iso_dt > snapshot_date:
                    continue
                collection_due = self._to_float(ws.cell(row_idx, 3).value) or 0.0
                new_sales = self._to_float(ws.cell(row_idx, 4).value) or 0.0
                total = self._to_float(ws.cell(row_idx, 5).value) or 0.0
                dld = self._to_float(ws.cell(row_idx, 6).value) or 0.0
                working_day = ws.cell(row_idx, 1).value

                daily_rows.append({
                    "source_code": self.source_code,
                    "reporting_basis": "Finance",
                    "entity_key": "group",
                    "collection_date": iso_dt,
                    "working_day_no": int(working_day) if isinstance(working_day, (int, float)) else None,
                    "collection_due_aed": collection_due,
                    "finance_da_aed": collection_due,
                    "new_sales_aed": new_sales,
                    "total_collections_aed": total,
                    "dld_oqood_aed": dld,
                    "source_row": row_idx,
                    "source_cells": {
                        "collection_due_aed": f"C{row_idx}",
                        "new_sales_aed": f"D{row_idx}",
                        "total_collections_aed": f"E{row_idx}",
                        "dld_oqood_aed": f"F{row_idx}",
                    },
                })
                daily_sums["collection_due_aed"] += collection_due
                daily_sums["new_sales_aed"] += new_sales
                daily_sums["total_collections_aed"] += total
                daily_sums["dld_oqood_aed"] += dld
                check = self._within_pct_tolerance(collection_due + new_sales, total, total)
                if not check["passed"]:
                    day_total_failures.append({"row": row_idx, **check})

            source_totals = {
                "mtd_da_total": self._to_float(ws.cell(total_row, 3).value),
                "mtd_ns_total": self._to_float(ws.cell(total_row, 4).value),
                "mtd_total_collections": self._to_float(ws.cell(total_row, 5).value),
                "mtd_dld_oqood_total": self._to_float(ws.cell(total_row, 6).value),
                "month_target_da": self._to_float(ws.cell(target_row, 3).value),
                "month_target_ns": self._to_float(ws.cell(target_row, 4).value),
                "month_target_total": self._to_float(ws.cell(target_row, 5).value),
                "mtd_prorata_da_target": self._to_float(ws.cell(prorata_row, 3).value),
                "mtd_prorata_ns_target": self._to_float(ws.cell(prorata_row, 4).value),
                "mtd_prorata_total_target": self._to_float(ws.cell(prorata_row, 5).value),
            }

            metrics = {
                "SNAPSHOT_DATE": snapshot_date,
                "DAILY_DAYS": [self._format_day(r["collection_date"]) for r in daily_rows],
                "DAILY_FINANCE_DA_SERIES": [r["finance_da_aed"] for r in daily_rows],
                "DAILY_NEW_SALES_SERIES": [r["new_sales_aed"] for r in daily_rows],
                "DAILY_TOTAL_COLLECTIONS_SERIES": [r["total_collections_aed"] for r in daily_rows],
                "mtd_da_total": source_totals["mtd_da_total"],
                # Backward-compatible alias. Label clearly as Finance collection-due / D+A.
                "mtd_dues_total": source_totals["mtd_da_total"],
                "mtd_advance_total": None,
                "mtd_advance_status": "unavailable_in_R04_no_advance_split",
                "mtd_ns_total": source_totals["mtd_ns_total"],
                "mtd_total_collections": source_totals["mtd_total_collections"],
                "mtd_dld_oqood_total": source_totals["mtd_dld_oqood_total"],
                **source_totals,
            }
            result.metrics = metrics
            result.rows = daily_rows

            result.validations["R04_DAILY_SUM_EQUALS_TOTAL_MTD_DA"] = self._within_pct_tolerance(
                daily_sums["collection_due_aed"], source_totals["mtd_da_total"], source_totals["mtd_da_total"]
            )
            result.validations["R04_DAILY_SUM_EQUALS_TOTAL_MTD_NEW_SALES"] = self._within_pct_tolerance(
                daily_sums["new_sales_aed"], source_totals["mtd_ns_total"], source_totals["mtd_ns_total"]
            )
            result.validations["R04_DAILY_SUM_EQUALS_TOTAL_MTD_TOTAL_COLLECTIONS"] = self._within_pct_tolerance(
                daily_sums["total_collections_aed"], source_totals["mtd_total_collections"], source_totals["mtd_total_collections"]
            )
            result.validations["R04_MTD_DA_PLUS_NS_EQUALS_TOTAL"] = self._within_pct_tolerance(
                (source_totals["mtd_da_total"] or 0) + (source_totals["mtd_ns_total"] or 0),
                source_totals["mtd_total_collections"],
                source_totals["mtd_total_collections"],
            )
            result.validations["R04_DAILY_DA_PLUS_NS_EQUALS_TOTAL_BY_DAY"] = {
                "passed": not day_total_failures,
                "failures": day_total_failures[:20],
                "failure_count": len(day_total_failures),
            }
            result.validations["R04_FINANCE_TARGET_LABEL_PRESENT"] = {
                "passed": self._norm(ws.cell(4, 3).value) == "collection due" and self._norm(ws.cell(4, 4).value) == "new sales",
                "labels": {"C4": ws.cell(4, 3).value, "D4": ws.cell(4, 4).value, "E4": ws.cell(4, 5).value},
            }
            if month_collection_row:
                result.validations["R04_COLLECTION_FOR_MONTH_EQUALS_TOTAL_MTD"] = self._within_pct_tolerance(
                    self._to_float(ws.cell(month_collection_row, 5).value),
                    source_totals["mtd_total_collections"],
                    source_totals["mtd_total_collections"],
                )

            failed = [k for k, v in result.validations.items() if not v.get("passed", False)]
            result.status = "warning" if failed else "ok"
            confidence = "live_warning" if failed else "live_validated"
            validation_status = "warning" if failed else "passed"
            if failed:
                result.warnings.append("Failed validations: " + ", ".join(failed))

            lineage_specs = {
                "mtd_da_total": ("MTD Finance D+A / collection-due collections", "C" + str(total_row), source_totals["mtd_da_total"], "Finance"),
                "mtd_dues_total": ("Backward-compatible alias for R04 Finance D+A / collection-due; not pure MDO dues", "C" + str(total_row), source_totals["mtd_da_total"], "Finance"),
                "mtd_ns_total": ("MTD new-sales collections", "D" + str(total_row), source_totals["mtd_ns_total"], "Finance"),
                "mtd_total_collections": ("MTD total collections = Finance D+A plus new-sales", "E" + str(total_row), source_totals["mtd_total_collections"], "Finance"),
                "mtd_dld_oqood_total": ("MTD DLD/Oqood collections", "F" + str(total_row), source_totals["mtd_dld_oqood_total"], "Finance"),
                "month_target_da": ("Monthly Finance D+A / collection-due target", "C5", source_totals["month_target_da"], "Finance"),
                "month_target_ns": ("Monthly new-sales target", "D5", source_totals["month_target_ns"], "Finance"),
                "month_target_total": ("Monthly total collections target", "E5", source_totals["month_target_total"], "Finance"),
                "mtd_prorata_da_target": ("MTD pro-rata Finance D+A / collection-due target", "C6", source_totals["mtd_prorata_da_target"], "Finance"),
                "mtd_prorata_ns_target": ("MTD pro-rata new-sales target", "D6", source_totals["mtd_prorata_ns_target"], "Finance"),
                "mtd_prorata_total_target": ("MTD pro-rata total target", "E6", source_totals["mtd_prorata_total_target"], "Finance"),
            }
            for metric_key, (definition, cell, value, basis) in lineage_specs.items():
                result.lineage[metric_key] = self._lineage(metric_key, metric_key, value, self.daily_sheet, cell, snapshot_date, definition, validation_status, confidence, basis)

            result.lineage["mtd_advance_total"] = self._lineage(
                "mtd_advance_total",
                "MTD advance collections unavailable in R04",
                None,
                self.daily_sheet,
                "not exposed by R04 daily workbook",
                snapshot_date,
                "R04 does not expose pure advance split. Use R02 MDO advance actuals or R08 advance summary for advance analysis.",
                "unavailable",
                "unavailable",
                "Finance",
            )
            result.metrics["_daily_sum_actuals"] = daily_sums
            result.metrics["_total_row"] = total_row
            return result
        finally:
            try:
                wb.close()
            except Exception:
                pass

    def _lineage(self, metric_key, metric_label, value, sheet, cell, snapshot_date, definition, validation_status, confidence, reporting_basis):
        rec = asdict(LineageRecord(
            metric_key=metric_key,
            metric_label=metric_label,
            value=value,
            unit="AED" if value is not None else "n/a",
            source_code=self.source_code,
            source_file=self.path.name,
            sheet=sheet,
            cell_or_range=cell,
            snapshot_date=snapshot_date,
            extraction_method="positional_daily_sheet_scan",
            entity_scope="group",
            business_definition=definition,
            validation_status=validation_status,
            confidence_state=confidence,
            last_loaded_at=self.loaded_at,
        ))
        rec["reporting_basis"] = reporting_basis
        return rec

    def _find_row_by_first_value(self, ws, text: str) -> Optional[int]:
        target = text.strip().lower()
        for row in ws.iter_rows():
            v = self._norm(row[0].value if row else None)
            if v == target:
                return row[0].row
        return None

    def _find_first_date(self, ws, max_row=5) -> Optional[str]:
        for row in ws.iter_rows(min_row=1, max_row=max_row):
            for cell in row:
                iso = self._date_to_iso(cell.value)
                if iso:
                    return iso
        return None

    @staticmethod
    def _format_day(iso_date: str) -> str:
        try:
            return datetime.strptime(iso_date, "%Y-%m-%d").strftime("%d %b")
        except Exception:
            return iso_date


class R02MDOAdapter(BaseAdapter):
    """Extract MDO monthly targets and actuals from R02 MDO Dynamic matrix."""

    source_code = "R02"
    sheet_name = "MDO Dynamic"

    section_entity_map = {
        "total collections": "group",
        "siniya total collections": "siniya",
        "sobha total collections": "sobha_dubai",
        "downtown uaq total collections": "downtown_uaq",
        "dowtown uaq total collections": "downtown_uaq",
        "sobha auh total collections": "sobha_auh",
    }

    metric_map = {
        "dues target": "dues_target_aed",
        "dues": "dues_actual_aed",
        "advance target": "advance_target_aed",
        "current year advance": "current_year_advance_actual_aed",
        "future year advance": "future_year_advance_actual_aed",
        "advance": "advance_actual_aed",
        "dues+ advance target": "da_target_aed",
        "dues+ advance": "da_actual_aed",
        "new sales target": "new_sales_target_aed",
        "new sales": "new_sales_actual_aed",
        "total collections target": "total_collections_target_aed",
        "mtd total collections": "total_collections_actual_aed",
    }

    month_aliases = {
        "jan": "jan", "january": "jan",
        "feb": "feb", "february": "feb",
        "mar": "mar", "march": "mar",
        "apr": "apr", "april": "apr",
        "may": "may",
        "jun": "jun", "june": "jun",
        "jul": "jul", "july": "jul",
        "aug": "aug", "august": "aug",
        "sep": "sep", "sept": "sep", "september": "sep",
        "oct": "oct", "october": "oct",
        "nov": "nov", "november": "nov",
        "dec": "dec", "december": "dec",
        "q1": "q1", "q2": "q2", "q3": "q3", "h1": "h1", "total": "total",
    }

    def extract(self) -> AdapterResult:
        result = AdapterResult(source_code=self.source_code, status="unknown")
        if not self.path.exists():
            result.status = "unavailable"
            result.errors.append(f"File not found: {self.path}")
            return result

        try:
            wb = load_workbook(self.path, data_only=True, read_only=False)
        except Exception as exc:
            result.status = "unavailable"
            result.errors.append(f"Could not open workbook: {exc}")
            return result

        try:
            if self.sheet_name not in wb.sheetnames:
                result.status = "unavailable"
                result.errors.append(f"Missing sheet: {self.sheet_name}")
                result.validations["R02_REQUIRED_SHEETS_PRESENT"] = {"passed": False, "missing": [self.sheet_name]}
                return result

            ws = wb[self.sheet_name]
            snapshot_date = self._find_first_date(ws, max_row=5)
            sections = self._parse_sections(ws)
            result.rows = self._rows_from_sections(ws, sections, snapshot_date)
            result.metrics = self._build_metrics(ws, sections, snapshot_date)
            result.validations = self._build_validations(sections)

            failed = [k for k, v in result.validations.items() if not v.get("passed", False)]
            result.status = "warning" if failed else "ok"
            confidence = "live_warning" if failed else "live_validated"
            validation_status = "warning" if failed else "passed"
            if failed:
                result.warnings.append("Failed validations: " + ", ".join(failed))

            for metric_key, meta in result.metrics.get("_lineage_specs", {}).items():
                result.lineage[metric_key] = self._lineage(
                    metric_key=metric_key,
                    metric_label=meta["label"],
                    value=meta["value"],
                    sheet=self.sheet_name,
                    cell=meta["cell"],
                    snapshot_date=snapshot_date,
                    definition=meta["definition"],
                    validation_status=validation_status,
                    confidence=confidence,
                    entity_scope=meta.get("entity", "group"),
                    reporting_basis="MDO",
                    extraction_method=meta.get("extraction_method", "positional_monthly_matrix"),
                )

            # Keep payload clean for app consumers; smoke script can inspect lineage instead.
            result.metrics["lineage_metric_count"] = len(result.lineage)
            result.metrics.pop("_lineage_specs", None)
            return result
        finally:
            try:
                wb.close()
            except Exception:
                pass

    def _parse_sections(self, ws) -> Dict[str, Dict[str, Any]]:
        sections: Dict[str, Dict[str, Any]] = {}
        row = 1
        while row <= ws.max_row:
            title = self._norm(ws.cell(row, 2).value)
            if title in self.section_entity_map:
                entity = self.section_entity_map[title]
                header_row = row + 1
                month_cols = self._month_columns(ws, header_row)
                section_metrics: Dict[str, Dict[str, Any]] = {}
                scan = header_row + 1
                while scan <= ws.max_row:
                    label_norm = self._norm(ws.cell(scan, 2).value)
                    if label_norm in self.section_entity_map and scan != row:
                        break
                    if label_norm.startswith("*note"):
                        break
                    if label_norm in self.metric_map:
                        metric_key = self.metric_map[label_norm]
                        section_metrics[metric_key] = {
                            "row": scan,
                            "label": ws.cell(scan, 2).value,
                            "values": {mk: self._to_float(ws.cell(scan, col).value) for mk, col in month_cols.items()},
                            "cells": {mk: f"{get_column_letter(col)}{scan}" for mk, col in month_cols.items()},
                        }
                    scan += 1
                sections[entity] = {
                    "title": ws.cell(row, 2).value,
                    "title_row": row,
                    "header_row": header_row,
                    "month_cols": month_cols,
                    "metrics": section_metrics,
                }
                row = scan
            else:
                row += 1
        return sections

    def _month_columns(self, ws, header_row: int) -> Dict[str, int]:
        cols: Dict[str, int] = {}
        for col in range(1, ws.max_column + 1):
            raw = ws.cell(header_row, col).value
            if raw is None:
                continue
            key = self.month_aliases.get(self._norm(raw))
            if key:
                cols[key] = col
        return cols

    def _rows_from_sections(self, ws, sections, snapshot_date: Optional[str]) -> List[Dict[str, Any]]:
        rows: List[Dict[str, Any]] = []
        for entity, sec in sections.items():
            for metric_key, m in sec["metrics"].items():
                for period_key, value in m["values"].items():
                    rows.append({
                        "source_code": self.source_code,
                        "reporting_basis": "MDO",
                        "entity_key": entity,
                        "parent_key": ENTITY_HIERARCHY.get(entity, {}).get("parent"),
                        "period_key": period_key,
                        "metric_key": metric_key,
                        "metric_label": m["label"],
                        "value_aed": value,
                        "source_sheet": self.sheet_name,
                        "source_cell": m["cells"].get(period_key),
                        "snapshot_date": snapshot_date,
                    })
        return rows

    def _build_metrics(self, ws, sections, snapshot_date: Optional[str]) -> Dict[str, Any]:
        metrics: Dict[str, Any] = {"SNAPSHOT_DATE": snapshot_date}
        specs: Dict[str, Dict[str, Any]] = {}

        # Aggregate metrics needed by existing data_loader and Batch 2 smoke tests.
        targets = [
            ("fy_dues_target_group", "group", "dues_target_aed", "total", "FY MDO dues target for Group"),
            ("fy_advance_target_group", "group", "advance_target_aed", "total", "FY MDO advance target for Group"),
            ("fy_da_target_group", "group", "da_target_aed", "total", "FY MDO D+A target for Group"),
            ("fy_new_sales_target_group", "group", "new_sales_target_aed", "total", "FY MDO new-sales target for Group"),
            ("fy_total_collections_target_group", "group", "total_collections_target_aed", "total", "FY MDO total collections target for Group"),
            ("may_dues_target_group", "group", "dues_target_aed", "may", "May MDO dues target for Group"),
            ("may_advance_target_group", "group", "advance_target_aed", "may", "May MDO advance target for Group"),
            ("may_da_target_group", "group", "da_target_aed", "may", "May MDO D+A target for Group"),
            ("may_new_sales_target_group", "group", "new_sales_target_aed", "may", "May MDO new-sales target for Group"),
            ("may_total_collections_target_group", "group", "total_collections_target_aed", "may", "May MDO total collections target for Group"),
            ("may_dues_actual_group", "group", "dues_actual_aed", "may", "May MDO dues actual for Group"),
            ("may_current_year_advance_actual_group", "group", "current_year_advance_actual_aed", "may", "May MDO current-year advance actual for Group"),
            ("may_future_year_advance_actual_group", "group", "future_year_advance_actual_aed", "may", "May MDO future-year advance actual for Group"),
            ("may_da_actual_group", "group", "da_actual_aed", "may", "May MDO D+A actual for Group"),
            ("may_new_sales_actual_group", "group", "new_sales_actual_aed", "may", "May MDO new-sales actual for Group"),
            ("may_total_collections_actual_group", "group", "total_collections_actual_aed", "may", "May MDO total collections actual for Group"),
        ]
        for out_key, entity, src_metric, period, definition in targets:
            val, cell = self._metric_cell(sections, entity, src_metric, period)
            metrics[out_key] = val
            specs[out_key] = {
                "label": out_key,
                "value": val,
                "cell": cell or "not found",
                "definition": definition,
                "entity": entity,
            }

        cy = metrics.get("may_current_year_advance_actual_group") or 0.0
        fy = metrics.get("may_future_year_advance_actual_group") or 0.0
        metrics["may_advance_actual_group"] = cy + fy
        specs["may_advance_actual_group"] = {
            "label": "may_advance_actual_group",
            "value": metrics["may_advance_actual_group"],
            "cell": "rollup: may_current_year_advance_actual_group + may_future_year_advance_actual_group",
            "definition": "May MDO advance actual equals current-year advance plus future-year advance.",
            "entity": "group",
            "extraction_method": "derived_rollup",
        }

        metrics["_lineage_specs"] = specs
        metrics["_sections_detected"] = {k: {"title": v["title"], "title_row": v["title_row"]} for k, v in sections.items()}
        return metrics

    def _metric_cell(self, sections, entity: str, metric: str, period: str) -> tuple[Optional[float], Optional[str]]:
        sec = sections.get(entity, {})
        item = sec.get("metrics", {}).get(metric, {})
        return item.get("values", {}).get(period), item.get("cells", {}).get(period)

    def _build_validations(self, sections) -> Dict[str, Any]:
        validations: Dict[str, Any] = {}
        required = ["group", "siniya", "sobha_dubai", "downtown_uaq", "sobha_auh"]
        missing = [e for e in required if e not in sections]
        validations["R02_REQUIRED_SECTIONS_PRESENT"] = {"passed": not missing, "missing": missing, "seen": sorted(sections)}

        # Roll up leaf entities to Group for key May and FY target metrics.
        leaf_entities = ["sobha_dubai", "sobha_auh", "siniya", "downtown_uaq"]
        checks = [
            ("R02_MAY_DUES_TARGET_GROUP_EQUALS_LEAF_ROLLUP", "dues_target_aed", "may"),
            ("R02_MAY_ADVANCE_TARGET_GROUP_EQUALS_LEAF_ROLLUP", "advance_target_aed", "may"),
            ("R02_MAY_DA_TARGET_GROUP_EQUALS_LEAF_ROLLUP", "da_target_aed", "may"),
            ("R02_MAY_TOTAL_COLLECTIONS_TARGET_GROUP_EQUALS_LEAF_ROLLUP", "total_collections_target_aed", "may"),
            ("R02_FY_DUES_TARGET_GROUP_EQUALS_LEAF_ROLLUP", "dues_target_aed", "total"),
            ("R02_FY_ADVANCE_TARGET_GROUP_EQUALS_LEAF_ROLLUP", "advance_target_aed", "total"),
            ("R02_FY_DA_TARGET_GROUP_EQUALS_LEAF_ROLLUP", "da_target_aed", "total"),
            ("R02_FY_TOTAL_COLLECTIONS_TARGET_GROUP_EQUALS_LEAF_ROLLUP", "total_collections_target_aed", "total"),
        ]
        for rule_id, metric, period in checks:
            group_val, _ = self._metric_cell(sections, "group", metric, period)
            leaf_sum = 0.0
            missing_vals = []
            for entity in leaf_entities:
                val, _ = self._metric_cell(sections, entity, metric, period)
                if val is None:
                    missing_vals.append(entity)
                else:
                    leaf_sum += val
            check = self._within_pct_tolerance(leaf_sum, group_val, group_val)
            check["missing_values"] = missing_vals
            validations[rule_id] = check
        return validations

    def _find_first_date(self, ws, max_row=5) -> Optional[str]:
        for row in ws.iter_rows(min_row=1, max_row=max_row):
            for cell in row:
                iso = self._date_to_iso(cell.value)
                if iso:
                    return iso
        return None

    def _lineage(self, metric_key, metric_label, value, sheet, cell, snapshot_date, definition, validation_status, confidence, entity_scope, reporting_basis, extraction_method):
        rec = asdict(LineageRecord(
            metric_key=metric_key,
            metric_label=metric_label,
            value=value,
            unit="AED",
            source_code=self.source_code,
            source_file=self.path.name,
            sheet=sheet,
            cell_or_range=cell,
            snapshot_date=snapshot_date,
            extraction_method=extraction_method,
            entity_scope=entity_scope,
            business_definition=definition,
            validation_status=validation_status,
            confidence_state=confidence,
            last_loaded_at=self.loaded_at,
        ))
        rec["reporting_basis"] = reporting_basis
        return rec


class SheetPresenceAdapter(BaseAdapter):
    """Validation-only scaffold used by future adapters when not yet implemented."""

    required_sheets: List[str] = []

    def extract(self) -> AdapterResult:
        result = AdapterResult(source_code=self.source_code, status="unknown")
        if not self.path.exists():
            result.status = "unavailable"
            result.errors.append(f"File not found: {self.path}")
            return result
        try:
            wb = load_workbook(self.path, data_only=True, read_only=True)
            missing = [s for s in self.required_sheets if s not in wb.sheetnames]
            result.metrics["sheetnames"] = wb.sheetnames
            result.validations[f"{self.source_code}_REQUIRED_SHEETS_PRESENT"] = {"passed": not missing, "missing": missing}
            result.status = "ok" if not missing else "warning"
            return result
        except Exception as exc:
            result.status = "unavailable"
            result.errors.append(str(exc))
            return result
        finally:
            try:
                wb.close()
            except Exception:
                pass


class R08AdvanceAdapter(SheetPresenceAdapter):
    source_code = "R08"
    required_sheets = ["Summary", "Rebate Summary", "Advance 2026CYFY", "Siniya CY FY", "DT CY FY"]


class R36MilestoneCohortAdapter(SheetPresenceAdapter):
    source_code = "R36"
    required_sheets = ["Total", "Active", "Sobha", "Siniya", "DT"]


ADAPTERS = {
    "R18": R18OverdueAdapter,
    "R02": R02MDOAdapter,
    "R04": R04FinanceDailyAdapter,
    "R08": R08AdvanceAdapter,
    "R36": R36MilestoneCohortAdapter,
}


def run_adapter(r_code: str, path: str | Path) -> AdapterResult:
    adapter_cls = ADAPTERS.get(r_code.upper())
    if not adapter_cls:
        raise ValueError(f"No adapter registered for {r_code}")
    return adapter_cls(path).extract()
