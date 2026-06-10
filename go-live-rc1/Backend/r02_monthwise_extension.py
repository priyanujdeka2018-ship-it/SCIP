"""
SCIP — R02 Monthwise MDO extension (planned-target basis)
==========================================================
Additive post-processing for R02MDOAdapter. Parses the "Monthwise MDO" sheet
(ORIGINAL PLANNED targets) and emits governed planned-target metrics plus
Q1 achievement-vs-planned percentages, closing the M060/D2 governance gap:
"MDO Dynamic" backfills elapsed-month targets to actuals, so achievement
computed from it reads ~100%. The true denominator is the planned basis.

Wiring (surgical, one import + one call):
    # source_adapters.py — top of file
    from r02_monthwise_extension import extend_r02_with_monthwise
    # R02MDOAdapter.extract() — inside the try block, immediately before
    # the final `return result` (after metrics/validations are built):
    extend_r02_with_monthwise(result, self.path, self.loaded_at)

Hard rules honoured:
  - Purely additive: only ADDS metrics/lineage/validations keys. Never
    mutates or removes existing keys (asserted in the smoke).
  - No silent fallback: missing sheet / unreadable cell -> value None,
    confidence_state "unavailable", validation failed/disclosed. Never a
    substituted number. Zero denominators -> achievement suppressed.
  - All values AED base units. Lineage carries the full contract per key.

Runtime deps: stdlib + openpyxl only.
"""

from __future__ import annotations

from datetime import datetime, date
from pathlib import Path
from typing import Any, Dict, Optional

from openpyxl import load_workbook

MONTHWISE_SHEET = "Monthwise MDO"
TOLERANCE_PCT = 0.0005  # locked 0.05% reconciliation tolerance

# Entity block titles (col B, lowercased) -> entity key. Group only is
# promoted in this drop; entity planned blocks are parsed but not promoted.
ENTITY_TITLE_MAP = {
    "total collections": "group",
    "siniya total collections": "siniya",
    "sobha total collections": "sobha_dubai",
    "downtown uaq total collections": "downtown_uaq",
    "dowtown uaq total collections": "downtown_uaq",
    "sobha auh total collections": "sobha_auh",
}

# Row labels (col B, lowercased) -> normalised metric. Monthwise group block
# has no CY/FY advance split — single "advance" row.
METRIC_ROW_MAP = {
    "dues target": "dues_target_aed",
    "dues": "dues_actual_aed",
    "advance target": "advance_target_aed",
    "advance": "advance_actual_aed",
    "dues+ advance target": "da_target_aed",
    "dues+ advance": "da_actual_aed",
    "new sales target": "new_sales_target_aed",
    "new sales": "new_sales_actual_aed",
    "total collections target": "total_collections_target_aed",
    "mtd total collections": "total_collections_actual_aed",
}

PERIOD_COLUMNS = {
    "jan": "C", "feb": "D", "mar": "E", "q1": "F", "apr": "G", "may": "H",
    "jun": "I", "q2": "J", "h1": "K", "jul": "L", "aug": "M", "sep": "N",
    "q3": "O", "oct": "P", "nov": "Q", "dec": "R", "total": "S",
}

SNAPSHOT_CELL = "S3"

# Governed surface for this drop: Group planned targets at q1/may/fy, the
# Q1 actuals needed for achievement, and the achievement pcts themselves.
_TARGET_METRICS = (
    ("dues_target_aed", "dues"),
    ("advance_target_aed", "advance"),
    ("da_target_aed", "da"),
    ("new_sales_target_aed", "ns"),
    ("total_collections_target_aed", "total_collections"),
)
_ACTUAL_METRICS = (
    ("dues_actual_aed", "dues"),
    ("advance_actual_aed", "advance"),
    ("da_actual_aed", "da"),
    ("new_sales_actual_aed", "ns"),
    ("total_collections_actual_aed", "total_collections"),
)
_ACHIEVEMENT_CATEGORIES = ("dues", "advance", "ns", "total_collections")


def _norm(v: Any) -> str:
    return str(v).strip().lower() if v is not None else ""


def _num(v: Any) -> Optional[float]:
    # data_only=True yields None for formulas referencing absent sheets —
    # treat None as unreadable, never zero.
    if isinstance(v, bool):
        return None
    if isinstance(v, (int, float)):
        return float(v)
    return None


def _date_iso(v: Any) -> Optional[str]:
    if isinstance(v, datetime):
        return v.date().isoformat()
    if isinstance(v, date):
        return v.isoformat()
    return None


def _lineage_record(
    metric_key: str,
    label: str,
    value: Optional[float],
    cell_or_range: str,
    snapshot_date: Optional[str],
    definition: str,
    validation_status: str,
    confidence_state: str,
    source_file: str,
    loaded_at: str,
    unit: str = "AED",
) -> Dict[str, Any]:
    return {
        "metric_key": metric_key,
        "metric_label": label,
        "value": value,
        "unit": unit if value is not None else "n/a",
        "source_code": "R02",
        "source_file": source_file,
        "sheet": MONTHWISE_SHEET,
        "cell_or_range": cell_or_range,
        "snapshot_date": snapshot_date,
        "extraction_method": "section_iterator_monthwise",
        "entity_scope": "group",
        "business_definition": definition,
        "validation_status": validation_status,
        "confidence_state": confidence_state,
        "last_loaded_at": loaded_at,
        "reporting_basis": "MDO planned",
    }


def _parse_sections(ws) -> Dict[str, Dict[str, Any]]:
    """Scan col B for entity block titles; map metric rows inside each block."""
    sections: Dict[str, Dict[str, Any]] = {}
    current: Optional[Dict[str, Any]] = None
    for row in range(2, ws.max_row + 1):
        raw = ws.cell(row=row, column=2).value
        label = _norm(raw)
        if not label:
            continue
        if label in ENTITY_TITLE_MAP:
            current = {"title": str(raw).strip(), "title_row": row, "metrics": {}}
            sections[ENTITY_TITLE_MAP[label]] = current
            continue
        if current is None:
            continue
        metric = METRIC_ROW_MAP.get(label)
        if metric and metric not in current["metrics"]:
            current["metrics"][metric] = row
    return sections


def _cell(ws, row: int, period: str):
    col = PERIOD_COLUMNS[period]
    ref = f"{col}{row}"
    return _num(ws[ref].value), ref


def _mark_all_unavailable(result, source_file: str, loaded_at: str, reason: str) -> None:
    """Sheet missing/unreadable: every extension key is explicitly None."""
    keys = []
    for _, short in _TARGET_METRICS:
        for period in ("q1", "may", "fy"):
            keys.append(f"planned_{short}_target_{period}_group")
    for _, short in _ACTUAL_METRICS:
        keys.append(f"q1_{short}_actual_monthwise_group")
    for cat in _ACHIEVEMENT_CATEGORIES:
        keys.append(f"q1_achievement_planned_{cat}_pct_group")
    keys.append("q1_mdo_achievement_by_category")

    for key in keys:
        _add_metric(result, key, None)
        unit = "pct" if key.endswith("_pct_group") or key == "q1_mdo_achievement_by_category" else "AED"
        _add_lineage(result, key, _lineage_record(
            key, key, None, "sheet not readable", None,
            f"Unavailable: {reason}", "warning", "unavailable",
            source_file, loaded_at, unit=unit))


def _add_metric(result, key: str, value) -> None:
    if key in result.metrics:
        result.warnings.append(f"r02_monthwise: refusing to overwrite existing metric {key}")
        return
    result.metrics[key] = value


def _add_lineage(result, key: str, record: Dict[str, Any]) -> None:
    if key in result.lineage:
        result.warnings.append(f"r02_monthwise: refusing to overwrite existing lineage {key}")
        return
    result.lineage[key] = record


def extend_r02_with_monthwise(result, path, loaded_at: Optional[str] = None):
    """Additively extend an R02 AdapterResult with Monthwise planned metrics.

    `result` is duck-typed: needs .metrics (dict), .lineage (dict),
    .validations (dict), .warnings (list). Never raises on data problems.
    """
    path = Path(path)
    loaded_at = loaded_at or datetime.utcnow().isoformat()
    source_file = path.name

    try:
        wb = load_workbook(path, data_only=True, read_only=False)
    except Exception as exc:
        result.validations["R02_MONTHWISE_SHEET_PRESENT"] = {"passed": False, "error": str(exc)}
        _mark_all_unavailable(result, source_file, loaded_at, f"workbook unreadable: {exc}")
        return result

    try:
        if MONTHWISE_SHEET not in wb.sheetnames:
            result.validations["R02_MONTHWISE_SHEET_PRESENT"] = {
                "passed": False, "missing": [MONTHWISE_SHEET]}
            _mark_all_unavailable(result, source_file, loaded_at,
                                  f"sheet '{MONTHWISE_SHEET}' not present")
            return result

        ws = wb[MONTHWISE_SHEET]
        result.validations["R02_MONTHWISE_SHEET_PRESENT"] = {"passed": True}
        snapshot = _date_iso(ws[SNAPSHOT_CELL].value)

        sections = _parse_sections(ws)
        result.metrics["_monthwise_sections_detected"] = {
            k: {"title": v["title"], "title_row": v["title_row"]} for k, v in sections.items()}
        group = sections.get("group")
        result.validations["R02_MONTHWISE_GROUP_BLOCK_PRESENT"] = {"passed": group is not None}
        if group is None:
            _mark_all_unavailable(result, source_file, loaded_at,
                                  "group block not found on Monthwise sheet")
            return result

        rows = group["metrics"]
        planned: Dict[str, Dict[str, Optional[float]]] = {}
        # ---- planned targets: q1 / may / fy(total) -------------------------
        for metric, short in _TARGET_METRICS:
            planned[short] = {}
            row = rows.get(metric)
            for period, out_period in (("q1", "q1"), ("may", "may"), ("total", "fy")):
                key = f"planned_{short}_target_{out_period}_group"
                if row is None:
                    value, ref = None, f"row '{metric}' not present"
                else:
                    value, ref = _cell(ws, row, period)
                planned[short][out_period] = value
                _add_metric(result, key, value)
                status = "passed" if value is not None else "warning"
                conf = "live_validated" if value is not None else "unavailable"
                _add_lineage(result, key, _lineage_record(
                    key, f"Planned MDO {short} target ({out_period}, Group)", value,
                    ref, snapshot,
                    "Original planned MDO target from Monthwise sheet — the true "
                    "achievement denominator; MDO Dynamic targets are reconciled "
                    "to actuals for elapsed months.",
                    status, conf, source_file, loaded_at))

        # ---- Q1 actuals (Monthwise basis; reconciled == Dynamic Q1) --------
        q1_actuals: Dict[str, Optional[float]] = {}
        for metric, short in _ACTUAL_METRICS:
            key = f"q1_{short}_actual_monthwise_group"
            row = rows.get(metric)
            if row is None:
                value, ref = None, f"row '{metric}' not present"
            else:
                value, ref = _cell(ws, row, "q1")
            q1_actuals[short] = value
            _add_metric(result, key, value)
            status = "disclosed" if value is not None else "warning"
            conf = "disclosed" if value is not None else "unavailable"
            _add_lineage(result, key, _lineage_record(
                key, f"Q1 MDO {short} actual (Group)", value, ref, snapshot,
                "Q1 actual per R02 note: Q1 months adjusted against actual "
                "collection figures; subject to reconciliation on clearing of "
                "unidentifieds. Disclosed, not pure-source.",
                status, conf, source_file, loaded_at))

        # ---- internal consistency: Q1 col == Jan+Feb+Mar -------------------
        consistency = {}
        for metric, short in _TARGET_METRICS:
            row = rows.get(metric)
            if row is None:
                continue
            q1v, _ = _cell(ws, row, "q1")
            parts = [_cell(ws, row, m)[0] for m in ("jan", "feb", "mar")]
            if q1v is None or any(p is None for p in parts):
                consistency[short] = {"passed": False, "reason": "unreadable cell"}
                continue
            ssum = sum(parts)
            ok = abs(q1v - ssum) <= max(abs(q1v), 1.0) * TOLERANCE_PCT
            consistency[short] = {"passed": ok, "q1_cell": q1v, "jan_feb_mar_sum": ssum}
        result.validations["R02_MONTHWISE_Q1_INTERNAL_CONSISTENCY"] = {
            "passed": all(v.get("passed") for v in consistency.values()) if consistency else False,
            "detail": consistency}

        # ---- achievement vs planned (zero/None denominator suppressed) -----
        ach: Dict[str, Optional[float]] = {}
        for cat in _ACHIEVEMENT_CATEGORIES:
            key = f"q1_achievement_planned_{cat}_pct_group"
            target = planned.get(cat, {}).get("q1")
            actual = q1_actuals.get(cat)
            if target is None or actual is None:
                value, status, conf, why = None, "warning", "unavailable", \
                    "numerator or planned denominator unreadable"
            elif target == 0:
                value, status, conf, why = None, "warning", "unavailable", \
                    "planned target is zero — achievement undefined, suppressed"
            else:
                value = round(actual / target * 100.0, 1)
                status, conf = "disclosed", "disclosed"
                why = ("Q1 actual / Q1 PLANNED MDO target x 100. Disclosed basis: "
                       "Q1 actuals reconciled per source note.")
            ach[cat] = value
            _add_metric(result, key, value)
            _add_lineage(result, key, _lineage_record(
                key, f"Q1 MDO achievement vs planned — {cat} (Group)", value,
                f"Q1 actual ÷ planned Q1 target ({MONTHWISE_SHEET})", snapshot,
                why, status, conf, source_file, loaded_at, unit="pct"))

        # ---- composite headline (binding constraint = worst category) ------
        defined = {c: v for c, v in ach.items() if v is not None}
        comp_key = "q1_mdo_achievement_by_category"
        if defined:
            worst_cat = min(defined, key=defined.get)
            comp_val = defined[worst_cat]
            detail = ", ".join(f"{c} {v}%" for c, v in defined.items())
            _add_metric(result, comp_key, comp_val)
            _add_lineage(result, comp_key, _lineage_record(
                comp_key, "Q1 MDO achievement by category (planned basis)",
                comp_val, f"min over categories ({MONTHWISE_SHEET} Q1)", snapshot,
                f"Worst-category Q1 achievement vs PLANNED MDO target "
                f"(binding constraint = {worst_cat}). All categories: {detail}.",
                "disclosed", "disclosed", source_file, loaded_at, unit="pct"))
        else:
            _add_metric(result, comp_key, None)
            _add_lineage(result, comp_key, _lineage_record(
                comp_key, "Q1 MDO achievement by category (planned basis)",
                None, "no category resolvable", snapshot,
                "Unavailable: no category had both a readable actual and a "
                "non-zero planned target.", "warning", "unavailable",
                source_file, loaded_at, unit="pct"))

        # ---- planned vs dynamic FY parity (warning, disclosed) -------------
        parity = {}
        if "MDO Dynamic" in wb.sheetnames:
            dws = wb["MDO Dynamic"]
            dyn_sections = _parse_dynamic_group_rows(dws)
            for metric, short in _TARGET_METRICS:
                prow, drow = rows.get(metric), dyn_sections.get(metric)
                if prow is None or drow is None:
                    continue
                pv, _ = _cell(ws, prow, "total")
                dv = _num(dws[f"{PERIOD_COLUMNS['total']}{drow}"].value)
                if pv is None or dv is None:
                    parity[short] = {"passed": False, "reason": "unreadable"}
                    continue
                ok = abs(pv - dv) <= max(abs(dv), 1.0) * TOLERANCE_PCT
                parity[short] = {"passed": ok, "planned_fy": pv, "dynamic_fy": dv}
            diverged = sorted(k for k, v in parity.items() if v.get("passed") is False)
            result.validations["R02_PLANNED_VS_DYNAMIC_FY_PARITY"] = {
                "passed": bool(parity), "severity": "warning", "disclosed": True,
                "diverged_categories": diverged,
                "note": ("Planned (Monthwise) FY vs Dynamic FY compared per category. "
                         "Divergence means Dynamic carries REVISED future-month "
                         "targets — disclosed, never silently reconciled."),
                "detail": parity}

        # ---- disclose the basis divergence finding --------------------------
        ns_pct = ach.get("ns")
        result.validations["R02_PLANNED_BASIS_GAP_DISCLOSED"] = {
            "passed": True, "disclosed": True,
            "note": ("Elapsed-month achievement must use the PLANNED (Monthwise) "
                     "basis; MDO Dynamic targets are reconciled to actuals and "
                     "read ~100% by construction."
                     + (f" Q1 new-sales vs planned = {ns_pct}%." if ns_pct is not None else ""))}

        return result
    finally:
        try:
            wb.close()
        except Exception:
            pass


def _parse_dynamic_group_rows(ws) -> Dict[str, int]:
    """Locate the Group block's target rows on MDO Dynamic for FY parity."""
    rows: Dict[str, int] = {}
    in_group = False
    for row in range(2, min(ws.max_row, 80) + 1):
        label = _norm(ws.cell(row=row, column=2).value)
        if not label:
            continue
        if label in ENTITY_TITLE_MAP:
            in_group = ENTITY_TITLE_MAP[label] == "group"
            continue
        if in_group and label in METRIC_ROW_MAP:
            metric = METRIC_ROW_MAP[label]
            rows.setdefault(metric, row)
    return rows
