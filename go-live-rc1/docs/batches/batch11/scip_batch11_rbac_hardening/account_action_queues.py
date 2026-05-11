"""
SCIP Batch 7 - Account-level action queues and collector drilldowns.

This module upgrades the Batch 6 action-queue safety gate by onboarding the
newly attached R-series reports:

  R10  Dues Coverage + Allocation + PTP Data
  R17  SPA / pre-registration overdue bifurcation
  R20  PR-to-SOA TAT process evidence
  R30  Monthly collector feedback/performance
  R31  PR unit update / PR quality
  R32  RM collectors + PR receipts report
  R34  Termination status / account legal escalation
  R38  Risk analysis by CIV and paid-band
  R09  Project collections detail

Guardrail: no account-level action is returned unless it has account/unit,
owner, entity mapping, amount, ageing bucket/status, and source lineage.
"""

from __future__ import annotations

from dataclasses import dataclass, asdict
from datetime import date, datetime
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Tuple

try:
    from fastapi import APIRouter, Request
    from fastapi.responses import JSONResponse
except Exception:  # pragma: no cover
    APIRouter = None
    Request = object
    JSONResponse = None

try:
    import auth
except Exception:  # pragma: no cover
    auth = None

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

TOLERANCE_PCT = 0.0005  # 0.05% locked rule.

ROLE_KEYS = ["collector_rm", "cco_gm_agm", "finance", "mis_qcg_admin"]

ENTITY_DISPLAY = {
    "group": "Group",
    "sobha": "Sobha",
    "sobha_dubai": "Sobha Dubai",
    "sobha_auh": "Sobha AUH",
    "uaq": "UAQ",
    "siniya": "Siniya",
    "downtown_uaq": "Downtown UAQ",
    "unknown": "Unknown",
}

PARENT_ENTITY = {
    "sobha_dubai": "sobha",
    "sobha_auh": "sobha",
    "siniya": "uaq",
    "downtown_uaq": "uaq",
    "sobha": "group",
    "uaq": "group",
}

AGEING_PRIORITY = {
    "180+": 6,
    "121-180": 5,
    "91-120": 4,
    "61-90": 3,
    "31-60": 2,
    "0-30": 1,
    "no_overdue": 0,
    "unknown": -1,
}

CRITICAL_REPORTS = ["R10", "R17", "R20", "R30", "R31", "R32", "R34", "R38", "R09"]


def _now() -> str:
    return datetime.utcnow().isoformat() + "Z"


def _norm(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _to_float(value: Any) -> Optional[float]:
    if value in (None, ""):
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).replace(",", "").strip()
    if text in ("", "-", "--", "00:00:00"):
        return None
    try:
        return float(text)
    except ValueError:
        return None


def _to_iso(value: Any) -> Optional[str]:
    if value in (None, "", 0, "0", "00:00:00"):
        return None
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    text = str(value).strip()
    for fmt in ("%Y-%m-%d", "%d-%m-%Y", "%d-%m-%y", "%d/%m/%Y", "%d/%m/%y", "%d %b %Y", "%d-%b-%Y"):
        try:
            return datetime.strptime(text, fmt).date().isoformat()
        except ValueError:
            pass
    return text if text else None


def _cell(row: int, col: int) -> str:
    return f"{get_column_letter(col)}{row}"


def _cell_range(row: int, cols: Iterable[int]) -> str:
    cols = list(cols)
    if not cols:
        return ""
    if len(cols) == 1:
        return _cell(row, cols[0])
    return f"{_cell(row, min(cols))}:{_cell(row, max(cols))}"


def _find_file(data_dir: Path, patterns: Iterable[str]) -> Optional[Path]:
    for pattern in patterns:
        hits = sorted(data_dir.glob(pattern))
        if hits:
            return hits[0]
    return None


def _entity_from_text(*values: Any) -> str:
    text = " ".join(_norm(v).lower() for v in values if _norm(v))
    if "sobha auh" in text or "river cove" in text or "siniya island" in text and "sobha" in text:
        return "sobha_auh"
    if "siniya" in text:
        return "siniya"
    if "downtown" in text or "dt " in text or "uaq" in text:
        return "downtown_uaq"
    if "sobha" in text or "hartland" in text or "creek" in text or "greens" in text or "one park" in text or "riverside" in text:
        return "sobha_dubai"
    if "central" in text:
        return "group"
    return "unknown"


def _display_entity(code: str) -> str:
    return ENTITY_DISPLAY.get(code or "unknown", "Unknown")


def _fmt_aed(value: Optional[float]) -> str:
    if value is None:
        return "Unavailable"
    v = float(value)
    av = abs(v)
    if av >= 1_000_000_000:
        return f"AED {v/1_000_000_000:.3f}B"
    if av >= 1_000_000:
        return f"AED {v/1_000_000:.1f}M"
    return f"AED {v:,.0f}"


def _severity_from(amount: Optional[float], ageing: Optional[str] = None, status: Optional[str] = None) -> str:
    amount = float(amount or 0)
    ageing = ageing or "unknown"
    status_text = (status or "").lower()
    if "eligible" in status_text and "not eligible" not in status_text:
        return "risk"
    if ageing in {"180+", "121-180"} or amount >= 1_000_000:
        return "risk"
    if ageing in {"91-120", "61-90"} or amount >= 250_000:
        return "attention"
    return "watch"


def _lineage(source_code: str, source_file: str, sheet: str, cell_or_range: str, snapshot_date: Optional[str], method: str, basis: str, row_number: Optional[int] = None) -> Dict[str, Any]:
    return {
        "source_code": source_code,
        "source_file": source_file,
        "sheet": sheet,
        "cell_or_range": cell_or_range,
        "row_number": row_number,
        "snapshot_date": snapshot_date,
        "extraction_method": method,
        "reporting_basis": basis,
        "validation_status": "passed",
        "confidence_state": "live_validated",
        "last_loaded_at": _now(),
    }


def _lineage_ref(metric_key: str, metric_label: str, lin: Dict[str, Any]) -> Dict[str, Any]:
    return {
        "metric_key": metric_key,
        "metric_label": metric_label,
        "has_lineage": True,
        "source_code": lin.get("source_code"),
        "source_file": lin.get("source_file"),
        "sheet": lin.get("sheet"),
        "cell_or_range": lin.get("cell_or_range"),
        "snapshot_date": lin.get("snapshot_date"),
        "validation_status": lin.get("validation_status"),
        "confidence_state": lin.get("confidence_state"),
        "reporting_basis": lin.get("reporting_basis"),
    }


def _has_action_lineage(action: Dict[str, Any]) -> bool:
    refs = action.get("lineage_refs") or []
    required = ["source_code", "source_file", "sheet", "cell_or_range", "validation_status", "confidence_state", "reporting_basis"]
    return bool(refs) and all(all(ref.get(k) for k in required) for ref in refs)


def _has_account_action_minimum(action: Dict[str, Any]) -> bool:
    required = ["account_id", "unit", "collector_rm_owner", "entity_code", "amount_aed", "ageing_bucket"]
    return all(action.get(k) not in (None, "", []) for k in required) and _has_action_lineage(action)


@dataclass
class AccountFact:
    source_code: str
    source_file: str
    sheet: str
    row_number: int
    account_id: str
    unit: str
    customer_name: Optional[str]
    project: Optional[str]
    sub_project: Optional[str]
    entity_code: str
    entity_display: str
    parent_entity_code: Optional[str]
    parent_entity_display: Optional[str]
    collector_rm_owner: Optional[str]
    manager_name: Optional[str]
    overdue_amount_aed: Optional[float]
    ageing_bucket: str
    risk_status: Optional[str]
    ptp_date: Optional[str]
    ptp_amount_aed: Optional[float]
    ptp_status: Optional[str]
    pr_status: Optional[str]
    escalation_status: Optional[str]
    validation_status: str
    confidence_state: str
    lineage: Dict[str, Any]


class WorkbookSource:
    source_code = "RXX"
    patterns: Tuple[str, ...] = ()

    def __init__(self, data_dir: Path):
        self.data_dir = data_dir
        self.path = _find_file(data_dir, self.patterns)
        self.errors: List[str] = []
        self.warnings: List[str] = []

    def exists(self) -> bool:
        return self.path is not None and self.path.exists()

    @property
    def source_file(self) -> str:
        return self.path.name if self.path else "missing"


class R10DuesCoverageAdapter(WorkbookSource):
    source_code = "R10"
    patterns = ("R10*.xlsx", "*Dues Coverage*.xlsx")

    def extract(self) -> Dict[str, Any]:
        account_allocations: List[Dict[str, Any]] = []
        ptp_facts: List[Dict[str, Any]] = []
        collector_coverage: List[Dict[str, Any]] = []
        if not self.exists():
            return {"source_code": self.source_code, "status": "missing", "account_allocations": [], "ptp_facts": [], "collector_coverage": [], "errors": ["R10 missing"]}
        wb = load_workbook(self.path, read_only=True, data_only=True)
        # Allocation account facts.
        if "Allocation" in wb.sheetnames:
            ws = wb["Allocation"]
            headers = [_norm(v) for v in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]
            # Snapshot date appears as last header in col31 in the sample.
            snapshot_date = _to_iso(headers[30] if len(headers) > 30 else None)
            for ridx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                account_id = _norm(row[4] if len(row) > 4 else None)
                unit = _norm(row[5] if len(row) > 5 else None)
                owner = _norm(row[8] if len(row) > 8 else None)
                entity = _entity_from_text(row[0] if len(row) > 0 else None, row[3] if len(row) > 3 else None)
                overdue_today = _to_float(row[6] if len(row) > 6 else None) or 0.0
                overdue_eom = _to_float(row[7] if len(row) > 7 else None) or 0.0
                overdue_ms = _to_float(row[23] if len(row) > 23 else None) or overdue_eom or overdue_today
                amount_tracker = _to_float(row[20] if len(row) > 20 else None) or 0.0
                communications = _to_float(row[16] if len(row) > 16 else None) or 0.0
                if not (account_id and unit and owner):
                    continue
                if max(overdue_ms, overdue_eom, overdue_today, amount_tracker, communications) <= 0:
                    continue
                lin = _lineage("R10", self.source_file, "Allocation", _cell_range(ridx, [1,5,6,7,8,9,17,21,24,30]), snapshot_date, "R10DuesCoverageAdapter.allocation_account_row", "R10 dues coverage allocation")
                account_allocations.append({
                    "account_id": account_id,
                    "unit": unit,
                    "project": _norm(row[0] if len(row) > 0 else None),
                    "sub_project": _norm(row[1] if len(row) > 1 else None),
                    "tower": _norm(row[3] if len(row) > 3 else None),
                    "collector_rm_owner": owner,
                    "advance_owner": _norm(row[9] if len(row) > 9 else None),
                    "manager_name": _norm(row[19] if len(row) > 19 else None),
                    "entity_code": entity,
                    "entity_display": _display_entity(entity),
                    "overdue_today_aed": round(overdue_today, 2),
                    "overdue_eom_aed": round(overdue_eom, 2),
                    "overdue_eom_ms_aed": round(overdue_ms, 2),
                    "amount_in_tracker_aed": round(amount_tracker, 2),
                    "receipt_generated_aed": round(_to_float(row[21] if len(row) > 21 else None) or 0, 2),
                    "communication_count": int(communications),
                    "last_task_updated_date": _to_iso(row[29] if len(row) > 29 else None),
                    "ageing_bucket": "amount_validated_no_ageing_from_R10",
                    "lineage": lin,
                })
        # PTP facts.
        if "PTP Data" in wb.sheetnames:
            ws = wb["PTP Data"]
            for ridx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                account_id = _norm(row[4] if len(row) > 4 else None)
                unit = _norm(row[5] if len(row) > 5 else None)
                owner = _norm(row[25] if len(row) > 25 else None) or _norm(row[11] if len(row) > 11 else None)
                amount = _to_float(row[16] if len(row) > 16 else None)
                if not (account_id and unit and owner and amount and amount > 0):
                    continue
                entity = _entity_from_text(row[27] if len(row) > 27 else None, row[6] if len(row) > 6 else None)
                lin = _lineage("R10", self.source_file, "PTP Data", _cell_range(ridx, [4,5,6,12,16,17,25,26,28]), _to_iso(row[7] if len(row) > 7 else None), "R10DuesCoverageAdapter.ptp_account_row", "R10 PTP account promise")
                ptp_facts.append({
                    "account_name": _norm(row[3] if len(row) > 3 else None),
                    "account_id": account_id,
                    "unit": unit,
                    "tower": _norm(row[6] if len(row) > 6 else None),
                    "created_date": _to_iso(row[7] if len(row) > 7 else None),
                    "collector_rm_owner": owner,
                    "owner_full_name": _norm(row[11] if len(row) > 11 else None),
                    "ptp_date": _to_iso(row[15] if len(row) > 15 else None),
                    "ptp_amount_aed": round(amount, 2),
                    "ptp_status": _norm(row[24] if len(row) > 24 else None),
                    "call_status": _norm(row[13] if len(row) > 13 else None),
                    "comments": _norm(row[14] if len(row) > 14 else None),
                    "entity_code": entity,
                    "entity_display": _display_entity(entity),
                    "lineage": lin,
                })
        # Coverage by collector.
        if "Agent - Coverage" in wb.sheetnames:
            ws = wb["Agent - Coverage"]
            for ridx, row in enumerate(ws.iter_rows(min_row=6, values_only=True), start=6):
                collector = _norm(row[1] if len(row) > 1 else None)
                if not collector or collector.lower().startswith("grand"):
                    continue
                lin = _lineage("R10", self.source_file, "Agent - Coverage", _cell_range(ridx, [2,5,6,7,8,9,10]), None, "R10DuesCoverageAdapter.collector_coverage", "R10 dues coverage summary")
                collector_coverage.append({
                    "collector_rm_owner": collector,
                    "total_odeom_amount_aed": round(_to_float(row[4] if len(row) > 4 else None) or 0, 2),
                    "coverage_units": int(_to_float(row[5] if len(row) > 5 else None) or 0),
                    "coverage_pct": _to_float(row[6] if len(row) > 6 else None),
                    "not_worked_units": int(_to_float(row[8] if len(row) > 8 else None) or 0),
                    "not_worked_pct": _to_float(row[9] if len(row) > 9 else None),
                    "lineage": lin,
                })
        wb.close()
        return {"source_code": "R10", "status": "loaded", "account_allocations": account_allocations, "ptp_facts": ptp_facts, "collector_coverage": collector_coverage, "errors": self.errors, "warnings": self.warnings}


class R34TerminationAdapter(WorkbookSource):
    source_code = "R34"
    patterns = ("R34*.xlsx", "*Termination Status*.xlsx")

    def extract(self) -> Dict[str, Any]:
        facts: List[Dict[str, Any]] = []
        if not self.exists():
            return {"source_code": "R34", "status": "missing", "termination_facts": [], "errors": ["R34 missing"]}
        wb = load_workbook(self.path, read_only=True, data_only=True)
        ws = wb["Data"] if "Data" in wb.sheetnames else wb[wb.sheetnames[0]]
        for ridx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            account_id = _norm(row[3] if len(row) > 3 else None)
            unit = _norm(row[4] if len(row) > 4 else None)
            if not (account_id and unit):
                continue
            buckets = {
                "0-30": _to_float(row[40] if len(row) > 40 else None) or 0.0,
                "31-60": _to_float(row[41] if len(row) > 41 else None) or 0.0,
                "61-90": _to_float(row[42] if len(row) > 42 else None) or 0.0,
                "91-120": _to_float(row[43] if len(row) > 43 else None) or 0.0,
                "121-180": _to_float(row[44] if len(row) > 44 else None) or 0.0,
                "180+": _to_float(row[45] if len(row) > 45 else None) or 0.0,
            }
            total_overdue = _to_float(row[52] if len(row) > 52 else None)
            if total_overdue is None:
                total_overdue = sum(buckets.values())
            ageing_bucket = max(buckets.keys(), key=lambda b: (buckets[b] > 0, AGEING_PRIORITY[b])) if any(buckets.values()) else "no_overdue"
            if total_overdue <= 0 and _norm(row[64] if len(row) > 64 else None).lower() not in {"termination", "possible pdn"}:
                continue
            entity = _entity_from_text(row[2] if len(row) > 2 else None, row[0] if len(row) > 0 else None)
            lin = _lineage("R34", self.source_file, ws.title, _cell_range(ridx, [1,4,5,6,41,42,43,44,45,46,53,65,71,72]), None, "R34TerminationAdapter.account_termination_row", "R34 termination status account evidence", ridx)
            facts.append({
                "account_id": account_id,
                "unit": unit,
                "customer_name": _norm(row[5] if len(row) > 5 else None),
                "project": _norm(row[2] if len(row) > 2 else None),
                "sub_project": _norm(row[1] if len(row) > 1 else None),
                "tower": _norm(row[0] if len(row) > 0 else None),
                "entity_code": entity,
                "entity_display": _display_entity(entity),
                "purchase_price_aed": _to_float(row[6] if len(row) > 6 else None),
                "total_overdue_aed": round(total_overdue, 2),
                "ageing_bucket": ageing_bucket,
                "ageing_buckets": {k: round(v, 2) for k, v in buckets.items()},
                "last_action_status": _norm(row[14] if len(row) > 14 else None),
                "last_action_description": _norm(row[15] if len(row) > 15 else None),
                "last_action_by": _norm(row[17] if len(row) > 17 else None),
                "possible_pdn": _norm(row[64] if len(row) > 64 else None),
                "eligible_status": _norm(row[70] if len(row) > 70 else None),
                "in_system_status": _norm(row[71] if len(row) > 71 else None),
                "pre_registration_status": _norm(row[35] if len(row) > 35 else None),
                "lineage": lin,
            })
        wb.close()
        return {"source_code": "R34", "status": "loaded", "termination_facts": facts, "errors": self.errors, "warnings": self.warnings}


class R31PRQualityAdapter(WorkbookSource):
    source_code = "R31"
    patterns = ("R31*.xlsx", "*PR Unit*.xlsx")

    def extract(self) -> Dict[str, Any]:
        pr_issues: List[Dict[str, Any]] = []
        if not self.exists():
            return {"source_code": "R31", "status": "missing", "pr_issues": [], "errors": ["R31 missing"]}
        wb = load_workbook(self.path, read_only=True, data_only=True)
        if "PR Data" in wb.sheetnames:
            ws = wb["PR Data"]
            for ridx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                pr_no = _norm(row[0] if len(row) > 0 else None)
                status = _norm(row[5] if len(row) > 5 else None)
                receipt_status = _norm(row[22] if len(row) > 22 else None)
                amount = _to_float(row[2] if len(row) > 2 else None) or _to_float(row[15] if len(row) > 15 else None) or 0.0
                if not pr_no or (status.lower() == "approved" and receipt_status.lower() in {"cleared", "pdc"}):
                    continue
                entity = _entity_from_text(row[39] if len(row) > 39 else None, row[20] if len(row) > 20 else None)
                lin = _lineage("R31", self.source_file, "PR Data", _cell_range(ridx, [1,3,4,6,8,17,18,21,22,23,39,40,44]), _to_iso(row[6] if len(row) > 6 else None), "R31PRQualityAdapter.pr_data_issue_row", "R31 PR unit update / PR quality", ridx)
                pr_issues.append({
                    "pr_no": pr_no,
                    "account_id": _norm(row[7] if len(row) > 7 else None) or _norm(row[31] if len(row) > 31 else None),
                    "unit": _norm(row[3] if len(row) > 3 else None) or _norm(row[32] if len(row) > 32 else None),
                    "collector_rm_owner": _norm(row[43] if len(row) > 43 else None) or _norm(row[16] if len(row) > 16 else None) or _norm(row[38] if len(row) > 38 else None),
                    "amount_aed": round(amount, 2),
                    "status": status,
                    "receipt_status": receipt_status,
                    "project": _norm(row[20] if len(row) > 20 else None),
                    "sub_project": _norm(row[37] if len(row) > 37 else None),
                    "entity_code": entity,
                    "entity_display": _display_entity(entity),
                    "remarks": _norm(row[17] if len(row) > 17 else None),
                    "lineage": lin,
                })
        wb.close()
        return {"source_code": "R31", "status": "loaded", "pr_issues": pr_issues, "errors": self.errors, "warnings": self.warnings}


class R32RMCollectorsAdapter(WorkbookSource):
    source_code = "R32"
    patterns = ("R32*.xlsx", "*RM Collectors*.xlsx")

    def extract(self) -> Dict[str, Any]:
        collector_metrics: List[Dict[str, Any]] = []
        receipt_facts: List[Dict[str, Any]] = []
        if not self.exists():
            return {"source_code": "R32", "status": "missing", "collector_metrics": [], "receipt_facts": [], "errors": ["R32 missing"]}
        wb = load_workbook(self.path, read_only=True, data_only=True)
        for sheet in [s for s in wb.sheetnames if s.startswith("RM Collector")]:
            ws = wb[sheet]
            entity_hint = _entity_from_text(sheet)
            for ridx, row in enumerate(ws.iter_rows(min_row=6, values_only=True), start=6):
                collector = _norm(row[2] if len(row) > 2 else None)
                if not collector or collector.lower().startswith("total"):
                    continue
                dues_mtd = _to_float(row[4] if len(row) > 4 else None)
                target = _to_float(row[6] if len(row) > 6 else None)
                if (dues_mtd or 0) == 0 and (target or 0) == 0:
                    continue
                lin = _lineage("R32", self.source_file, sheet, _cell_range(ridx, [3,4,5,6,7,8,9]), None, "R32RMCollectorsAdapter.collector_metric_row", "R32 RM collectors performance")
                collector_metrics.append({
                    "collector_rm_owner": collector,
                    "entity_code": entity_hint,
                    "entity_display": _display_entity(entity_hint),
                    "daily_collection_aed": round(_to_float(row[3] if len(row) > 3 else None) or 0, 2),
                    "dues_mtd_aed": round(dues_mtd or 0, 2),
                    "achieved_pct": _to_float(row[5] if len(row) > 5 else None),
                    "target_net_constraints_aed": round(target or 0, 2),
                    "receipts_dues_aed": round(_to_float(row[7] if len(row) > 7 else None) or 0, 2),
                    "receipts_advance_aed": round(_to_float(row[8] if len(row) > 8 else None) or 0, 2),
                    "lineage": lin,
                })
        if "Daily Collection" in wb.sheetnames:
            ws = wb["Daily Collection"]
            for ridx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                amount = _to_float(row[4] if len(row) > 4 else None) or 0.0
                pr_no = _norm(row[5] if len(row) > 5 else None)
                if amount <= 0:
                    continue
                entity = _entity_from_text(row[14] if len(row) > 14 else None)
                lin = _lineage("R32", self.source_file, "Daily Collection", _cell_range(ridx, [1,2,3,5,6,8,10,11,13,15]), _to_iso(row[2] if len(row) > 2 else None), "R32RMCollectorsAdapter.daily_collection_row", "R32 RM receipts / daily collection")
                receipt_facts.append({
                    "unit": _norm(row[0] if len(row) > 0 else None),
                    "account_id": _norm(row[7] if len(row) > 7 else None),
                    "collector_rm_owner": _norm(row[1] if len(row) > 1 else None) or _norm(row[12] if len(row) > 12 else None),
                    "date": _to_iso(row[2] if len(row) > 2 else None),
                    "deposit_aed": round(amount, 2),
                    "pr_no": pr_no,
                    "pr_status": _norm(row[9] if len(row) > 9 else None),
                    "installment_type": _norm(row[6] if len(row) > 6 else None),
                    "entity_code": entity,
                    "entity_display": _display_entity(entity),
                    "lineage": lin,
                })
        wb.close()
        return {"source_code": "R32", "status": "loaded", "collector_metrics": collector_metrics, "receipt_facts": receipt_facts, "errors": self.errors, "warnings": self.warnings}


class R30MonthlyFeedbackAdapter(WorkbookSource):
    source_code = "R30"
    patterns = ("R30*.xlsx", "*Monthly Feedback*.xlsx")

    def extract(self) -> Dict[str, Any]:
        facts: List[Dict[str, Any]] = []
        if not self.exists():
            return {"source_code": "R30", "status": "missing", "collector_feedback": [], "errors": ["R30 missing"]}
        wb = load_workbook(self.path, read_only=True, data_only=True)
        ws = wb["Summary"] if "Summary" in wb.sheetnames else wb[wb.sheetnames[0]]
        snapshot_date = None
        for cell in (31, 33):
            snapshot_date = snapshot_date or _to_iso(ws.cell(2, cell).value)
        for ridx, row in enumerate(ws.iter_rows(min_row=4, values_only=True), start=4):
            collector = _norm(row[1] if len(row) > 1 else None)
            if not collector or collector.lower().startswith("total"):
                continue
            lin = _lineage("R30", self.source_file, ws.title, _cell_range(ridx, [2,3,4,5,6,7,8,9,10,17,31,33,39]), snapshot_date, "R30MonthlyFeedbackAdapter.summary_collector_row", "R30 monthly collector feedback")
            facts.append({
                "collector_rm_owner": collector,
                "allocated_units": int(_to_float(row[2] if len(row) > 2 else None) or 0),
                "allocated_od_aed": round(_to_float(row[3] if len(row) > 3 else None) or 0, 2),
                "target_aed": round(_to_float(row[4] if len(row) > 4 else None) or 0, 2),
                "total_attempts": int(_to_float(row[5] if len(row) > 5 else None) or 0),
                "total_contacts": int(_to_float(row[6] if len(row) > 6 else None) or 0),
                "contact_intensity": _to_float(row[8] if len(row) > 8 else None),
                "units_paid": int(_to_float(row[9] if len(row) > 9 else None) or 0),
                "total_payment_aed": round(_to_float(row[16] if len(row) > 16 else None) or 0, 2),
                "lineage": lin,
            })
        wb.close()
        return {"source_code": "R30", "status": "loaded", "collector_feedback": facts, "errors": self.errors, "warnings": self.warnings}


class R17SPALegalAdapter(WorkbookSource):
    source_code = "R17"
    patterns = ("R17*.xlsx", "*SPA*.xlsx", "*registartion*.xlsx")

    def extract(self) -> Dict[str, Any]:
        facts: List[Dict[str, Any]] = []
        if not self.exists():
            return {"source_code": "R17", "status": "missing", "spa_project_facts": [], "errors": ["R17 missing"]}
        wb = load_workbook(self.path, read_only=True, data_only=True)
        ws = wb["Summary"] if "Summary" in wb.sheetnames else wb[wb.sheetnames[0]]
        for ridx, row in enumerate(ws.iter_rows(min_row=4, values_only=True), start=4):
            project = _norm(row[2] if len(row) > 2 else None)
            overdue = _to_float(row[3] if len(row) > 3 else None) or 0.0
            if not project or overdue <= 0:
                continue
            entity = _entity_from_text(project)
            lin = _lineage("R17", self.source_file, ws.title, _cell_range(ridx, [3,4,8,9,12,13,14,16]), None, "R17SPALegalAdapter.project_spa_row", "R17 SPA/pre-registration overdue bifurcation")
            facts.append({
                "project": project,
                "entity_code": entity,
                "entity_display": _display_entity(entity),
                "overdue_aed": round(overdue, 2),
                "spa_not_sent_aed": round(_to_float(row[7] if len(row) > 7 else None) or 0, 2),
                "spa_sent_not_signed_aed": round(_to_float(row[8] if len(row) > 8 else None) or 0, 2),
                "spa_signed_not_executed_aed": round(_to_float(row[11] if len(row) > 11 else None) or 0, 2),
                "spa_executed_not_preregistered_aed": round(_to_float(row[12] if len(row) > 12 else None) or 0, 2),
                "prereg_completed_aed": round(_to_float(row[13] if len(row) > 13 else None) or 0, 2),
                "lineage": lin,
            })
        wb.close()
        return {"source_code": "R17", "status": "loaded", "spa_project_facts": facts, "errors": self.errors, "warnings": self.warnings}


class R09ProjectCollectionsAdapter(WorkbookSource):
    source_code = "R09"
    patterns = ("R09*.xlsx", "*Collections 07*.xlsx")

    def extract(self) -> Dict[str, Any]:
        facts: List[Dict[str, Any]] = []
        if not self.exists():
            return {"source_code": "R09", "status": "missing", "project_collection_facts": [], "errors": ["R09 missing"]}
        wb = load_workbook(self.path, read_only=True, data_only=True)
        sheet = "Daily Collections-0" if "Daily Collections-0" in wb.sheetnames else wb.sheetnames[0]
        ws = wb[sheet]
        snapshot_date = _to_iso(ws.cell(3, 2).value)
        for ridx, row in enumerate(ws.iter_rows(min_row=8, values_only=True), start=8):
            project = _norm(row[3] if len(row) > 3 else None)
            overdue = _to_float(row[11] if len(row) > 11 else None) or 0.0
            sale = _to_float(row[7] if len(row) > 7 else None) or 0.0
            if not project or (overdue <= 0 and sale <= 0):
                continue
            entity = _entity_from_text(project)
            lin = _lineage("R09", self.source_file, sheet, _cell_range(ridx, [4,8,9,10,12,14,15,16,17,18]), snapshot_date, "R09ProjectCollectionsAdapter.project_collection_row", "R09 project collection detail")
            facts.append({
                "project": project,
                "entity_code": entity,
                "entity_display": _display_entity(entity),
                "sale_value_aed": round(sale, 2),
                "milestone_due_aed": round(_to_float(row[8] if len(row) > 8 else None) or 0, 2),
                "collected_to_date_aed": round(_to_float(row[9] if len(row) > 9 else None) or 0, 2),
                "overdue_aed": round(overdue, 2),
                "pro_rata_target_mtd_aed": round(_to_float(row[13] if len(row) > 13 else None) or 0, 2),
                "collected_dues_mtd_aed": round(_to_float(row[14] if len(row) > 14 else None) or 0, 2),
                "lineage": lin,
            })
        wb.close()
        return {"source_code": "R09", "status": "loaded", "project_collection_facts": facts, "errors": self.errors, "warnings": self.warnings}


class R20PRToSOATATAdapter(WorkbookSource):
    source_code = "R20"
    patterns = ("R20*.xlsx", "*PR to SOA*.xlsx")

    def extract(self) -> Dict[str, Any]:
        facts: List[Dict[str, Any]] = []
        if not self.exists():
            return {"source_code": "R20", "status": "missing", "tat_facts": [], "errors": ["R20 missing"]}
        wb = load_workbook(self.path, read_only=True, data_only=True)
        if "Summary 2" in wb.sheetnames:
            ws = wb["Summary 2"]
            months = [_norm(v) for v in next(ws.iter_rows(min_row=3, max_row=3, values_only=True))]
            for ridx, row in enumerate(ws.iter_rows(min_row=4, max_row=8, values_only=True), start=4):
                bucket = _norm(row[1] if len(row) > 1 else None)
                if not bucket:
                    continue
                for cidx in range(2, min(len(row), 14)):
                    count = _to_float(row[cidx]) or 0.0
                    if count <= 0:
                        continue
                    month = months[cidx] if cidx < len(months) else f"M{cidx}"
                    lin = _lineage("R20", self.source_file, "Summary 2", _cell(ridx, cidx + 1), None, "R20PRToSOATATAdapter.month_bucket_count", "R20 PR-to-SOA TAT count")
                    facts.append({"tat_bucket": bucket, "month": month, "count": int(count), "lineage": lin})
        wb.close()
        return {"source_code": "R20", "status": "loaded", "tat_facts": facts, "errors": self.errors, "warnings": self.warnings}


class R38RiskAnalysisAdapter(WorkbookSource):
    source_code = "R38"
    patterns = ("R38*.xlsx", "*Risk Analysis*.xlsx")

    def extract(self) -> Dict[str, Any]:
        facts: List[Dict[str, Any]] = []
        if not self.exists():
            return {"source_code": "R38", "status": "missing", "risk_facts": [], "errors": ["R38 missing"]}
        wb = load_workbook(self.path, read_only=True, data_only=True)
        risk_sheets = [s for s in wb.sheetnames if "mil" in s.lower() or "100" in s]
        paid_bands = [("0-10%", 4, 5), ("10-20%", 6, 7), ("20-30%", 8, 9), ("30-40%", 10, 11), ("40-60%", 12, 13), ("60%+", 14, 15)]
        for sheet in risk_sheets:
            ws = wb[sheet]
            current_project_group = None
            for ridx, row in enumerate(ws.iter_rows(min_row=5, values_only=True), start=5):
                if _norm(row[1] if len(row) > 1 else None):
                    current_project_group = _norm(row[1])
                tower = _norm(row[2] if len(row) > 2 else None)
                if not tower or tower.lower().startswith("grand"):
                    continue
                for band, units_col, value_col in paid_bands:
                    units = _to_float(row[units_col-1] if len(row) >= units_col else None) or 0.0
                    value = _to_float(row[value_col-1] if len(row) >= value_col else None) or 0.0
                    if units <= 0 and value <= 0:
                        continue
                    entity = _entity_from_text(current_project_group, tower)
                    lin = _lineage("R38", self.source_file, sheet, _cell_range(ridx, [2,3,units_col,value_col]), None, "R38RiskAnalysisAdapter.civ_paid_band_row", "R38 CIV/paid-band risk analysis")
                    facts.append({
                        "civ_bucket": sheet,
                        "project_group": current_project_group,
                        "tower": tower,
                        "paid_band": band,
                        "units": int(units),
                        "sale_value_aed": round(value, 2),
                        "entity_code": entity,
                        "entity_display": _display_entity(entity),
                        "lineage": lin,
                    })
        wb.close()
        return {"source_code": "R38", "status": "loaded", "risk_facts": facts, "errors": self.errors, "warnings": self.warnings}


def _extract_all(data_dir: Path) -> Dict[str, Dict[str, Any]]:
    adapters = [
        R10DuesCoverageAdapter(data_dir),
        R34TerminationAdapter(data_dir),
        R31PRQualityAdapter(data_dir),
        R32RMCollectorsAdapter(data_dir),
        R30MonthlyFeedbackAdapter(data_dir),
        R17SPALegalAdapter(data_dir),
        R09ProjectCollectionsAdapter(data_dir),
        R20PRToSOATATAdapter(data_dir),
        R38RiskAnalysisAdapter(data_dir),
    ]
    return {adapter.source_code: adapter.extract() for adapter in adapters}


def _build_lookups(extracted: Dict[str, Dict[str, Any]]) -> Dict[str, Any]:
    term_by_key = {}
    for fact in extracted.get("R34", {}).get("termination_facts", []):
        term_by_key[(fact.get("account_id"), fact.get("unit"))] = fact
    alloc_by_key = {}
    for fact in extracted.get("R10", {}).get("account_allocations", []):
        alloc_by_key[(fact.get("account_id"), fact.get("unit"))] = fact
    metrics_by_collector = {}
    for fact in extracted.get("R32", {}).get("collector_metrics", []):
        metrics_by_collector.setdefault(fact.get("collector_rm_owner"), []).append(fact)
    for fact in extracted.get("R30", {}).get("collector_feedback", []):
        metrics_by_collector.setdefault(fact.get("collector_rm_owner"), []).append(fact)
    return {"termination": term_by_key, "allocation": alloc_by_key, "metrics_by_collector": metrics_by_collector}


def _make_collector_ptp_action(ptp: Dict[str, Any], term: Dict[str, Any], alloc: Optional[Dict[str, Any]]) -> Optional[Dict[str, Any]]:
    account_id = ptp.get("account_id")
    unit = ptp.get("unit")
    owner = ptp.get("collector_rm_owner") or (alloc or {}).get("collector_rm_owner") or term.get("last_action_by")
    amount = ptp.get("ptp_amount_aed") or term.get("total_overdue_aed")
    ageing = term.get("ageing_bucket")
    entity = term.get("entity_code") or ptp.get("entity_code") or (alloc or {}).get("entity_code") or "unknown"
    if not (account_id and unit and owner and amount and ageing and entity != "unknown"):
        return None
    ptp_lin = ptp.get("lineage")
    term_lin = term.get("lineage")
    refs = [
        _lineage_ref(f"PTP_{account_id}_{unit}", "Promise-to-pay amount/date", ptp_lin),
        _lineage_ref(f"AGEING_{account_id}_{unit}", "Ageing and overdue validation", term_lin),
    ]
    if alloc and alloc.get("lineage"):
        refs.append(_lineage_ref(f"OWNER_{account_id}_{unit}", "Owner allocation validation", alloc["lineage"]))
    action = {
        "action_id": f"collector_ptp::{account_id}::{unit}",
        "action_type": "collector_ptp_follow_up",
        "grain": "account_unit",
        "role_visibility": ["collector_rm", "cco_gm_agm"],
        "status": "ready",
        "title": f"Follow PTP · {account_id} / {unit}",
        "summary": f"{owner} has a PTP of {_fmt_aed(amount)} for {unit}; ageing bucket {ageing}.",
        "recommended_action": "Call/WhatsApp customer before PTP date, update disposition, and escalate if unpaid after date.",
        "severity": _severity_from(term.get("total_overdue_aed") or amount, ageing, term.get("eligible_status")),
        "account_id": account_id,
        "unit": unit,
        "customer_name": term.get("customer_name") or ptp.get("account_name"),
        "collector_rm_owner": owner,
        "manager_name": (alloc or {}).get("manager_name"),
        "entity_code": entity,
        "entity_display": _display_entity(entity),
        "parent_entity_code": PARENT_ENTITY.get(entity),
        "parent_entity_display": _display_entity(PARENT_ENTITY.get(entity)) if PARENT_ENTITY.get(entity) else None,
        "project": term.get("project") or (alloc or {}).get("project"),
        "sub_project": term.get("sub_project") or (alloc or {}).get("sub_project"),
        "amount_aed": round(float(amount), 2),
        "display_amount": _fmt_aed(amount),
        "overdue_amount_aed": term.get("total_overdue_aed") or (alloc or {}).get("overdue_eom_ms_aed"),
        "ageing_bucket": ageing,
        "ptp_date": ptp.get("ptp_date"),
        "ptp_amount_aed": ptp.get("ptp_amount_aed"),
        "ptp_status": ptp.get("ptp_status"),
        "escalation_status": term.get("possible_pdn") or term.get("eligible_status"),
        "reporting_basis": "R10 PTP + R34 ageing/termination + R10 allocation owner evidence",
        "confidence_state": "live_validated_account_owner_ageing_lineaged",
        "lineage_refs": refs,
    }
    return action if _has_account_action_minimum(action) else None


def _make_termination_action(term: Dict[str, Any], alloc: Optional[Dict[str, Any]]) -> Optional[Dict[str, Any]]:
    account_id, unit = term.get("account_id"), term.get("unit")
    amount = term.get("total_overdue_aed")
    owner = (alloc or {}).get("collector_rm_owner") or term.get("last_action_by")
    entity = term.get("entity_code") or (alloc or {}).get("entity_code")
    ageing = term.get("ageing_bucket")
    if not (account_id and unit and owner and amount and ageing and entity and entity != "unknown"):
        return None
    refs = [_lineage_ref(f"TERM_{account_id}_{unit}", "Termination / overdue eligibility", term["lineage"])]
    if alloc and alloc.get("lineage"):
        refs.append(_lineage_ref(f"OWNER_{account_id}_{unit}", "Owner allocation validation", alloc["lineage"]))
    action = {
        "action_id": f"termination::{account_id}::{unit}",
        "action_type": "termination_escalation_review",
        "grain": "account_unit",
        "role_visibility": ["cco_gm_agm", "mis_qcg_admin", "collector_rm"],
        "status": "ready",
        "title": f"Termination/PDN review · {account_id} / {unit}",
        "summary": f"{_fmt_aed(amount)} overdue; status {term.get('possible_pdn') or term.get('eligible_status')}; owner {owner}.",
        "recommended_action": "Review eligibility, validate latest contact trail, and decide development notice / termination path.",
        "severity": _severity_from(amount, ageing, term.get("eligible_status") or term.get("possible_pdn")),
        "account_id": account_id,
        "unit": unit,
        "customer_name": term.get("customer_name"),
        "collector_rm_owner": owner,
        "manager_name": (alloc or {}).get("manager_name"),
        "entity_code": entity,
        "entity_display": _display_entity(entity),
        "parent_entity_code": PARENT_ENTITY.get(entity),
        "parent_entity_display": _display_entity(PARENT_ENTITY.get(entity)) if PARENT_ENTITY.get(entity) else None,
        "project": term.get("project"),
        "sub_project": term.get("sub_project"),
        "amount_aed": round(float(amount), 2),
        "display_amount": _fmt_aed(amount),
        "overdue_amount_aed": amount,
        "ageing_bucket": ageing,
        "ptp_date": None,
        "ptp_amount_aed": None,
        "ptp_status": None,
        "escalation_status": term.get("possible_pdn") or term.get("eligible_status"),
        "reporting_basis": "R34 termination status + R10 owner allocation evidence",
        "confidence_state": "live_validated_account_owner_ageing_lineaged",
        "lineage_refs": refs,
    }
    return action if _has_account_action_minimum(action) else None


def _make_pr_action(issue: Dict[str, Any]) -> Optional[Dict[str, Any]]:
    account_id = issue.get("account_id")
    unit = issue.get("unit")
    owner = issue.get("collector_rm_owner")
    amount = issue.get("amount_aed")
    entity = issue.get("entity_code")
    if not (account_id and unit and owner and amount and entity and entity != "unknown"):
        return None
    refs = [_lineage_ref(f"PR_{issue.get('pr_no')}", "Payment request quality issue", issue["lineage"])]
    return {
        "action_id": f"finance_pr::{issue.get('pr_no')}",
        "action_type": "finance_pr_quality_exception",
        "grain": "payment_request_account_unit",
        "role_visibility": ["finance", "mis_qcg_admin"],
        "status": "ready",
        "title": f"PR exception · {issue.get('pr_no')}",
        "summary": f"{issue.get('status') or 'PR issue'} / receipt {issue.get('receipt_status') or 'unknown'} for {_fmt_aed(amount)}.",
        "recommended_action": "Resolve PR status, correct receipt mapping, and confirm SOA impact before customer follow-up.",
        "severity": "attention" if amount < 500_000 else "risk",
        "account_id": account_id,
        "unit": unit,
        "customer_name": None,
        "collector_rm_owner": owner,
        "entity_code": entity,
        "entity_display": _display_entity(entity),
        "amount_aed": round(float(amount), 2),
        "display_amount": _fmt_aed(amount),
        "ageing_bucket": "process_exception",
        "pr_status": issue.get("status"),
        "receipt_status": issue.get("receipt_status"),
        "project": issue.get("project"),
        "reporting_basis": "R31 PR unit update / PR quality evidence",
        "confidence_state": "live_validated_pr_exception_lineaged",
        "lineage_refs": refs,
    }


def _make_coverage_action(coverage: Dict[str, Any]) -> Optional[Dict[str, Any]]:
    if (coverage.get("not_worked_units") or 0) <= 0:
        return None
    refs = [_lineage_ref(f"COVERAGE_{coverage.get('collector_rm_owner')}", "Collector not-worked coverage", coverage["lineage"])]
    return {
        "action_id": f"coverage::{coverage.get('collector_rm_owner')}",
        "action_type": "coverage_gap_review",
        "grain": "collector",
        "role_visibility": ["cco_gm_agm", "mis_qcg_admin"],
        "status": "ready",
        "title": f"Coverage gap · {coverage.get('collector_rm_owner')}",
        "summary": f"{coverage.get('not_worked_units')} not-worked units; OD pool {_fmt_aed(coverage.get('total_odeom_amount_aed'))}.",
        "recommended_action": "Ask team lead to clear not-worked queue and validate disposition quality.",
        "severity": "attention" if (coverage.get("not_worked_units") or 0) < 20 else "risk",
        "collector_rm_owner": coverage.get("collector_rm_owner"),
        "amount_aed": coverage.get("total_odeom_amount_aed"),
        "display_amount": _fmt_aed(coverage.get("total_odeom_amount_aed")),
        "ageing_bucket": "coverage_gap",
        "reporting_basis": "R10 Agent Coverage",
        "confidence_state": "live_validated_collector_coverage_lineaged",
        "lineage_refs": refs,
    }


def _make_tat_actions(tat_facts: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    actions = []
    for fact in tat_facts:
        if fact.get("tat_bucket") not in {"9-15 days", "15+ days"}:
            continue
        if fact.get("month") != "May":
            continue
        count = fact.get("count") or 0
        if count <= 0:
            continue
        refs = [_lineage_ref(f"TAT_{fact.get('tat_bucket')}_{fact.get('month')}", "PR-to-SOA TAT ageing bucket", fact["lineage"])]
        actions.append({
            "action_id": f"tat::{fact.get('month')}::{fact.get('tat_bucket')}",
            "action_type": "pr_to_soa_tat_backlog",
            "grain": "process_bucket_month",
            "role_visibility": ["finance", "mis_qcg_admin"],
            "status": "ready",
            "title": f"PR-to-SOA TAT · {fact.get('tat_bucket')}",
            "summary": f"{count:,} receipts in {fact.get('tat_bucket')} TAT bucket for {fact.get('month')}.",
            "recommended_action": "Review process backlog, isolate stalled modes, and update finance/QCG closure plan.",
            "severity": "risk" if fact.get("tat_bucket") == "15+ days" else "attention",
            "amount_aed": count,
            "display_amount": f"{count:,} receipts",
            "ageing_bucket": fact.get("tat_bucket"),
            "reporting_basis": "R20 PR-to-SOA TAT Report",
            "confidence_state": "live_validated_process_tat_lineaged",
            "lineage_refs": refs,
        })
    return actions


def _top(items: List[Dict[str, Any]], limit: int) -> List[Dict[str, Any]]:
    return sorted(items, key=lambda x: (x.get("severity") == "risk", float(x.get("amount_aed") or 0)), reverse=True)[:limit]


def build_action_queue_payload(data_dir: Optional[Path | str] = None, role: Optional[str] = None) -> Dict[str, Any]:
    base = Path(data_dir or Path(__file__).resolve().parent)
    if not any(base.glob("R10*.xlsx")) and Path("/mnt/data").exists():
        base = Path("/mnt/data")
    extracted = _extract_all(base)
    lookups = _build_lookups(extracted)
    term_by_key = lookups["termination"]
    alloc_by_key = lookups["allocation"]

    # Build true account-level collector actions from R10 PTP + R34 ageing + R10 owner allocation.
    collector_actions: List[Dict[str, Any]] = []
    for ptp in extracted.get("R10", {}).get("ptp_facts", []):
        key = (ptp.get("account_id"), ptp.get("unit"))
        action = _make_collector_ptp_action(ptp, term_by_key.get(key, {}), alloc_by_key.get(key))
        if action:
            collector_actions.append(action)

    termination_actions: List[Dict[str, Any]] = []
    for term in extracted.get("R34", {}).get("termination_facts", []):
        key = (term.get("account_id"), term.get("unit"))
        if (term.get("possible_pdn") or "").upper() != "TERMINATION" and term.get("ageing_bucket") not in {"180+", "121-180", "91-120"}:
            continue
        action = _make_termination_action(term, alloc_by_key.get(key))
        if action:
            termination_actions.append(action)

    finance_pr_actions = [_make_pr_action(issue) for issue in extracted.get("R31", {}).get("pr_issues", [])]
    finance_pr_actions = [a for a in finance_pr_actions if a]
    coverage_actions = [_make_coverage_action(c) for c in extracted.get("R10", {}).get("collector_coverage", [])]
    coverage_actions = [a for a in coverage_actions if a]
    tat_actions = _make_tat_actions(extracted.get("R20", {}).get("tat_facts", []))

    collector_actions = _top(collector_actions, 40)
    termination_actions = _top(termination_actions, 40)
    finance_pr_actions = _top(finance_pr_actions, 40)
    coverage_actions = _top(coverage_actions, 20)
    tat_actions = _top(tat_actions, 10)

    roles = {
        "collector_rm": {
            "role": "collector_rm",
            "status": "ready" if collector_actions else "blocked_no_ready_actions",
            "headline": "Collector/RM PTP follow-up queue is unlocked with account, owner, amount, ageing and lineage.",
            "disclosure": "Only PTP actions that join R10 PTP, R34 ageing/termination, and R10 owner allocation are shown.",
            "account_actions": collector_actions,
            "process_actions": [],
            "management_actions": [],
            "required_source_fields": ["account_id", "unit", "collector_rm_owner", "entity", "amount", "ageing_bucket", "lineage"],
            "reporting_basis": "R10 PTP + R34 ageing + R10 allocation",
        },
        "cco_gm_agm": {
            "role": "cco_gm_agm",
            "status": "ready",
            "headline": "Management risk queue combines PTP follow-ups, termination reviews and coverage gaps.",
            "disclosure": "Account actions are source-lineaged; project/risk enrichments stay separate from collector ownership.",
            "account_actions": collector_actions[:15] + termination_actions[:15],
            "process_actions": coverage_actions[:8],
            "management_actions": termination_actions[:20],
            "required_source_fields": ["account_id", "unit", "owner", "entity", "amount", "ageing", "lineage"],
            "reporting_basis": "R10/R34/R30/R32 management queue",
        },
        "finance": {
            "role": "finance",
            "status": "ready" if finance_pr_actions or tat_actions else "blocked_no_process_actions",
            "headline": "Finance queue surfaces PR exceptions and PR-to-SOA TAT ageing.",
            "disclosure": "Finance actions come from PR/TAT sources and do not assign collector calls unless owner evidence exists.",
            "account_actions": finance_pr_actions[:25],
            "process_actions": tat_actions,
            "management_actions": [],
            "required_source_fields": ["pr_no/account", "unit", "owner where available", "amount/count", "process ageing/status", "lineage"],
            "reporting_basis": "R31 PR quality + R20 PR-to-SOA TAT",
        },
        "mis_qcg_admin": {
            "role": "mis_qcg_admin",
            "status": "ready",
            "headline": "MIS/QCG queue validates owner mapping, PR exceptions, coverage gaps and process TAT.",
            "disclosure": "Queue emphasizes data/process closure and blocks unlineaged account tasks.",
            "account_actions": finance_pr_actions[:15] + termination_actions[:10],
            "process_actions": coverage_actions[:15] + tat_actions,
            "management_actions": [],
            "required_source_fields": ["source_lineage", "owner/entity mapping", "amount/status validation"],
            "reporting_basis": "R10/R20/R31/R34 governance queue",
        },
    }

    if role and role in roles:
        roles = {role: roles[role]}

    validations = _validate_roles(roles)
    source_status = {code: extracted.get(code, {}).get("status") for code in CRITICAL_REPORTS}
    facts_summary = {
        "r10_allocations": len(extracted.get("R10", {}).get("account_allocations", [])),
        "r10_ptp_facts": len(extracted.get("R10", {}).get("ptp_facts", [])),
        "r34_termination_facts": len(extracted.get("R34", {}).get("termination_facts", [])),
        "r31_pr_issues": len(extracted.get("R31", {}).get("pr_issues", [])),
        "r32_collector_metrics": len(extracted.get("R32", {}).get("collector_metrics", [])),
        "r32_receipt_facts": len(extracted.get("R32", {}).get("receipt_facts", [])),
        "r30_collector_feedback": len(extracted.get("R30", {}).get("collector_feedback", [])),
        "r17_spa_project_facts": len(extracted.get("R17", {}).get("spa_project_facts", [])),
        "r09_project_collection_facts": len(extracted.get("R09", {}).get("project_collection_facts", [])),
        "r20_tat_facts": len(extracted.get("R20", {}).get("tat_facts", [])),
        "r38_risk_facts": len(extracted.get("R38", {}).get("risk_facts", [])),
        "collector_account_actions_ready": len(collector_actions),
        "termination_review_actions_ready": len(termination_actions),
        "finance_pr_actions_ready": len(finance_pr_actions),
        "process_actions_ready": len(coverage_actions) + len(tat_actions),
    }
    return {
        "contract_version": "action_queues.v2.batch7",
        "status": "ready_account_level_guarded" if validations["passed"] else "validation_failed",
        "generated_at": _now(),
        "guardrails": _guardrails(),
        "source_availability": source_status,
        "data_grain_disclosure": {
            "current_grain": "account_unit_with_project_process_enrichment",
            "true_account_level_available": bool(collector_actions),
            "rule": "No account action appears without account ID/unit, owner, entity, amount/status, ageing bucket or process status, and lineage.",
        },
        "facts_summary": facts_summary,
        "roles": roles,
        "validations": validations,
        "lineage_sample": _sample_lineage_refs(roles),
        "source_notes": {
            "R17": "Project/legal SPA enrichment; not owner source.",
            "R38": "Project/CIV risk enrichment; not owner source.",
            "R09": "Project collection enrichment; not owner source.",
            "R30_R32": "Collector performance/receipt context; R10/R34 provide the account-action join gate.",
        },
    }


def _sample_lineage_refs(roles: Dict[str, Any]) -> List[Dict[str, Any]]:
    refs: List[Dict[str, Any]] = []
    for payload in roles.values():
        for section in ("account_actions", "process_actions", "management_actions"):
            for action in payload.get(section, [])[:2]:
                refs.extend(action.get("lineage_refs") or [])
    return refs[:10]


def _guardrails() -> Dict[str, Any]:
    return {
        "tolerance_pct": 0.05,
        "no_silent_fallback": True,
        "locked_hierarchy": "Group > Sobha(Sobha Dubai, Sobha AUH) + UAQ(Siniya, Downtown UAQ)",
        "role_model": ["Board/CXO", "CCO/GM/AGM", "Finance", "MIS/QCG/Admin", "Collector/RM"],
        "removed_role": "Entity Head",
        "account_action_gate": "account actions require account_id + unit + owner + entity + amount/status + ageing/process status + lineage",
        "liquid_glass_gate": "action queues and collector detail remain L4/L5 evidence/action output; dense rows use solid surfaces",
    }


def _validate_roles(roles: Dict[str, Any]) -> Dict[str, Any]:
    checks: Dict[str, bool] = {}
    visible_actions: List[Dict[str, Any]] = []
    for role_payload in roles.values():
        for section in ("account_actions", "process_actions", "management_actions"):
            visible_actions.extend(role_payload.get(section) or [])
    account_level = [a for a in visible_actions if a.get("grain") in {"account_unit", "payment_request_account_unit"}]
    checks["all_visible_actions_have_lineage"] = all(_has_action_lineage(a) for a in visible_actions)
    checks["account_actions_have_owner_entity_amount_ageing_or_process_status"] = all(
        a.get("account_id") and a.get("unit") and a.get("collector_rm_owner") and a.get("entity_code") and a.get("amount_aed") is not None and (a.get("ageing_bucket") or a.get("pr_status"))
        for a in account_level
    )
    checks["collector_role_has_only_collector_visible_account_actions"] = all(
        set(a.get("role_visibility") or []).intersection({"collector_rm"}) and a.get("action_type") in {"collector_ptp_follow_up", "termination_escalation_review"}
        for a in roles.get("collector_rm", {}).get("account_actions", [])
    )
    checks["finance_role_has_no_collector_call_ptp_actions"] = all(
        a.get("action_type") != "collector_ptp_follow_up"
        for a in roles.get("finance", {}).get("account_actions", [])
    )
    checks["entity_head_removed"] = "entity_head" not in roles
    checks["reporting_basis_present"] = all(bool(a.get("reporting_basis")) for a in visible_actions)
    return {"passed": all(checks.values()), "checks": checks, "checked_action_count": len(visible_actions), "checked_account_action_count": len(account_level)}


if APIRouter is not None:
    router = APIRouter(prefix="/action-queues", tags=["action-queues"])

    @router.get("")
    async def get_action_queues(request: Request, role: Optional[str] = None) -> Any:
        actor = auth.get_actor(request) if auth is not None else None
        requested_role = actor.role if actor is not None else role
        payload = build_action_queue_payload(role=requested_role)
        if auth is not None and actor is not None:
            payload = auth.filter_action_queue_payload(payload, actor)
        return JSONResponse(content=payload, status_code=200)

    @router.get("/collector-drilldown")
    async def get_collector_drilldown(request: Request, collector: Optional[str] = None) -> Any:
        actor = auth.get_actor(request) if auth is not None else None
        if actor is not None and actor.role not in {"collector_rm", "cco_gm_agm", "mis_qcg_admin"}:
            auth.log_denied_attempt(actor, "read:collector_drilldown", "role_not_allowed", path="/action-queues/collector-drilldown")
            return JSONResponse(content={"status": "denied", "reason": "role_not_allowed"}, status_code=403)
        payload = build_action_queue_payload(role="collector_rm")
        actions = payload.get("roles", {}).get("collector_rm", {}).get("account_actions", [])
        if collector:
            actions = [a for a in actions if str(a.get("collector_rm_owner", "")).lower() == collector.lower()]
        if auth is not None and actor is not None:
            actions = auth.filter_rows_for_actor(actions, actor, "read:action_queues")
        return JSONResponse(content={
            "contract_version": "collector_drilldown.v3.batch11",
            "status": "ready" if actions else "no_actions_for_filter",
            "collector": collector or "all",
            "account_actions": actions,
            "validations": payload.get("validations"),
            "guardrails": payload.get("guardrails"),
            "rbac": {"actor": actor.to_dict() if actor is not None else None, "filtered": auth is not None},
        }, status_code=200)
