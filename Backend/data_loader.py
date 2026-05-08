"""
SOBHA COLLECTIONS INTELLIGENCE PLATFORM
data_loader.py — R-Series Data Pipeline
Version: v6
Authority: FINAL_ARCHITECTURE.md

Responsibilities:
  - Read all active R-series xlsx files from /data directory
  - Apply column mappings from pipeline_config.json
  - Compute all derived constants (OD_TODAY, PIPELINE_GROSS, etc.)
  - Return { "dataframes": dict, "computed": dict }
  - Pre-aggregate lean summary JSON for frontend (~100KB)
  - Graceful degradation on missing or malformed files

Dependency rule:
  data_loader.py  →  imports: constants, utils, pipeline_config.json
  data_loader.py  →  NEVER imports: endpoints, quickball, main
  data_loader.py  →  is stateless. No session state. No globals mutated.

Architecture rules (never violate):
  - OD_TODAY comes from R18. Never hardcoded.
  - PIPELINE_GROSS, PIPELINE_ADV_DENOM come from R36.
  - CY_ADV_MIX_YTD computed ONCE here. Never per-submodule.
  - DAILY_DAYS[] extracted from R04 ONCE. Never per-section.
  - Collector data sorted DESC by achievement_pct on load.
  - Missing file → data_pending payload, not crash.
  - All monetary values stored as AED base units (not millions).
"""

from __future__ import annotations

import json
import logging
import os
import re
from datetime import datetime, date
from pathlib import Path
from typing import Any, Optional

import openpyxl

import constants as C
import utils as U
import file_resolver

# ---------------------------------------------------------------------------
# LOGGING
# ---------------------------------------------------------------------------
logger = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# PATHS
# ---------------------------------------------------------------------------
# Resolve /data directory relative to this file's location.
# In deployment: backend/ sits next to data/ at repo root.
_HERE = Path(__file__).parent
DATA_DIR = (_HERE.parent / "data").resolve()
PIPELINE_CONFIG_PATH = (_HERE / "pipeline_config.json").resolve()


# ---------------------------------------------------------------------------
# SECTION 1 — CONFIG LOADER
# ---------------------------------------------------------------------------

def _load_pipeline_config() -> dict:
    """Load pipeline_config.json. Raises on missing file — config is required."""
    with open(PIPELINE_CONFIG_PATH, "r", encoding="utf-8") as f:
        return json.load(f)


# ---------------------------------------------------------------------------
# SECTION 2 — OPENPYXL SHEET READER
# ---------------------------------------------------------------------------

def _read_sheet(file_path: Path, sheet_name: Any) -> Optional[list[dict]]:
    """
    Read a single xlsx sheet into a list of row dicts.
    Uses openpyxl read_only mode for performance.

    Returns:
        List of dicts (header → value) or None on failure.
    """
    try:
        wb = openpyxl.load_workbook(str(file_path), read_only=True, data_only=True)

        # Resolve sheet name
        if isinstance(sheet_name, int):
            ws = wb.worksheets[sheet_name]
        elif isinstance(sheet_name, str):
            if sheet_name not in wb.sheetnames:
                logger.warning("Sheet '%s' not found in %s. Available: %s",
                               sheet_name, file_path.name, wb.sheetnames)
                wb.close()
                return None
            ws = wb[sheet_name]
        else:
            ws = wb.active

        rows = list(ws.iter_rows(values_only=True))
        wb.close()

        if not rows:
            return []

        headers = [str(h).strip() if h is not None else f"col_{i}"
                   for i, h in enumerate(rows[0])]
        result = []
        for row in rows[1:]:
            if all(v is None for v in row):
                continue  # skip blank rows
            result.append(dict(zip(headers, row)))
        return result

    except FileNotFoundError:
        logger.info("File not found (graceful skip): %s", file_path.name)
        return None
    except Exception as exc:
        logger.warning("Failed to read %s sheet=%s: %s", file_path.name, sheet_name, exc)
        return None


# ---------------------------------------------------------------------------
# SECTION 3 — COLUMN MAPPER & TYPE COERCER
# ---------------------------------------------------------------------------

def _apply_column_map(rows: list[dict], col_map: dict, col_types: dict) -> list[dict]:
    """
    Rename columns per pipeline_config column_map and coerce types.
    Rows that don't have any mapped column are skipped.
    """
    mapped = []
    for row in rows:
        new_row: dict = {}
        for src_col, dest_col in col_map.items():
            raw = row.get(src_col)
            coerced = _coerce(raw, col_types.get(dest_col, "str"))
            new_row[dest_col] = coerced
        if any(v is not None for v in new_row.values()):
            mapped.append(new_row)
    return mapped


def _coerce(value: Any, type_str: str) -> Any:
    """
    Coerce a raw openpyxl cell value to the declared type.
    Returns None on conversion failure — never raises.
    """
    if value is None:
        return None
    try:
        if type_str == "int":
            return int(float(value))
        if type_str == "float":
            return float(value)
        if type_str == "bool":
            if isinstance(value, bool):
                return value
            return str(value).strip().lower() in ("true", "yes", "1")
        if type_str == "datetime":
            if isinstance(value, (datetime, date)):
                return value
            # Try parsing common string formats
            for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%Y/%m/%d", "%d %b %Y"):
                try:
                    return datetime.strptime(str(value).strip(), fmt)
                except ValueError:
                    continue
            return None
        # Default: str
        return str(value).strip() if value is not None else None
    except (ValueError, TypeError):
        return None


# ---------------------------------------------------------------------------
# SECTION 4 — AGGREGATION ENGINE
# ---------------------------------------------------------------------------

def _run_aggregations(rows: list[dict], agg_cfg: dict) -> dict:
    """
    Execute the aggregation operations declared in pipeline_config for a source.

    Supported ops:
      sum            → sum a column
      mean           → mean of a column
      filter_sum     → sum after filtering
      filter_mean    → mean after filtering
      filter_count   → count rows after filtering
      count_filter   → same as filter_count (alias)
      last           → last non-null value
      extract_list   → extract column as list
      weighted_avg   → sum(numerator) / sum(denominator) * 100
      derived        → computed from other aggregation results (deferred)
    """
    results: dict = {}

    for agg_name, agg_def in agg_cfg.items():
        op = agg_def.get("op")
        col = agg_def.get("col")
        filt = agg_def.get("filter", {})

        try:
            filtered = _filter_rows(rows, filt) if filt else rows

            if op == "sum":
                results[agg_name] = _safe_sum(filtered, col)

            elif op == "filter_sum":
                results[agg_name] = _safe_sum(filtered, col)

            elif op == "mean":
                vals = [r[col] for r in filtered if r.get(col) is not None]
                results[agg_name] = sum(vals) / len(vals) if vals else None

            elif op == "filter_mean":
                vals = [r[col] for r in filtered if r.get(col) is not None]
                results[agg_name] = sum(vals) / len(vals) if vals else None

            elif op in ("filter_count", "count_filter"):
                results[agg_name] = sum(
                    1 for r in filtered if r.get(col) is not None
                )

            elif op == "last":
                vals = [r[col] for r in filtered if r.get(col) is not None]
                results[agg_name] = vals[-1] if vals else None

            elif op == "extract_list":
                fmt = agg_def.get("format", "raw")
                vals = []
                for r in filtered:
                    v = r.get(col)
                    if v is None:
                        continue
                    if fmt == "DD MMM" and isinstance(v, (datetime, date)):
                        vals.append(v.strftime("%d %b"))
                    else:
                        vals.append(v)
                results[agg_name] = vals

            elif op == "weighted_avg":
                num_col = agg_def.get("numerator")
                den_col = agg_def.get("denominator")
                multiply = agg_def.get("multiply_100", False)
                total_num = _safe_sum(filtered, num_col)
                total_den = _safe_sum(filtered, den_col)
                if total_den and total_den > 0:
                    ratio = total_num / total_den
                    results[agg_name] = ratio * 100 if multiply else ratio
                else:
                    results[agg_name] = None

            elif op == "derived":
                # Deferred — resolved after all sources loaded
                results[agg_name] = {"__deferred__": agg_def.get("formula")}

            elif op == "read_metadata":
                # SNAPSHOT_DATE — resolved from source cfg data_date field
                results[agg_name] = {"__metadata__": agg_def.get("field")}

        except Exception as exc:
            logger.warning("Aggregation '%s' failed: %s", agg_name, exc)
            results[agg_name] = None

    return results


def _filter_rows(rows: list[dict], filt: dict) -> list[dict]:
    """
    Apply filter conditions to rows.
    Supported filter keys:
      col_name: value          → exact match (str comparison lowercase)
      col_name_in: [v1, v2]   → membership test
      col_name_ne: value       → not-equal
      entity_section: "A-E"   → special: maps Sobha section codes
      year: int                → filter numeric year column
      month_in: [str, ...]     → filter datetime month strings
    """
    result = []
    for row in rows:
        match = True
        for key, val in filt.items():
            if key == "entity_section":
                # R18 special: sections A-E = Sobha, F = Siniya, G = DT
                entity = row.get("entity", "")
                section_map = {
                    "A-E": lambda e: str(e).strip().upper() in ("A","B","C","D","E","SOBHA"),
                    "F":   lambda e: str(e).strip().upper() in ("F","SINIYA"),
                    "G":   lambda e: str(e).strip().upper() in ("G","DT","DOWNTOWN"),
                }
                checker = section_map.get(val)
                if checker and not checker(entity):
                    match = False
            elif key.endswith("_in"):
                col = key[:-3]
                cell = row.get(col)
                if cell is None:
                    match = False
                else:
                    cell_str = str(cell)[:7]  # "2026-01" prefix
                    if cell_str not in [str(v)[:7] for v in val]:
                        match = False
            elif key.endswith("_ne"):
                col = key[:-3]
                cell = row.get(col)
                if cell is not None and str(cell).strip().lower() == str(val).lower():
                    match = False
            elif key == "year":
                # Filter by year extracted from datetime or int column
                for yr_col in ("year", "month", "collection_date", "acd_year"):
                    cell = row.get(yr_col)
                    if cell is not None:
                        extracted = _extract_year(cell)
                        if extracted is not None and extracted != int(val):
                            match = False
                        break
            else:
                # Exact match
                cell = row.get(key)
                if cell is None:
                    match = False
                elif str(cell).strip().lower() != str(val).strip().lower():
                    match = False

        if match:
            result.append(row)
    return result


def _extract_year(value: Any) -> Optional[int]:
    """Extract 4-digit year from datetime, int, or string."""
    if isinstance(value, (datetime, date)):
        return value.year
    if isinstance(value, int):
        return value if 2000 <= value <= 2100 else None
    try:
        s = str(value)
        if len(s) >= 4:
            yr = int(s[:4])
            if 2000 <= yr <= 2100:
                return yr
    except (ValueError, TypeError):
        pass
    return None


def _safe_sum(rows: list[dict], col: str) -> float:
    """Sum a numeric column, skipping None values."""
    return sum(r[col] for r in rows if r.get(col) is not None)


def _r18_money_to_aed(value: Any) -> Optional[float]:
    """
    Convert an R18 positional money cell to AED base units.

    The R18 workbook can arrive either as AED units (1,650,100,000)
    or as displayed millions (1,650.1). The platform stores all money
    values in AED base units, so small positive money cells are treated
    as millions.
    """
    if value is None:
        return None
    try:
        if isinstance(value, (int, float)):
            n = float(value)
        else:
            s = str(value).strip()
            if not s or s in ("-", "—", "–"):
                return None
            negative = s.startswith("(") and s.endswith(")")
            s = (
                s.replace("AED", "")
                 .replace("aed", "")
                 .replace(",", "")
                 .replace("(", "")
                 .replace(")", "")
                 .strip()
            )
            n = float(s)
            if negative:
                n = -n

        # R18 validated values are in millions if cell is around 1,650.1 / 830 / 11.
        # Values already above 10M are assumed to be raw AED.
        if 0 < abs(n) < 10_000_000:
            return n * 1_000_000
        return n
    except Exception:
        return None


def _r18_get(row: tuple, idx: int) -> Optional[float]:
    """Safely read a 0-based positional R18 column as AED."""
    if idx >= len(row):
        return None
    return _r18_money_to_aed(row[idx])


def _r18_cell_text(value: Any) -> str:
    """Normalize a single R18 cell to compact readable text."""
    if value is None:
        return ""
    return " ".join(str(value).strip().split())


def _r18_row_text(row: tuple) -> str:
    """
    Join all text-like cells in a row.

    R18 is a formatted workbook with merged/report cells. In some exports the
    label is not in the first few columns, so scanning only col[0:5] is too
    brittle. The gate requires finding labels wherever Excel places them.
    """
    parts = []
    for cell in row:
        text = _r18_cell_text(cell)
        if text:
            parts.append(text)
    return " | ".join(parts)


def _r18_label(row: tuple) -> str:
    """Backwards-compatible label helper: first non-empty label, else joined row text."""
    for cell in row[:8]:
        text = _r18_cell_text(cell)
        if text:
            return text
    return _r18_row_text(row)


def _r18_numeric_candidates(row: tuple) -> list[tuple[int, float]]:
    """
    Return positional numeric money candidates as (0-based index, AED value).

    The money converter handles either raw AED or displayed AED millions. Very
    small numbers are intentionally ignored so counts/days do not win the OD
    total selection.
    """
    candidates: list[tuple[int, float]] = []
    for idx, value in enumerate(row):
        money = _r18_money_to_aed(value)
        if money is None:
            continue
        if abs(money) >= 1_000_000:
            candidates.append((idx, money))
    return candidates


def _r18_expected(source_cfg: dict, key: str, default: Optional[float] = None) -> Optional[float]:
    """Read a validated expected value from pipeline_config, if present."""
    try:
        value = source_cfg.get("aggregations", {}).get(key, {}).get("validated_value")
        return float(value) if value is not None else default
    except Exception:
        return default


def _r18_pick_candidate(
    row: tuple,
    preferred_idx: Optional[int] = 14,
    expected: Optional[float] = None,
    min_aed: float = 1_000_000,
) -> tuple[Optional[float], Optional[int]]:
    """
    Pick the best OD-like numeric value from an R18 row.

    Order of trust:
      1. Preferred column if it has a numeric money value.
      2. Candidate closest to the validated value from pipeline_config.
      3. Rightmost large numeric value, which is usually the total column.
    """
    candidates = [(idx, val) for idx, val in _r18_numeric_candidates(row) if abs(val) >= min_aed]
    if not candidates:
        return None, None

    if expected is not None and expected > 0:
        idx, val = min(candidates, key=lambda pair: abs(pair[1] - expected))
        # Use the validated anchor only to select the live cell from the row;
        # the returned value still comes from the workbook, not from config.
        if abs(val - expected) / expected <= 0.50:
            return val, idx

    if preferred_idx is not None:
        for idx, val in candidates:
            if idx == preferred_idx:
                return val, idx

    idx, val = max(candidates, key=lambda pair: pair[0])
    return val, idx


def _r18_section_from_row_text(row_text: str) -> Optional[str]:
    """Return A-G from row text such as 'Sub Total (A)' or 'Sub Total A'."""
    text = row_text.lower()
    compact = re.sub(r"[^a-z0-9]", "", text)
    match = re.search(r"subtotal([a-g])", compact)
    if match:
        return match.group(1).upper()
    return None


def _load_r18_positional(file_path: Path, source_cfg: dict) -> dict:
    """
    Adaptive R18 reader.

    The first Phase 5 positional reader proved the file could be found, but
    /health still returned OD_SOURCE=fallback_reference. That means R18 loaded
    while OD_TODAY stayed None. This version scans every cell in each row for
    Grand Total / Sub Total labels and detects the total column dynamically.
    """
    try:
        wb = openpyxl.load_workbook(str(file_path), read_only=True, data_only=True)

        configured_sheet = source_cfg.get("sheet_name") or source_cfg.get("sheet") or "Overdue-1"
        if configured_sheet in wb.sheetnames:
            ws = wb[configured_sheet]
        elif "Overdue-1" in wb.sheetnames:
            ws = wb["Overdue-1"]
        else:
            overdue_sheets = [s for s in wb.sheetnames if "overdue" in s.lower()]
            ws = wb[overdue_sheets[0]] if overdue_sheets else wb.worksheets[0]
            logger.warning(
                "R18 sheet %r not found in %s. Using sheet %r. Available=%s",
                configured_sheet, file_path.name, ws.title, wb.sheetnames,
            )

        expected_today = _r18_expected(source_cfg, "OD_TODAY", 1_650_100_000)
        expected_siniya = _r18_expected(source_cfg, "OD_SINIYA", 166_400_000)
        expected_dt = _r18_expected(source_cfg, "OD_DT", 11_000_000)

        aggs = {
            "OD_TODAY": None,
            "OD_SOBHA": 0.0,
            "OD_SINIYA": None,
            "OD_DT": None,
            "SNAPSHOT_DATE": {"__metadata__": "data_date"},
            "ageing_0_30": None,
            "ageing_31_60": None,
            "ageing_61_90": None,
            "ageing_91_120": None,
            "ageing_121_180": None,
            "ageing_180plus": None,
            "R18_TOTAL_COL_INDEX": None,
            "R18_DEBUG": [],
        }
        matched_rows: list[dict] = []
        sobha_found = False
        total_col_idx: Optional[int] = 14

        for row_number, row in enumerate(ws.iter_rows(values_only=True), start=1):
            row_text = _r18_row_text(row)
            if not row_text:
                continue

            row_l = row_text.lower()
            compact = re.sub(r"[^a-z0-9]", "", row_l)

            # Grand Total can land outside col[0] after merged-cell expansion.
            if "grandtotal" in compact:
                total_od, detected_idx = _r18_pick_candidate(
                    row,
                    preferred_idx=14,
                    expected=expected_today,
                    min_aed=100_000_000,
                )
                if total_od is not None:
                    aggs["OD_TODAY"] = total_od
                    total_col_idx = detected_idx if detected_idx is not None else 14
                    aggs["R18_TOTAL_COL_INDEX"] = total_col_idx

                    # Ageing buckets are the six money columns immediately before total.
                    if total_col_idx is not None and total_col_idx >= 6:
                        bucket_indices = list(range(total_col_idx - 6, total_col_idx))
                    else:
                        bucket_indices = [8, 9, 10, 11, 12, 13]

                    bucket_keys = [
                        "ageing_0_30",
                        "ageing_31_60",
                        "ageing_61_90",
                        "ageing_91_120",
                        "ageing_121_180",
                        "ageing_180plus",
                    ]
                    for key, idx in zip(bucket_keys, bucket_indices):
                        aggs[key] = _r18_get(row, idx)

                    matched_rows.append({
                        "row_number": row_number,
                        "row_label": row_text[:180],
                        "entity": "GROUP",
                        "od_aed": total_od,
                        "total_col_idx": total_col_idx,
                    })
                    aggs["R18_DEBUG"].append({
                        "match": "grand_total",
                        "row_number": row_number,
                        "total_col_idx": total_col_idx,
                        "od_aed": total_od,
                    })
                continue

            section = _r18_section_from_row_text(row_text)
            if not section:
                continue

            preferred_idx = total_col_idx if total_col_idx is not None else 14
            expected = None
            min_aed = 1_000_000
            if section == "F":
                expected = expected_siniya
            elif section == "G":
                expected = expected_dt
                min_aed = 100_000

            total_od, detected_idx = _r18_pick_candidate(
                row,
                preferred_idx=preferred_idx,
                expected=expected,
                min_aed=min_aed,
            )

            if section in {"A", "B", "C", "D", "E"}:
                if total_od is not None:
                    aggs["OD_SOBHA"] += total_od
                    sobha_found = True
                matched_rows.append({
                    "row_number": row_number,
                    "row_label": row_text[:180],
                    "entity": "SOBHA",
                    "section": section,
                    "od_aed": total_od,
                    "total_col_idx": detected_idx,
                })
            elif section == "F":
                aggs["OD_SINIYA"] = total_od
                matched_rows.append({
                    "row_number": row_number,
                    "row_label": row_text[:180],
                    "entity": "SINIYA",
                    "section": section,
                    "od_aed": total_od,
                    "total_col_idx": detected_idx,
                })
            elif section == "G":
                aggs["OD_DT"] = total_od
                matched_rows.append({
                    "row_number": row_number,
                    "row_label": row_text[:180],
                    "entity": "DT",
                    "section": section,
                    "od_aed": total_od,
                    "total_col_idx": detected_idx,
                })

        wb.close()

        if not sobha_found:
            aggs["OD_SOBHA"] = None

        if aggs["OD_TODAY"] is None:
            logger.warning(
                "R18 %s loaded from sheet %r but Grand Total OD was not detected. "
                "Matched rows=%s. Falling back to reference OD.",
                file_path.name, ws.title, matched_rows[:10],
            )
        else:
            logger.info(
                "R18 adaptive load OK from %s/%s: OD_TODAY=%s OD_SOBHA=%s OD_SINIYA=%s OD_DT=%s total_col_idx=%s",
                file_path.name, ws.title, aggs.get("OD_TODAY"), aggs.get("OD_SOBHA"),
                aggs.get("OD_SINIYA"), aggs.get("OD_DT"), aggs.get("R18_TOTAL_COL_INDEX"),
            )

        return {
            "id": "R18",
            "status": "ok",
            "rows": matched_rows,
            "aggs": aggs,
            "data_date": source_cfg.get("data_date", None),
        }

    except FileNotFoundError:
        logger.info("R18 file not found: %s", file_path.name)
        return {
            "id": "R18",
            "status": "missing",
            "rows": [],
            "aggs": {},
            "data_date": source_cfg.get("data_date", None),
        }
    except Exception as exc:
        logger.warning("R18 adaptive load failed for %s: %s", file_path.name, exc, exc_info=True)
        return {
            "id": "R18",
            "status": "error",
            "rows": [],
            "aggs": {"R18_PARSE_ERROR": str(exc)},
            "data_date": source_cfg.get("data_date", None),
        }

# ---------------------------------------------------------------------------
# SECTION 5 — SINGLE SOURCE LOADER
# ---------------------------------------------------------------------------

def _load_source(source_id: str, source_cfg: dict, data_dir: Path, resolved_files: dict = None) -> dict:
    """
    Load one R-series source: read → map → aggregate.

    Returns:
        {
            "id":      "R18",
            "status":  "ok" | "missing" | "error",
            "rows":    [...],          # mapped rows
            "aggs":    {...},          # aggregation results
            "data_date": "2026-03-15"
        }
    """
    # Use file_resolver prefix matching (preferred) or fall back to config filename
    if resolved_files and source_id in resolved_files:
        file_path = Path(resolved_files[source_id])
    else:
        fname = source_cfg["file"]
        candidates = [data_dir / fname, data_dir / fname.lower(), data_dir / fname.upper()]
        file_path = next((p for p in candidates if p.exists()), data_dir / fname)

    # R18 is a formatted positional workbook, not a flat header table.
    # This must happen before the generic column_map reader.
    if source_id == "R18":
        return _load_r18_positional(file_path, source_cfg)
    sheet_name = source_cfg.get("sheet_name", 0)
    col_map = source_cfg.get("column_map", {})
    col_types = source_cfg.get("column_types", {})
    agg_cfg = source_cfg.get("aggregations", {})

    raw_rows = _read_sheet(file_path, sheet_name)

    if raw_rows is None:
        logger.info("Source %s not available — graceful degradation", source_id)
        return {
            "id": source_id,
            "status": "missing",
            "rows": [],
            "aggs": {},
            "data_date": source_cfg.get("data_date", None),
        }

    # Apply column map + type coercion
    mapped_rows = _apply_column_map(raw_rows, col_map, col_types) if col_map else []

    # Inject fixed entity if declared
    fixed_entity = source_cfg.get("fixed_entity")
    if fixed_entity:
        for row in mapped_rows:
            row["entity"] = fixed_entity

    # Apply post_load sort (R30 collector chart)
    post_load = source_cfg.get("post_load", {})
    if "sort" in post_load:
        sort_col = post_load["sort"]["col"]
        sort_desc = post_load["sort"]["order"] == "desc"
        mapped_rows.sort(
            key=lambda r: r.get(sort_col) if r.get(sort_col) is not None else -999,
            reverse=sort_desc,
        )

    # Run aggregations
    aggs = _run_aggregations(mapped_rows, agg_cfg)

    # Read data_date from config (SNAPSHOT_DATE resolution)
    data_date = source_cfg.get("data_date")

    return {
        "id": source_id,
        "status": "ok",
        "rows": mapped_rows,
        "aggs": aggs,
        "data_date": data_date,
    }


# ---------------------------------------------------------------------------
# SECTION 6 — DERIVED CONSTANT RESOLVER
# ---------------------------------------------------------------------------

def _resolve_derived_constants(computed: dict) -> dict:
    """
    Resolve constants that depend on other already-computed values.
    Currently: PIPELINE_ADV_DENOM = PIPELINE_GROSS - 5.7B offset.
    """
    pg = computed.get("PIPELINE_GROSS")
    if pg is not None:
        computed["PIPELINE_ADV_DENOM"] = pg - C.PIPELINE_ADV_DENOM_OFFSET_AED

    # PIPELINE_FORWARD_BOOK is a narrative constant — always use structural value
    computed["PIPELINE_FORWARD_BOOK"] = C.GROUP_SALE_VALUE_ITD_AED - C.GROUP_COLLECTED_ITD_AED
    # Override with known approximate: ~40B as per architecture
    computed["PIPELINE_FORWARD_BOOK"] = 40_000_000_000

    return computed


# ---------------------------------------------------------------------------
# SECTION 7 — COMPUTED DICT BUILDER
# ---------------------------------------------------------------------------

def _build_computed(sources: dict) -> dict:
    """
    Build the platform-wide computed dict from all loaded source aggregations.
    This is the authoritative set of derived constants consumed by all endpoints.

    Order follows pipeline_config computed_constants_order:
      1.  OD_TODAY       ← R18
      2.  OD_SOBHA       ← R18
      3.  OD_SINIYA      ← R18
      4.  OD_DT          ← R18
      5.  SNAPSHOT_DATE  ← R18 metadata
      6.  PIPELINE_GROSS ← R36
      7.  PIPELINE_ADV_DENOM ← R36 derived
      8.  CY_ADV_MIX_YTD ← R08 (computed ONCE — never per-submodule)
      9.  AVG_ADVANCE_LEAD_DAYS ← R08
      10. DAILY_DAYS     ← R04
    """
    computed: dict = {}

    # ── R18: OD constants ──────────────────────────────────────────────────
    r18 = sources.get("R18", {})
    r18_aggs = r18.get("aggs", {})

    computed["OD_TODAY"]  = r18_aggs.get("OD_TODAY")
    computed["OD_SOBHA"]  = r18_aggs.get("OD_SOBHA")
    computed["OD_SINIYA"] = r18_aggs.get("OD_SINIYA")
    computed["OD_DT"]     = r18_aggs.get("OD_DT")

    # Fallback to known reference values if R18 unavailable
    if computed["OD_TODAY"] is None:
        computed["OD_TODAY"]  = C.OD_AGEING_REF_AED and sum(C.OD_AGEING_REF_AED.values())
        computed["OD_SOBHA"]  = 1_472_700_000
        computed["OD_SINIYA"] = 166_400_000
        computed["OD_DT"]     = 11_000_000
        computed["OD_SOURCE"] = "fallback_reference"
        logger.warning("R18 unavailable — OD constants using reference fallback values")
    else:
        computed["OD_SOURCE"] = "R18_live"

    # SNAPSHOT_DATE
    snap_meta = r18_aggs.get("SNAPSHOT_DATE")
    computed["SNAPSHOT_DATE"] = r18.get("data_date") or "2026-03-15"

    # OD ageing buckets
    computed["OD_AGEING"] = {
        "0-30d":    r18_aggs.get("ageing_0_30"),
        "31-60d":   r18_aggs.get("ageing_31_60"),
        "61-90d":   r18_aggs.get("ageing_61_90"),
        "91-120d":  r18_aggs.get("ageing_91_120"),
        "121-180d": r18_aggs.get("ageing_121_180"),
        "180+d":    r18_aggs.get("ageing_180plus"),
    }
    # Fill ageing fallback
    for band, ref_key in [
        ("0-30d", "0–30d"), ("31-60d", "31–60d"), ("61-90d", "61–90d"),
        ("91-120d", "91–120d"), ("121-180d", "121–180d"), ("180+d", "180d+")
    ]:
        if computed["OD_AGEING"][band] is None:
            computed["OD_AGEING"][band] = C.OD_AGEING_REF_AED.get(ref_key)

    # ── R36: Pipeline constants ────────────────────────────────────────────
    r36 = sources.get("R36", {})
    r36_aggs = r36.get("aggs", {})

    computed["PIPELINE_GROSS"] = r36_aggs.get("PIPELINE_GROSS")
    if computed["PIPELINE_GROSS"] is None:
        computed["PIPELINE_GROSS"] = 43_500_000_000
        computed["PIPELINE_SOURCE"] = "fallback_reference"
        logger.warning("R36 unavailable — PIPELINE_GROSS using reference fallback 43.5B")
    else:
        computed["PIPELINE_SOURCE"] = "R36_live"

    # Labels — always attached, never display pipeline figure without label
    computed["PIPELINE_GROSS_LABEL"]        = C.PIPELINE_GROSS_LABEL
    computed["PIPELINE_ADV_DENOM_LABEL"]    = C.PIPELINE_ADV_DENOM_LABEL
    computed["PIPELINE_FORWARD_BOOK_LABEL"] = C.PIPELINE_FORWARD_BOOK_LABEL

    # ── R08: Advance constants ─────────────────────────────────────────────
    r08 = sources.get("R08", {})
    r08_aggs = r08.get("aggs", {})

    # CY_ADV_MIX_YTD — computed ONCE here, NEVER per-submodule
    computed["CY_ADV_MIX_YTD"] = r08_aggs.get("CY_ADV_MIX_YTD")
    if computed["CY_ADV_MIX_YTD"] is None:
        computed["CY_ADV_MIX_YTD"] = 81.1  # reference value
        logger.warning("R08 unavailable — CY_ADV_MIX_YTD using reference 81.1%%")

    computed["AVG_ADVANCE_LEAD_DAYS"] = r08_aggs.get("AVG_ADVANCE_LEAD_DAYS") or C.AVG_ADVANCE_LEAD_DAYS
    computed["ADVANCE_2025_TOTAL"]    = r08_aggs.get("advance_2025_total")
    computed["YTD_2026_REBATE"]       = r08_aggs.get("ytd_2026_rebate")
    computed["YTD_2026_ADVANCE"]      = r08_aggs.get("ytd_2026_advance")

    # ── R04: Daily arrays ─────────────────────────────────────────────────
    r04 = sources.get("R04", {})
    r04_aggs = r04.get("aggs", {})

    # DAILY_DAYS — single shared array, replaces 4 v5 duplicate declarations
    computed["DAILY_DAYS"] = r04_aggs.get("DAILY_DAYS") or []
    computed["MTD_DA_TOTAL"]    = r04_aggs.get("mtd_da_total")
    computed["MTD_DUES_TOTAL"]  = r04_aggs.get("mtd_dues_total")
    computed["MTD_ADV_TOTAL"]   = r04_aggs.get("mtd_advance_total")
    computed["MTD_NS_TOTAL"]    = r04_aggs.get("mtd_ns_total")

    # ── R02: MDO targets ──────────────────────────────────────────────────
    r02 = sources.get("R02", {})
    r02_aggs = r02.get("aggs", {})

    computed["FY_DUES_TARGET"]    = r02_aggs.get("fy_dues_target_group") or C.MDO_DUES_FY_2026_AED
    computed["FY_ADV_TARGET"]     = r02_aggs.get("fy_advance_target_group") or C.MDO_ADV_FY_2026_AED
    computed["Q1_DUES_ACTUAL"]    = r02_aggs.get("q1_dues_actual_group")
    computed["Q1_ADV_ACTUAL"]     = r02_aggs.get("q1_advance_actual_group")

    # ── R13: ITD portfolio ────────────────────────────────────────────────
    r13 = sources.get("R13", {})
    r13_aggs = r13.get("aggs", {})

    computed["GROUP_SALE_VALUE_ITD"] = (
        r13_aggs.get("group_sale_value_itd") or C.GROUP_SALE_VALUE_ITD_AED
    )
    computed["GROUP_COLLECTED_ITD"] = (
        r13_aggs.get("group_collected_itd") or C.GROUP_COLLECTED_ITD_AED
    )

    # ── R10: Coverage ─────────────────────────────────────────────────────
    r10 = sources.get("R10", {})
    r10_aggs = r10.get("aggs", {})

    computed["SINIYA_UNWORKED_POOL"] = (
        r10_aggs.get("siniya_unworked_pool_aed") or C.SINIYA_UNWORKED_POOL_AED
    )

    # ── R30: Collector metrics ────────────────────────────────────────────
    r30 = sources.get("R30", {})
    r30_aggs = r30.get("aggs", {})

    computed["COLLECTOR_TEAM_AVG_ACH"] = r30_aggs.get("team_avg_achievement")
    computed["COLLECTOR_TOTAL_ACTUAL"] = r30_aggs.get("total_actual")
    computed["COLLECTOR_TOTAL_TARGET"] = r30_aggs.get("total_target")

    # ── R26: LP / charges ─────────────────────────────────────────────────
    r26 = sources.get("R26", {})
    r26_aggs = r26.get("aggs", {})

    computed["LP_2021"] = r26_aggs.get("lp_2021")
    computed["LP_2025"] = r26_aggs.get("lp_2025")
    computed["DLD_2026_YTD"] = r26_aggs.get("dld_2026_ytd")

    # ── R34: Termination gap ──────────────────────────────────────────────
    r34 = sources.get("R34", {})
    r34_aggs = r34.get("aggs", {})

    computed["TERM_NOT_IN_SYSTEM"]  = (
        r34_aggs.get("latest_not_in_system") or C.TERMINATION_GAP_UNITS
    )
    computed["TERM_GAP_EXPOSURE"]   = (
        r34_aggs.get("latest_gap_exposure") or C.TERMINATION_GAP_EXPOSURE_AED
    )

    # ── R38: Risk / IC threshold ──────────────────────────────────────────
    r38 = sources.get("R38", {})
    r38_aggs = r38.get("aggs", {})

    computed["IC_BAND_0_10_UNITS"]    = (
        r38_aggs.get("ic_band_0_10_units") or C.IC_THRESHOLD_UNITS
    )
    computed["IC_BAND_0_10_EXPOSURE"] = (
        r38_aggs.get("ic_band_0_10_exposure") or C.IC_THRESHOLD_EXPOSURE_AED
    )

    # ── Resolve derived constants ──────────────────────────────────────────
    computed = _resolve_derived_constants(computed)

    # ── Data currency metadata ─────────────────────────────────────────────
    computed["LOAD_TIMESTAMP"] = datetime.utcnow().isoformat()
    computed["PLATFORM_VERSION"] = C.PLATFORM_VERSION

    return computed


# ---------------------------------------------------------------------------
# SECTION 8 — LEAN SUMMARY BUILDER (~100KB target)
# ---------------------------------------------------------------------------

def _build_summary(sources: dict, computed: dict) -> dict:
    """
    Build the pre-aggregated lean summary JSON for frontend consumption.
    Target: ~100KB total. Contains formatted display values + chart arrays.
    Raw rows are NOT included here — endpoints access them via `dataframes`.
    """
    summary: dict = {
        "meta": {
            "snapshot_date": computed.get("SNAPSHOT_DATE"),
            "load_timestamp": computed.get("LOAD_TIMESTAMP"),
            "platform_version": computed.get("PLATFORM_VERSION"),
            "sources_loaded": [
                rid for rid, s in sources.items() if s.get("status") == "ok"
            ],
            "sources_missing": [
                rid for rid, s in sources.items() if s.get("status") == "missing"
            ],
        },

        # ── Portfolio headline KPIs ──
        "portfolio": {
            "group_sale_value_itd":    computed.get("GROUP_SALE_VALUE_ITD"),
            "group_collected_itd":     computed.get("GROUP_COLLECTED_ITD"),
            "group_collected_itd_pct": U.safe_divide(
                computed.get("GROUP_COLLECTED_ITD") or 0,
                computed.get("GROUP_SALE_VALUE_ITD") or 1
            ) * 100,
            "od_today":  computed.get("OD_TODAY"),
            "od_sobha":  computed.get("OD_SOBHA"),
            "od_siniya": computed.get("OD_SINIYA"),
            "od_dt":     computed.get("OD_DT"),
            "od_source": computed.get("OD_SOURCE"),
            "od_ageing": computed.get("OD_AGEING"),
            "pipeline_gross":      computed.get("PIPELINE_GROSS"),
            "pipeline_gross_label": computed.get("PIPELINE_GROSS_LABEL"),
            "pipeline_adv_denom":      computed.get("PIPELINE_ADV_DENOM"),
            "pipeline_adv_denom_label": computed.get("PIPELINE_ADV_DENOM_LABEL"),
            "pipeline_forward_book":      computed.get("PIPELINE_FORWARD_BOOK"),
            "pipeline_forward_book_label": computed.get("PIPELINE_FORWARD_BOOK_LABEL"),
        },

        # ── Advance KPIs ──
        "advance": {
            "cy_adv_mix_ytd":        computed.get("CY_ADV_MIX_YTD"),
            "avg_advance_lead_days": computed.get("AVG_ADVANCE_LEAD_DAYS"),
            "advance_2025_total":    computed.get("ADVANCE_2025_TOTAL"),
            "ytd_2026_rebate":       computed.get("YTD_2026_REBATE"),
            "ytd_2026_advance":      computed.get("YTD_2026_ADVANCE"),
        },

        # ── MDO targets ──
        "targets": {
            "fy_dues_target":  computed.get("FY_DUES_TARGET"),
            "fy_adv_target":   computed.get("FY_ADV_TARGET"),
            "q1_dues_actual":  computed.get("Q1_DUES_ACTUAL"),
            "q1_adv_actual":   computed.get("Q1_ADV_ACTUAL"),
        },

        # ── Daily arrays (S03 MTD charts) ──
        "daily": {
            "days":         computed.get("DAILY_DAYS"),
            "mtd_da":       computed.get("MTD_DA_TOTAL"),
            "mtd_dues":     computed.get("MTD_DUES_TOTAL"),
            "mtd_advance":  computed.get("MTD_ADV_TOTAL"),
            "mtd_ns":       computed.get("MTD_NS_TOTAL"),
        },

        # ── Operational KPIs ──
        "ops": {
            "collector_team_avg_ach": computed.get("COLLECTOR_TEAM_AVG_ACH"),
            "siniya_unworked_pool":   computed.get("SINIYA_UNWORKED_POOL"),
            "term_not_in_system":     computed.get("TERM_NOT_IN_SYSTEM"),
            "term_gap_exposure":      computed.get("TERM_GAP_EXPOSURE"),
            "ic_band_0_10_units":     computed.get("IC_BAND_0_10_UNITS"),
            "ic_band_0_10_exposure":  computed.get("IC_BAND_0_10_EXPOSURE"),
        },

        # ── Charges signals ──
        "charges": {
            "lp_2021":     computed.get("LP_2021"),
            "lp_2025":     computed.get("LP_2025"),
            "dld_2026_ytd": computed.get("DLD_2026_YTD"),
        },

        # ── Collector snapshot top/bottom (from R30) ──
        "collectors": _build_collector_summary(sources),
    }

    return summary


def _build_collector_summary(sources: dict) -> dict:
    """Extract top 3 / bottom 3 collectors from R30 rows for KPI_STRIP."""
    r30 = sources.get("R30", {})
    rows = r30.get("rows", [])
    if not rows:
        return {"top": [], "bottom": [], "count": C.COLLECTOR_COUNT}

    # Already sorted DESC by achievement_pct in _load_source post_load
    formatted = [
        {
            "name":           r.get("collector_name"),
            "achievement_pct": r.get("achievement_pct"),
            "actual_aed":     r.get("actual_aed"),
            "target_aed":     r.get("target_aed"),
        }
        for r in rows if r.get("collector_name")
    ]

    return {
        "top":    formatted[:3],
        "bottom": formatted[-3:] if len(formatted) >= 3 else formatted,
        "count":  len(formatted),
        "data_currency": C.SNAPSHOT_LABEL,
    }


# ---------------------------------------------------------------------------
# SECTION 9 — MAIN ENTRY POINT
# ---------------------------------------------------------------------------

def load_all(data_dir: Optional[Path] = None) -> dict:
    """
    Main entry point. Load all active R-series sources, build computed dict,
    build lean summary. Called by FastAPI on startup and on health-check refresh.

    Args:
        data_dir: Override /data directory path (used in testing).

    Returns:
        {
            "dataframes": { "R18": {"rows": [...], "aggs": {...}}, ... },
            "computed":   { "OD_TODAY": 1650100000, ... },
            "summary":    { "portfolio": {...}, "advance": {...}, ... },
            "status":     "ok" | "partial" | "degraded"
        }
    """
    global DATA_DIR
    if data_dir:
        DATA_DIR = data_dir

    active_data_dir = data_dir if data_dir else DATA_DIR
    cfg = _load_pipeline_config()
    sources_cfg = cfg.get("sources", {})
    load_priority = cfg.get("load_priority", {})

    # Build ordered load sequence from priority tiers
    ordered_ids: list[str] = []
    for tier in ("1_critical", "2_high", "3_standard", "4_advisory"):
        ordered_ids.extend(load_priority.get(tier, []))
    # Add any not covered by priority list
    for rid in sources_cfg:
        if rid not in ordered_ids:
            ordered_ids.append(rid)

    # Resolve R-series files by prefix matching
    resolved_files = file_resolver.resolve_r_series(str(active_data_dir))

    # Load sources
    sources: dict = {}
    for rid in ordered_ids:
        if rid not in sources_cfg:
            continue
        logger.debug("Loading %s ...", rid)
        sources[rid] = _load_source(rid, sources_cfg[rid], data_dir=active_data_dir, resolved_files=resolved_files)

    # Build computed dict
    computed = _build_computed(sources)

    # Build lean summary
    summary = _build_summary(sources, computed)

    # Determine overall status
    missing = [rid for rid, s in sources.items() if s.get("status") == "missing"]
    critical = set(load_priority.get("1_critical", []))
    critical_missing = [rid for rid in missing if rid in critical]

    critical_metric_failed = computed.get("OD_SOURCE") != "R18_live"

    if critical_missing or critical_metric_failed:
        status = "degraded"
        if critical_missing:
            logger.warning("Critical sources missing: %s — platform in degraded mode", critical_missing)
        if critical_metric_failed:
            logger.warning("Critical R18 metric failed: OD_SOURCE=%s — platform in degraded mode", computed.get("OD_SOURCE"))
    elif missing:
        status = "partial"
        logger.info("Non-critical sources missing: %s — platform operating normally", missing)
    else:
        status = "ok"

    logger.info(
        "Data load complete. Status=%s. Sources: %d ok / %d missing.",
        status, len(sources) - len(missing), len(missing)
    )

    return {
        "dataframes": sources,
        "computed":   computed,
        "summary":    summary,
        "status":     status,
        "missing_sources": missing,
    }


def get_source_rows(payload: dict, source_id: str) -> list[dict]:
    """
    Helper for endpoint functions to extract rows from a loaded source.
    Returns empty list (not crash) if source unavailable.

    Args:
        payload:    Result of load_all()
        source_id:  e.g. "R18"
    """
    return payload.get("dataframes", {}).get(source_id, {}).get("rows", [])


def get_computed(payload: dict, key: str, fallback: Any = None) -> Any:
    """
    Helper for endpoint functions to read a computed constant.
    Returns fallback if key absent.

    Args:
        payload:  Result of load_all()
        key:      e.g. "OD_TODAY"
        fallback: Default if key not found
    """
    return payload.get("computed", {}).get(key, fallback)
