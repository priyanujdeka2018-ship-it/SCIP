"""
SOBHA COLLECTIONS INTELLIGENCE PLATFORM
data_loader.py — R-Series Data Pipeline
Version: v8.3 Batch 3
Authority: FINAL_ARCHITECTURE.md

Responsibilities:
  - Read all active R-series xlsx files from /data directory
  - Apply column mappings from pipeline_config.json
  - Compute all derived constants (OD_TODAY, PIPELINE_GROSS, etc.)
  - Return { "dataframes": dict, "computed": dict }
  - Pre-aggregate lean summary JSON for frontend (~100KB)
  - Graceful degradation on missing or malformed files

Dependency rule:
  data_loader.py  →  imports: constants, utils, pipeline_config.json
  data_loader.py  →  NEVER imports: endpoints, quickball, main
  data_loader.py  →  is stateless. No session state. No globals mutated.

Architecture rules (never violate):
  - OD_TODAY comes from R18. Never hardcoded.
  - PIPELINE_GROSS, PIPELINE_ADV_DENOM come from R36.
  - CY_ADV_MIX_YTD computed ONCE here. Never per-submodule.
  - DAILY_DAYS[] extracted from R04 ONCE. Never per-section.
  - Collector data sorted DESC by achievement_pct on load.
  - Missing file → data_pending payload, not crash.
  - All monetary values stored as AED base units (not millions).
"""

from __future__ import annotations

import copy
import json
import logging
import os
import threading
import time
from datetime import datetime, date
from pathlib import Path
from typing import Any, Optional

import openpyxl
try:  # package import path
    from .excel_reader import read_sheet as _read_sheet
    from .column_mapper import apply_column_map as _apply_column_map, coerce as _coerce
except ImportError:  # top-level runtime import path used by uvicorn main:app
    from excel_reader import read_sheet as _read_sheet
    from column_mapper import apply_column_map as _apply_column_map, coerce as _coerce

import constants as C
import utils as U
import file_resolver
import source_adapters

# Import aggregation helpers from the standalone module.  These aliases
# ensure that existing calls within this module continue to work while
# the underlying logic has been extracted for better modularity.
try:  # package import path
    from .aggregation import (
        run_aggregations as _run_aggregations,
        filter_rows as _filter_rows,
        extract_year as _extract_year,
        safe_sum as _safe_sum,
    )
except ImportError:  # top-level runtime import path used by uvicorn main:app
    from aggregation import (
        run_aggregations as _run_aggregations,
        filter_rows as _filter_rows,
        extract_year as _extract_year,
        safe_sum as _safe_sum,
    )

# ---------------------------------------------------------------------------
# LOGGING
# ---------------------------------------------------------------------------
logger = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# PATHS
# ---------------------------------------------------------------------------
# Resolve /data directory relative to this file's location.
# In deployment: backend/ sits next to data/ at repo root.
# Staging/Render can override using SCIP_SOURCE_ROOT or DATA_DIR.
_HERE = Path(__file__).parent
DEFAULT_DATA_DIR = (_HERE.parent / "data").resolve()
DATA_DIR = Path(
    os.environ.get("SCIP_SOURCE_ROOT")
    or os.environ.get("DATA_DIR")
    or DEFAULT_DATA_DIR
).resolve()
PIPELINE_CONFIG_PATH = (_HERE / "pipeline_config.json").resolve()
logger.info("SCIP data source directory resolved to: %s", DATA_DIR)

# ---------------------------------------------------------------------------
# STAGING/UAT CACHE
# ---------------------------------------------------------------------------
# Render Free + Google Drive bootstrap + Excel parsing can make every request
# expensive if endpoints call load_all() directly. The cache below keeps the
# lineage-bearing server payload in memory for a short TTL. It does not compute
# or fabricate business numbers; it only reuses the last trusted loader output.
_DATA_CACHE_LOCK = threading.RLock()
_DATA_CACHE_PAYLOAD: Optional[dict] = None
_DATA_CACHE_LOADED_AT: Optional[float] = None
_DATA_CACHE_DIR: Optional[str] = None
_DATA_CACHE_HITS = 0
_DATA_CACHE_MISSES = 0


def _cache_ttl_seconds() -> int:
    try:
        return int(os.environ.get("SCIP_DATA_CACHE_TTL_SECONDS", "900"))
    except ValueError:
        return 900


def _resolve_active_data_dir(data_dir: Optional[Path | str] = None) -> Path:
    if data_dir:
        return Path(data_dir).resolve()
    return Path(
        os.environ.get("SCIP_SOURCE_ROOT")
        or os.environ.get("DATA_DIR")
        or DATA_DIR
    ).resolve()


# ---------------------------------------------------------------------------
# W0 INSTRUMENTATION — load timing capture (log/status only; payload unchanged)
# ---------------------------------------------------------------------------
_LAST_LOAD_DURATION_SECONDS: Optional[float] = None
_LAST_LOAD_TIMINGS: dict = {}


def _last_load_timings_summary(top_n: int = 5) -> dict:
    if not _LAST_LOAD_TIMINGS:
        return {}
    sources_ms = _LAST_LOAD_TIMINGS.get("sources_ms", {})
    slowest = sorted(sources_ms.items(), key=lambda kv: kv[1], reverse=True)[:top_n]
    return {
        "total_ms": _LAST_LOAD_TIMINGS.get("total_ms"),
        "resolve_r_series_ms": _LAST_LOAD_TIMINGS.get("resolve_r_series_ms"),
        "build_computed_ms": _LAST_LOAD_TIMINGS.get("build_computed_ms"),
        "build_summary_ms": _LAST_LOAD_TIMINGS.get("build_summary_ms"),
        "slowest_sources": [{"source": rid, "ms": ms} for rid, ms in slowest],
    }


# ---------------------------------------------------------------------------
# SECTION 1 — CONFIG LOADER
# ---------------------------------------------------------------------------

def _load_pipeline_config() -> dict:
    """Load pipeline_config.json. Raises on missing file — config is required."""
    with open(PIPELINE_CONFIG_PATH, "r", encoding="utf-8") as f:
        return json.load(f)


def _load_source_with_adapter(source_id: str, file_path: Path, source_cfg: dict) -> dict:
    """
    Load complex R-series sources through source-specific adapters.

    Batch 3 implements R18, R04, R02, R08, and R36 as production adapters because
    they are positional/matrix based and cannot be trusted through the
    generic flat header-reader.
    """
    try:
        adapter_result = source_adapters.run_adapter(source_id, file_path).to_dict()
    except Exception as exc:
        logger.exception("Adapter load failed for %s", source_id)
        return {
            "id": source_id,
            "status": "error",
            "rows": [],
            "aggs": {},
            "lineage": {},
            "validations": {},
            "errors": [str(exc)],
            "data_date": source_cfg.get("data_date"),
        }

    status_map = {"ok": "ok", "warning": "ok", "unavailable": "missing", "error": "error"}
    metrics = adapter_result.get("metrics", {}) or {}
    return {
        "id": source_id,
        "status": status_map.get(adapter_result.get("status"), "error"),
        "adapter_status": adapter_result.get("status"),
        "rows": adapter_result.get("rows", []),
        "aggs": metrics,
        "lineage": adapter_result.get("lineage", {}),
        "validations": adapter_result.get("validations", {}),
        "warnings": adapter_result.get("warnings", []),
        "errors": adapter_result.get("errors", []),
        "data_date": metrics.get("SNAPSHOT_DATE") or source_cfg.get("data_date"),
    }


# ---------------------------------------------------------------------------
# SECTION 2 — OPENPYXL SHEET READER
# ---------------------------------------------------------------------------

def __old_read_sheet(file_path: Path, sheet_name: Any) -> Optional[list[dict]]:
    """
    Read a single xlsx sheet into a list of row dicts.
    Uses openpyxl read_only mode for performance.

    Returns:
        List of dicts (header → value) or None on failure.
    """
    try:
        wb = openpyxl.load_workbook(str(file_path), read_only=True, data_only=True)

        # Resolve sheet name
        if isinstance(sheet_name, int):
            ws = wb.worksheets[sheet_name]
        elif isinstance(sheet_name, str):
            if sheet_name not in wb.sheetnames:
                logger.warning("Sheet '%s' not found in %s. Available: %s",
                               sheet_name, file_path.name, wb.sheetnames)
                wb.close()
                return None
            ws = wb[sheet_name]
        else:
            ws = wb.active

        rows = list(ws.iter_rows(values_only=True))
        wb.close()

        if not rows:
            return []

        headers = [str(h).strip() if h is not None else f"col_{i}"
                   for i, h in enumerate(rows[0])]
        result = []
        for row in rows[1:]:
            if all(v is None for v in row):
                continue  # skip blank rows
            result.append(dict(zip(headers, row)))
        return result

    except FileNotFoundError:
        logger.info("File not found (graceful skip): %s", file_path.name)
        return None
    except Exception as exc:
        logger.warning("Failed to read %s sheet=%s: %s", file_path.name, sheet_name, exc)
        return None


# ---------------------------------------------------------------------------
# SECTION 3 — COLUMN MAPPER & TYPE COERCER
# ---------------------------------------------------------------------------

def __old_apply_column_map(rows: list[dict], col_map: dict, col_types: dict) -> list[dict]:
    """
    Rename columns per pipeline_config column_map and coerce types.
    Rows that don't have any mapped column are skipped.
    """
    mapped = []
    for row in rows:
        new_row: dict = {}
        for src_col, dest_col in col_map.items():
            raw = row.get(src_col)
            coerced = _coerce(raw, col_types.get(dest_col, "str"))
            new_row[dest_col] = coerced
        if any(v is not None for v in new_row.values()):
            mapped.append(new_row)
    return mapped


def __old_coerce(value: Any, type_str: str) -> Any:
    """
    Coerce a raw openpyxl cell value to the declared type.
    Returns None on conversion failure — never raises.
    """
    if value is None:
        return None
    try:
        if type_str == "int":
            return int(float(value))
        if type_str == "float":
            return float(value)
        if type_str == "bool":
            if isinstance(value, bool):
                return value
            return str(value).strip().lower() in ("true", "yes", "1")
        if type_str == "datetime":
            if isinstance(value, (datetime, date)):
                return value
            # Try parsing common string formats
            for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%Y/%m/%d", "%d %b %Y"):
                try:
                    return datetime.strptime(str(value).strip(), fmt)
                except ValueError:
                    continue
            return None
        # Default: str
        return str(value).strip() if value is not None else None
    except (ValueError, TypeError):
        return None


# ---------------------------------------------------------------------------
# SECTION 4 — AGGREGATION ENGINE
# ---------------------------------------------------------------------------

def __old_run_aggregations(rows: list[dict], agg_cfg: dict) -> dict:
    """
    Execute the aggregation operations declared in pipeline_config for a source.

    Supported ops:
      sum            → sum a column
      mean           → mean of a column
      filter_sum     → sum after filtering
      filter_mean    → mean after filtering
      filter_count   → count rows after filtering
      count_filter   → same as filter_count (alias)
      last           → last non-null value
      extract_list   → extract column as list
      weighted_avg   → sum(numerator) / sum(denominator) * 100
      derived        → computed from other aggregation results (deferred)
    """
    results: dict = {}

    for agg_name, agg_def in agg_cfg.items():
        op = agg_def.get("op")
        col = agg_def.get("col")
        filt = agg_def.get("filter", {})

        try:
            filtered = _filter_rows(rows, filt) if filt else rows

            if op == "sum":
                results[agg_name] = _safe_sum(filtered, col)

            elif op == "filter_sum":
                results[agg_name] = _safe_sum(filtered, col)

            elif op == "mean":
                vals = [r[col] for r in filtered if r.get(col) is not None]
                results[agg_name] = sum(vals) / len(vals) if vals else None

            elif op == "filter_mean":
                vals = [r[col] for r in filtered if r.get(col) is not None]
                results[agg_name] = sum(vals) / len(vals) if vals else None

            elif op in ("filter_count", "count_filter"):
                results[agg_name] = sum(
                    1 for r in filtered if r.get(col) is not None
                )

            elif op == "last":
                vals = [r[col] for r in filtered if r.get(col) is not None]
                results[agg_name] = vals[-1] if vals else None

            elif op == "extract_list":
                fmt = agg_def.get("format", "raw")
                vals = []
                for r in filtered:
                    v = r.get(col)
                    if v is None:
                        continue
                    if fmt == "DD MMM" and isinstance(v, (datetime, date)):
                        vals.append(v.strftime("%d %b"))
                    else:
                        vals.append(v)
                results[agg_name] = vals

            elif op == "weighted_avg":
                num_col = agg_def.get("numerator")
                den_col = agg_def.get("denominator")
                multiply = agg_def.get("multiply_100", False)
                total_num = _safe_sum(filtered, num_col)
                total_den = _safe_sum(filtered, den_col)
                if total_den and total_den > 0:
                    ratio = total_num / total_den
                    results[agg_name] = ratio * 100 if multiply else ratio
                else:
                    results[agg_name] = None

            elif op == "derived":
                # Deferred — resolved after all sources loaded
                results[agg_name] = {"__deferred__": agg_def.get("formula")}

            elif op == "read_metadata":
                # SNAPSHOT_DATE — resolved from source cfg data_date field
                results[agg_name] = {"__metadata__": agg_def.get("field")}

        except Exception as exc:
            logger.warning("Aggregation '%s' failed: %s", agg_name, exc)
            results[agg_name] = None

    return results


def __old_filter_rows(rows: list[dict], filt: dict) -> list[dict]:
    """
    Apply filter conditions to rows.
    Supported filter keys:
      col_name: value          → exact match (str comparison lowercase)
      col_name_in: [v1, v2]   → membership test
      col_name_ne: value       → not-equal
      entity_section: "A-E"   → special: maps Sobha section codes
      year: int                → filter numeric year column
      month_in: [str, ...]     → filter datetime month strings
    """
    result = []
    for row in rows:
        match = True
        for key, val in filt.items():
            if key == "entity_section":
                # R18 special: sections A-E = Sobha, F = Siniya, G = DT
                entity = row.get("entity", "")
                section_map = {
                    "A-E": lambda e: str(e).strip().upper() in ("A","B","C","D","E","SOBHA"),
                    "F":   lambda e: str(e).strip().upper() in ("F","SINIYA"),
                    "G":   lambda e: str(e).strip().upper() in ("G","DT","DOWNTOWN"),
                }
                checker = section_map.get(val)
                if checker and not checker(entity):
                    match = False
            elif key.endswith("_in"):
                col = key[:-3]
                cell = row.get(col)
                if cell is None:
                    match = False
                else:
                    cell_str = str(cell)[:7]  # "2026-01" prefix
                    if cell_str not in [str(v)[:7] for v in val]:
                        match = False
            elif key.endswith("_ne"):
                col = key[:-3]
                cell = row.get(col)
                if cell is not None and str(cell).strip().lower() == str(val).lower():
                    match = False
            elif key == "year":
                # Filter by year extracted from datetime or int column
                for yr_col in ("year", "month", "collection_date", "acd_year"):
                    cell = row.get(yr_col)
                    if cell is not None:
                        extracted = _extract_year(cell)
                        if extracted is not None and extracted != int(val):
                            match = False
                        break
            else:
                # Exact match
                cell = row.get(key)
                if cell is None:
                    match = False
                elif str(cell).strip().lower() != str(val).strip().lower():
                    match = False

        if match:
            result.append(row)
    return result


def __old_extract_year(value: Any) -> Optional[int]:
    """Extract 4-digit year from datetime, int, or string."""
    if isinstance(value, (datetime, date)):
        return value.year
    if isinstance(value, int):
        return value if 2000 <= value <= 2100 else None
    try:
        s = str(value)
        if len(s) >= 4:
            yr = int(s[:4])
            if 2000 <= yr <= 2100:
                return yr
    except (ValueError, TypeError):
        pass
    return None


def __old_safe_sum(rows: list[dict], col: str) -> float:
    """Sum a numeric column, skipping None values."""
    return sum(r[col] for r in rows if r.get(col) is not None)


# ---------------------------------------------------------------------------
# SECTION 5 — SINGLE SOURCE LOADER
# ---------------------------------------------------------------------------

def _load_source(source_id: str, source_cfg: dict, data_dir: Path, resolved_files: dict = None) -> dict:
    """
    Load one R-series source: read → map → aggregate.

    Returns:
        {
            "id":      "R18",
            "status":  "ok" | "missing" | "error",
            "rows":    [...],          # mapped rows
            "aggs":    {...},          # aggregation results
            "data_date": "2026-03-15"
        }
    """
    # Use file_resolver prefix matching (preferred) or fall back to config filename
    if resolved_files and source_id in resolved_files:
        file_path = Path(resolved_files[source_id])
    else:
        fname = source_cfg["file"]
        candidates = [data_dir / fname, data_dir / fname.lower(), data_dir / fname.upper()]
        file_path = next((p for p in candidates if p.exists()), data_dir / fname)
    adapter_name = source_cfg.get("adapter_class")
    if source_id.upper() in getattr(source_adapters, "ADAPTERS", {}) or adapter_name in {
        "R18OverdueAdapter",
        "R04FinanceDailyAdapter",
        "R02MDOAdapter",
        "R08AdvanceAdapter",
        "R36MilestoneCohortAdapter",
    }:
        return _load_source_with_adapter(source_id.upper(), file_path, source_cfg)

    sheet_name = source_cfg.get("sheet_name", 0)
    col_map = source_cfg.get("column_map", {})
    col_types = source_cfg.get("column_types", {})
    agg_cfg = source_cfg.get("aggregations", {})

    raw_rows = _read_sheet(file_path, sheet_name)

    if raw_rows is None:
        logger.info("Source %s not available — graceful degradation", source_id)
        return {
            "id": source_id,
            "status": "missing",
            "rows": [],
            "aggs": {},
            "data_date": source_cfg.get("data_date", None),
        }

    # Apply column map + type coercion
    mapped_rows = _apply_column_map(raw_rows, col_map, col_types) if col_map else []

    # Inject fixed entity if declared
    fixed_entity = source_cfg.get("fixed_entity")
    if fixed_entity:
        for row in mapped_rows:
            row["entity"] = fixed_entity

    # Apply post_load sort (R30 collector chart)
    post_load = source_cfg.get("post_load", {})
    if "sort" in post_load:
        sort_col = post_load["sort"]["col"]
        sort_desc = post_load["sort"]["order"] == "desc"
        mapped_rows.sort(
            key=lambda r: r.get(sort_col) if r.get(sort_col) is not None else -999,
            reverse=sort_desc,
        )

    # Run aggregations
    aggs = _run_aggregations(mapped_rows, agg_cfg)

    # Read data_date from config (SNAPSHOT_DATE resolution)
    data_date = source_cfg.get("data_date")

    return {
        "id": source_id,
        "status": "ok",
        "rows": mapped_rows,
        "aggs": aggs,
        "data_date": data_date,
    }


# ---------------------------------------------------------------------------
# SECTION 6 — DERIVED CONSTANT RESOLVER
# ---------------------------------------------------------------------------

def _resolve_derived_constants(computed: dict) -> dict:
    """
    Resolve constants that depend on other already-computed values.
    Currently: PIPELINE_ADV_DENOM = PIPELINE_GROSS - 5.7B offset.
    """
    pg = computed.get("PIPELINE_GROSS")
    if pg is not None:
        computed["PIPELINE_ADV_DENOM"] = pg - C.PIPELINE_ADV_DENOM_OFFSET_AED

    # PIPELINE_FORWARD_BOOK is a narrative constant — always use structural value
    computed["PIPELINE_FORWARD_BOOK"] = C.GROUP_SALE_VALUE_ITD_AED - C.GROUP_COLLECTED_ITD_AED
    # Override with known approximate: ~40B as per architecture
    computed["PIPELINE_FORWARD_BOOK"] = 40_000_000_000

    return computed


# ---------------------------------------------------------------------------
# SECTION 7 — COMPUTED DICT BUILDER
# ---------------------------------------------------------------------------

def _build_computed(sources: dict) -> dict:
    """
    Build the platform-wide computed dict from all loaded source aggregations.
    This is the authoritative set of derived constants consumed by all endpoints.

    Order follows pipeline_config computed_constants_order:
      1.  OD_TODAY       ← R18
      2.  OD_SOBHA       ← R18
      3.  OD_SINIYA      ← R18
      4.  OD_DT          ← R18
      5.  SNAPSHOT_DATE  ← R18 metadata
      6.  PIPELINE_GROSS ← R36
      7.  PIPELINE_ADV_DENOM ← R36 derived
      8.  CY_ADV_MIX_YTD ← R08 (computed ONCE — never per-submodule)
      9.  AVG_ADVANCE_LEAD_DAYS ← R08
      10. DAILY_DAYS     ← R04
    """
    computed: dict = {}

    # ── R18: OD constants ──────────────────────────────────────────────────
    r18 = sources.get("R18", {})
    r18_aggs = r18.get("aggs", {})

    # Batch 1 hierarchy: Group -> Sobha(Sobha Dubai, Sobha AUH) + UAQ(Siniya, Downtown UAQ)
    computed["OD_SOBHA_DUBAI"] = r18_aggs.get("OD_SOBHA_DUBAI")
    computed["OD_SOBHA_AUH"] = r18_aggs.get("OD_SOBHA_AUH")
    computed["OD_SOBHA"] = r18_aggs.get("OD_SOBHA")
    computed["OD_SINIYA"] = r18_aggs.get("OD_SINIYA")
    computed["OD_DOWNTOWN_UAQ"] = r18_aggs.get("OD_DOWNTOWN_UAQ")
    computed["OD_DT"] = r18_aggs.get("OD_DT") or computed.get("OD_DOWNTOWN_UAQ")  # backward-compatible alias
    computed["OD_UAQ"] = r18_aggs.get("OD_UAQ")
    computed["OD_GROUP"] = r18_aggs.get("OD_GROUP")
    computed["OD_TODAY"] = r18_aggs.get("OD_TODAY")

    # No silent fallback for Board/CXO trust layer. If R18 is missing, expose unavailable state.
    if computed["OD_TODAY"] is None:
        computed["OD_SOURCE"] = "unavailable"
        computed["OD_CONFIDENCE"] = "unavailable"
        computed["OD_FALLBACK_POLICY"] = "no_silent_fallback"
        logger.warning("R18 unavailable — OD metrics marked unavailable; no silent fallback applied")
    else:
        computed["OD_SOURCE"] = "R18_live"
        adapter_status = r18.get("adapter_status")
        computed["OD_CONFIDENCE"] = "live_validated" if adapter_status == "ok" else "live_warning"
        computed["OD_FALLBACK_POLICY"] = "live_or_labelled_unavailable"

    computed["OD_LINEAGE"] = r18.get("lineage", {})
    computed["OD_VALIDATIONS"] = r18.get("validations", {})

    # SNAPSHOT_DATE
    computed["SNAPSHOT_DATE"] = r18_aggs.get("SNAPSHOT_DATE") or r18.get("data_date")

    # OD ageing buckets
    computed["OD_AGEING"] = {
        "0-30d":    r18_aggs.get("ageing_0_30"),
        "31-60d":   r18_aggs.get("ageing_31_60"),
        "61-90d":   r18_aggs.get("ageing_61_90"),
        "91-120d":  r18_aggs.get("ageing_91_120"),
        "121-180d": r18_aggs.get("ageing_121_180"),
        "180+d":    r18_aggs.get("ageing_180plus"),
    }
    # No silent ageing fallback in Batch 1. Missing ageing buckets remain None and are
    # explained through OD_VALIDATIONS / OD_LINEAGE in the trust layer.

    # ── R36: Pipeline constants ────────────────────────────────────────────
    r36 = sources.get("R36", {})
    r36_aggs = r36.get("aggs", {})

    computed["PIPELINE_GROSS"] = r36_aggs.get("PIPELINE_GROSS")
    computed["PIPELINE_TOTAL_FORWARD_CALENDAR"] = r36_aggs.get("PIPELINE_TOTAL_FORWARD_CALENDAR")
    computed["PIPELINE_ACTIVE_TOTAL_PURCHASE_PRICE"] = r36_aggs.get("PIPELINE_ACTIVE_TOTAL_PURCHASE_PRICE")
    computed["PIPELINE_TOTAL_PURCHASE_PRICE"] = r36_aggs.get("PIPELINE_TOTAL_PURCHASE_PRICE")
    computed["FORWARD_COLLECTIBLE_CALENDAR"] = r36_aggs.get("FORWARD_COLLECTIBLE_CALENDAR") or {}
    computed["TOTAL_FORWARD_COLLECTIBLE_CALENDAR"] = r36_aggs.get("TOTAL_FORWARD_COLLECTIBLE_CALENDAR") or {}
    computed["R36_ENTITY_TOTALS"] = r36_aggs.get("R36_ENTITY_TOTALS") or {}
    computed["R36_LINEAGE"] = r36.get("lineage", {})
    computed["R36_VALIDATIONS"] = r36.get("validations", {})
    if computed["PIPELINE_GROSS"] is None:
        computed["PIPELINE_SOURCE"] = "unavailable_no_silent_fallback"
        computed["PIPELINE_CONFIDENCE"] = "unavailable"
        logger.warning("R36 unavailable — pipeline metrics marked unavailable; no silent fallback applied")
    else:
        computed["PIPELINE_SOURCE"] = "R36_live"
        computed["PIPELINE_CONFIDENCE"] = "live_validated" if r36.get("adapter_status") == "ok" else "live_warning"

    # Labels — always attached, never display pipeline figure without label
    computed["PIPELINE_GROSS_LABEL"]        = r36_aggs.get("PIPELINE_GROSS_LABEL") or C.PIPELINE_GROSS_LABEL
    computed["PIPELINE_ADV_DENOM_LABEL"]    = C.PIPELINE_ADV_DENOM_LABEL
    computed["PIPELINE_FORWARD_BOOK_LABEL"] = C.PIPELINE_FORWARD_BOOK_LABEL

    # ── R08: Advance constants ─────────────────────────────────────────────
    r08 = sources.get("R08", {})
    r08_aggs = r08.get("aggs", {})

    # CY_ADV_MIX_YTD — computed ONCE here, NEVER per-submodule.
    # Batch 3: no silent fallback for Board/CXO trust layer.
    computed["CY_ADV_MIX_YTD"] = r08_aggs.get("CY_ADV_MIX_YTD")
    computed["FY_ADV_MIX_YTD"] = r08_aggs.get("FY_ADV_MIX_YTD")
    computed["ADVANCE_2026_TOTAL"] = r08_aggs.get("ADVANCE_2026_TOTAL")
    computed["ADVANCE_2026_CY"] = r08_aggs.get("ADVANCE_2026_CY")
    computed["ADVANCE_2026_FY"] = r08_aggs.get("ADVANCE_2026_FY")
    if computed["CY_ADV_MIX_YTD"] is None:
        computed["R08_SOURCE"] = "unavailable_no_silent_fallback"
        computed["R08_CONFIDENCE"] = "unavailable"
        logger.warning("R08 unavailable — CY/FY advance mix marked unavailable; no silent fallback applied")
    else:
        computed["R08_SOURCE"] = "R08_live"
        computed["R08_CONFIDENCE"] = "live_validated" if r08.get("adapter_status") == "ok" else "live_warning"

    computed["AVG_ADVANCE_LEAD_DAYS"] = r08_aggs.get("AVG_ADVANCE_LEAD_DAYS")
    computed["AVG_ADVANCE_LEAD_DAYS_STATUS"] = "live" if computed["AVG_ADVANCE_LEAD_DAYS"] is not None else "unavailable_in_batch3_no_silent_fallback"
    computed["ADVANCE_2025_TOTAL"]    = r08_aggs.get("advance_2025_total")
    computed["YTD_2026_REBATE"]       = r08_aggs.get("ytd_2026_rebate")
    computed["YTD_2026_REBATE_ESTIMATED"] = r08_aggs.get("ytd_2026_rebate_estimated")
    computed["YTD_2026_ADVANCE_WITH_REBATES"] = r08_aggs.get("ytd_2026_advance_with_rebates")
    computed["YTD_2026_ADVANCE"]      = r08_aggs.get("ytd_2026_advance")
    computed["ADVANCE_2026_SOBHA_PARENT_TOTAL"] = r08_aggs.get("advance_2026_sobha_parent_total")
    computed["ADVANCE_2026_SINIYA_TOTAL"] = r08_aggs.get("advance_2026_siniya_total")
    computed["ADVANCE_2026_DOWNTOWN_UAQ_TOTAL"] = r08_aggs.get("advance_2026_downtown_uaq_total")
    computed["ADVANCE_2026_UAQ_TOTAL"] = r08_aggs.get("advance_2026_uaq_total")
    computed["R08_LINEAGE"] = r08.get("lineage", {})
    computed["R08_VALIDATIONS"] = r08.get("validations", {})

    # ── R04: Daily arrays ─────────────────────────────────────────────────
    r04 = sources.get("R04", {})
    r04_aggs = r04.get("aggs", {})

    # DAILY_DAYS — single shared array, replaces 4 v5 duplicate declarations
    # Batch 2: R04 is Finance basis. "Collection Due" is treated as Finance D+A / collection-due.
    computed["DAILY_DAYS"] = r04_aggs.get("DAILY_DAYS") or []
    computed["DAILY_FINANCE_DA_SERIES"] = r04_aggs.get("DAILY_FINANCE_DA_SERIES") or []
    computed["DAILY_NEW_SALES_SERIES"] = r04_aggs.get("DAILY_NEW_SALES_SERIES") or []
    computed["DAILY_TOTAL_COLLECTIONS_SERIES"] = r04_aggs.get("DAILY_TOTAL_COLLECTIONS_SERIES") or []
    computed["MTD_DA_TOTAL"]    = r04_aggs.get("mtd_da_total")
    computed["MTD_DUES_TOTAL"]  = r04_aggs.get("mtd_dues_total")  # compatibility alias for Finance D+A/collection-due
    computed["MTD_ADV_TOTAL"]   = r04_aggs.get("mtd_advance_total")
    computed["MTD_ADV_STATUS"]  = r04_aggs.get("mtd_advance_status")
    computed["MTD_NS_TOTAL"]    = r04_aggs.get("mtd_ns_total")
    computed["MTD_TOTAL_COLLECTIONS"] = r04_aggs.get("mtd_total_collections")
    computed["MTD_DLD_OQOOD_TOTAL"] = r04_aggs.get("mtd_dld_oqood_total")
    computed["R04_LINEAGE"] = r04.get("lineage", {})
    computed["R04_VALIDATIONS"] = r04.get("validations", {})
    computed["R04_REPORTING_BASIS"] = "Finance"

    # ── R02: MDO targets ──────────────────────────────────────────────────
    r02 = sources.get("R02", {})
    r02_aggs = r02.get("aggs", {})

    computed["FY_DUES_TARGET"]    = r02_aggs.get("fy_dues_target_group")
    computed["FY_ADV_TARGET"]     = r02_aggs.get("fy_advance_target_group")
    computed["FY_DA_TARGET"]      = r02_aggs.get("fy_da_target_group")
    computed["FY_NEW_SALES_TARGET"] = r02_aggs.get("fy_new_sales_target_group")
    computed["FY_TOTAL_COLLECTIONS_TARGET"] = r02_aggs.get("fy_total_collections_target_group")
    computed["MAY_DUES_TARGET"]   = r02_aggs.get("may_dues_target_group")
    computed["MAY_ADV_TARGET"]    = r02_aggs.get("may_advance_target_group")
    computed["MAY_DA_TARGET"]     = r02_aggs.get("may_da_target_group")
    computed["MAY_NEW_SALES_TARGET"] = r02_aggs.get("may_new_sales_target_group")
    computed["MAY_TOTAL_COLLECTIONS_TARGET"] = r02_aggs.get("may_total_collections_target_group")
    computed["MAY_DUES_ACTUAL_MDO"] = r02_aggs.get("may_dues_actual_group")
    computed["MAY_ADV_ACTUAL_MDO"] = r02_aggs.get("may_advance_actual_group")
    computed["MAY_DA_ACTUAL_MDO"] = r02_aggs.get("may_da_actual_group")
    computed["MAY_NEW_SALES_ACTUAL_MDO"] = r02_aggs.get("may_new_sales_actual_group")
    computed["MAY_TOTAL_COLLECTIONS_ACTUAL_MDO"] = r02_aggs.get("may_total_collections_actual_group")
    computed["Q1_DUES_ACTUAL"]    = r02_aggs.get("q1_dues_actual_group")
    computed["Q1_ADV_ACTUAL"]     = r02_aggs.get("q1_advance_actual_group")
    computed["R02_LINEAGE"] = r02.get("lineage", {})
    computed["R02_VALIDATIONS"] = r02.get("validations", {})
    computed["R02_REPORTING_BASIS"] = "MDO"

    # ── R13: ITD portfolio ────────────────────────────────────────────────
    r13 = sources.get("R13", {})
    r13_aggs = r13.get("aggs", {})

    computed["GROUP_SALE_VALUE_ITD"] = (
        r13_aggs.get("group_sale_value_itd") or C.GROUP_SALE_VALUE_ITD_AED
    )
    computed["GROUP_COLLECTED_ITD"] = (
        r13_aggs.get("group_collected_itd") or C.GROUP_COLLECTED_ITD_AED
    )

    # ── R10: Coverage ─────────────────────────────────────────────────────
    r10 = sources.get("R10", {})
    r10_aggs = r10.get("aggs", {})

    computed["SINIYA_UNWORKED_POOL"] = (
        r10_aggs.get("siniya_unworked_pool_aed") or C.SINIYA_UNWORKED_POOL_AED
    )

    # ── R30: Collector metrics ────────────────────────────────────────────
    r30 = sources.get("R30", {})
    r30_aggs = r30.get("aggs", {})

    computed["COLLECTOR_TEAM_AVG_ACH"] = r30_aggs.get("team_avg_achievement")
    computed["COLLECTOR_TOTAL_ACTUAL"] = r30_aggs.get("total_actual")
    computed["COLLECTOR_TOTAL_TARGET"] = r30_aggs.get("total_target")

    # ── R26: LP / charges ─────────────────────────────────────────────────
    r26 = sources.get("R26", {})
    r26_aggs = r26.get("aggs", {})

    computed["LP_2021"] = r26_aggs.get("lp_2021")
    computed["LP_2025"] = r26_aggs.get("lp_2025")
    computed["DLD_2026_YTD"] = r26_aggs.get("dld_2026_ytd")

    # ── R34: Termination gap ──────────────────────────────────────────────
    r34 = sources.get("R34", {})
    r34_aggs = r34.get("aggs", {})

    computed["TERM_NOT_IN_SYSTEM"]  = (
        r34_aggs.get("latest_not_in_system") or C.TERMINATION_GAP_UNITS
    )
    computed["TERM_GAP_EXPOSURE"]   = (
        r34_aggs.get("latest_gap_exposure") or C.TERMINATION_GAP_EXPOSURE_AED
    )

    # ── R38: Risk / IC threshold ──────────────────────────────────────────
    r38 = sources.get("R38", {})
    r38_aggs = r38.get("aggs", {})

    computed["IC_BAND_0_10_UNITS"]    = (
        r38_aggs.get("ic_band_0_10_units") or C.IC_THRESHOLD_UNITS
    )
    computed["IC_BAND_0_10_EXPOSURE"] = (
        r38_aggs.get("ic_band_0_10_exposure") or C.IC_THRESHOLD_EXPOSURE_AED
    )

    # ── Resolve derived constants ──────────────────────────────────────────
    computed = _resolve_derived_constants(computed)

    # ── Data currency metadata ─────────────────────────────────────────────
    computed["LOAD_TIMESTAMP"] = datetime.utcnow().isoformat()
    computed["PLATFORM_VERSION"] = C.PLATFORM_VERSION

    return computed


# ---------------------------------------------------------------------------
# SECTION 8 — LEAN SUMMARY BUILDER (~100KB target)
# ---------------------------------------------------------------------------

def _build_summary(sources: dict, computed: dict) -> dict:
    """
    Build the pre-aggregated lean summary JSON for frontend consumption.
    Target: ~100KB total. Contains formatted display values + chart arrays.
    Raw rows are NOT included here — endpoints access them via `dataframes`.
    """
    summary: dict = {
        "meta": {
            "snapshot_date": computed.get("SNAPSHOT_DATE"),
            "load_timestamp": computed.get("LOAD_TIMESTAMP"),
            "platform_version": computed.get("PLATFORM_VERSION"),
            "sources_loaded": [
                rid for rid, s in sources.items() if s.get("status") == "ok"
            ],
            "sources_missing": [
                rid for rid, s in sources.items() if s.get("status") == "missing"
            ],
        },

        # ── Portfolio headline KPIs ──
        "portfolio": {
            "group_sale_value_itd":    computed.get("GROUP_SALE_VALUE_ITD"),
            "group_collected_itd":     computed.get("GROUP_COLLECTED_ITD"),
            "group_collected_itd_pct": U.safe_divide(
                computed.get("GROUP_COLLECTED_ITD") or 0,
                computed.get("GROUP_SALE_VALUE_ITD") or 1
            ) * 100,
            "od_today":  computed.get("OD_TODAY"),
            "od_group":  computed.get("OD_GROUP"),
            "od_sobha":  computed.get("OD_SOBHA"),
            "od_sobha_dubai": computed.get("OD_SOBHA_DUBAI"),
            "od_sobha_auh":   computed.get("OD_SOBHA_AUH"),
            "od_uaq":    computed.get("OD_UAQ"),
            "od_siniya": computed.get("OD_SINIYA"),
            "od_downtown_uaq": computed.get("OD_DOWNTOWN_UAQ"),
            "od_dt":     computed.get("OD_DT"),
            "od_source": computed.get("OD_SOURCE"),
            "od_confidence": computed.get("OD_CONFIDENCE"),
            "od_fallback_policy": computed.get("OD_FALLBACK_POLICY"),
            "od_ageing": computed.get("OD_AGEING"),
            "od_validations": computed.get("OD_VALIDATIONS"),
            "od_lineage": computed.get("OD_LINEAGE"),
            "pipeline_gross":      computed.get("PIPELINE_GROSS"),
            "pipeline_gross_label": computed.get("PIPELINE_GROSS_LABEL"),
            "pipeline_adv_denom":      computed.get("PIPELINE_ADV_DENOM"),
            "pipeline_adv_denom_label": computed.get("PIPELINE_ADV_DENOM_LABEL"),
            "pipeline_forward_book":      computed.get("PIPELINE_FORWARD_BOOK"),
            "pipeline_forward_book_label": computed.get("PIPELINE_FORWARD_BOOK_LABEL"),
            "pipeline_source": computed.get("PIPELINE_SOURCE"),
            "pipeline_confidence": computed.get("PIPELINE_CONFIDENCE"),
            "pipeline_total_forward_calendar": computed.get("PIPELINE_TOTAL_FORWARD_CALENDAR"),
            "pipeline_active_total_purchase_price": computed.get("PIPELINE_ACTIVE_TOTAL_PURCHASE_PRICE"),
            "pipeline_total_purchase_price": computed.get("PIPELINE_TOTAL_PURCHASE_PRICE"),
            "forward_collectible_calendar": computed.get("FORWARD_COLLECTIBLE_CALENDAR"),
            "r36_entity_totals": computed.get("R36_ENTITY_TOTALS"),
            "r36_lineage": computed.get("R36_LINEAGE"),
            "r36_validations": computed.get("R36_VALIDATIONS"),
        },

        # ── Advance KPIs ──
        "advance": {
            "source": computed.get("R08_SOURCE"),
            "confidence": computed.get("R08_CONFIDENCE"),
            "cy_adv_mix_ytd":        computed.get("CY_ADV_MIX_YTD"),
            "fy_adv_mix_ytd":        computed.get("FY_ADV_MIX_YTD"),
            "avg_advance_lead_days": computed.get("AVG_ADVANCE_LEAD_DAYS"),
            "avg_advance_lead_days_status": computed.get("AVG_ADVANCE_LEAD_DAYS_STATUS"),
            "advance_2025_total":    computed.get("ADVANCE_2025_TOTAL"),
            "advance_2026_total":    computed.get("ADVANCE_2026_TOTAL"),
            "advance_2026_cy":       computed.get("ADVANCE_2026_CY"),
            "advance_2026_fy":       computed.get("ADVANCE_2026_FY"),
            "advance_2026_sobha_parent_total": computed.get("ADVANCE_2026_SOBHA_PARENT_TOTAL"),
            "advance_2026_siniya_total": computed.get("ADVANCE_2026_SINIYA_TOTAL"),
            "advance_2026_downtown_uaq_total": computed.get("ADVANCE_2026_DOWNTOWN_UAQ_TOTAL"),
            "advance_2026_uaq_total": computed.get("ADVANCE_2026_UAQ_TOTAL"),
            "ytd_2026_rebate":       computed.get("YTD_2026_REBATE"),
            "ytd_2026_rebate_estimated": computed.get("YTD_2026_REBATE_ESTIMATED"),
            "ytd_2026_advance_with_rebates": computed.get("YTD_2026_ADVANCE_WITH_REBATES"),
            "ytd_2026_advance":      computed.get("YTD_2026_ADVANCE"),
            "lineage": computed.get("R08_LINEAGE"),
            "validations": computed.get("R08_VALIDATIONS"),
        },

        # ── MDO targets ──
        "targets": {
            "reporting_basis": computed.get("R02_REPORTING_BASIS"),
            "fy_dues_target":  computed.get("FY_DUES_TARGET"),
            "fy_adv_target":   computed.get("FY_ADV_TARGET"),
            "fy_da_target":    computed.get("FY_DA_TARGET"),
            "fy_new_sales_target": computed.get("FY_NEW_SALES_TARGET"),
            "fy_total_collections_target": computed.get("FY_TOTAL_COLLECTIONS_TARGET"),
            "may_dues_target": computed.get("MAY_DUES_TARGET"),
            "may_adv_target":  computed.get("MAY_ADV_TARGET"),
            "may_da_target":   computed.get("MAY_DA_TARGET"),
            "may_new_sales_target": computed.get("MAY_NEW_SALES_TARGET"),
            "may_total_collections_target": computed.get("MAY_TOTAL_COLLECTIONS_TARGET"),
            "may_dues_actual_mdo": computed.get("MAY_DUES_ACTUAL_MDO"),
            "may_adv_actual_mdo": computed.get("MAY_ADV_ACTUAL_MDO"),
            "may_da_actual_mdo": computed.get("MAY_DA_ACTUAL_MDO"),
            "may_new_sales_actual_mdo": computed.get("MAY_NEW_SALES_ACTUAL_MDO"),
            "may_total_collections_actual_mdo": computed.get("MAY_TOTAL_COLLECTIONS_ACTUAL_MDO"),
            "q1_dues_actual":  computed.get("Q1_DUES_ACTUAL"),
            "q1_adv_actual":   computed.get("Q1_ADV_ACTUAL"),
            "lineage": computed.get("R02_LINEAGE"),
            "validations": computed.get("R02_VALIDATIONS"),
        },

        # ── Daily arrays (S03 MTD charts) ──
        "daily": {
            "reporting_basis": computed.get("R04_REPORTING_BASIS"),
            "days":         computed.get("DAILY_DAYS"),
            "finance_da_series": computed.get("DAILY_FINANCE_DA_SERIES"),
            "new_sales_series": computed.get("DAILY_NEW_SALES_SERIES"),
            "total_collections_series": computed.get("DAILY_TOTAL_COLLECTIONS_SERIES"),
            "mtd_da":       computed.get("MTD_DA_TOTAL"),
            "mtd_dues":     computed.get("MTD_DUES_TOTAL"),
            "mtd_advance":  computed.get("MTD_ADV_TOTAL"),
            "mtd_advance_status": computed.get("MTD_ADV_STATUS"),
            "mtd_ns":       computed.get("MTD_NS_TOTAL"),
            "mtd_total_collections": computed.get("MTD_TOTAL_COLLECTIONS"),
            "mtd_dld_oqood": computed.get("MTD_DLD_OQOOD_TOTAL"),
            "lineage": computed.get("R04_LINEAGE"),
            "validations": computed.get("R04_VALIDATIONS"),
        },

        # ── Operational KPIs ──
        "ops": {
            "collector_team_avg_ach": computed.get("COLLECTOR_TEAM_AVG_ACH"),
            "siniya_unworked_pool":   computed.get("SINIYA_UNWORKED_POOL"),
            "term_not_in_system":     computed.get("TERM_NOT_IN_SYSTEM"),
            "term_gap_exposure":      computed.get("TERM_GAP_EXPOSURE"),
            "ic_band_0_10_units":     computed.get("IC_BAND_0_10_UNITS"),
            "ic_band_0_10_exposure":  computed.get("IC_BAND_0_10_EXPOSURE"),
        },

        # ── Charges signals ──
        "charges": {
            "lp_2021":     computed.get("LP_2021"),
            "lp_2025":     computed.get("LP_2025"),
            "dld_2026_ytd": computed.get("DLD_2026_YTD"),
        },

        # ── Collector snapshot top/bottom (from R30) ──
        "collectors": _build_collector_summary(sources),
    }

    return summary


def _build_collector_summary(sources: dict) -> dict:
    """Extract top 3 / bottom 3 collectors from R30 rows for KPI_STRIP."""
    r30 = sources.get("R30", {})
    rows = r30.get("rows", [])
    if not rows:
        return {"top": [], "bottom": [], "count": C.COLLECTOR_COUNT}

    # Already sorted DESC by achievement_pct in _load_source post_load
    formatted = [
        {
            "name":           r.get("collector_name"),
            "achievement_pct": r.get("achievement_pct"),
            "actual_aed":     r.get("actual_aed"),
            "target_aed":     r.get("target_aed"),
        }
        for r in rows if r.get("collector_name")
    ]

    return {
        "top":    formatted[:3],
        "bottom": formatted[-3:] if len(formatted) >= 3 else formatted,
        "count":  len(formatted),
        "data_currency": C.SNAPSHOT_LABEL,
    }


# ---------------------------------------------------------------------------
# SECTION 9 — MAIN ENTRY POINT
# ---------------------------------------------------------------------------

def load_all(data_dir: Optional[Path] = None) -> dict:
    """
    Main entry point. Load all active R-series sources, build computed dict,
    build lean summary. Called by FastAPI on startup and on health-check refresh.

    Args:
        data_dir: Override /data directory path (used in testing).

    Returns:
        {
            "dataframes": { "R18": {"rows": [...], "aggs": {...}}, ... },
            "computed":   { "OD_TODAY": 1650100000, ... },
            "summary":    { "portfolio": {...}, "advance": {...}, ... },
            "status":     "ok" | "partial" | "degraded"
        }
    """
    global DATA_DIR, _LAST_LOAD_DURATION_SECONDS, _LAST_LOAD_TIMINGS
    if data_dir:
        DATA_DIR = data_dir

    _t_total = time.perf_counter()
    load_timings: dict = {"sources_ms": {}}

    active_data_dir = data_dir if data_dir else DATA_DIR
    cfg = _load_pipeline_config()
    sources_cfg = cfg.get("sources", {})
    load_priority = cfg.get("load_priority", {})

    # Build ordered load sequence from priority tiers
    ordered_ids: list[str] = []
    for tier in ("1_critical", "2_high", "3_standard", "4_advisory"):
        ordered_ids.extend(load_priority.get(tier, []))
    # Add any not covered by priority list
    for rid in sources_cfg:
        if rid not in ordered_ids:
            ordered_ids.append(rid)

    # Resolve R-series files by prefix matching
    _t = time.perf_counter()
    resolved_files = file_resolver.resolve_r_series(str(active_data_dir))
    load_timings["resolve_r_series_ms"] = round((time.perf_counter() - _t) * 1000, 1)

    # Load sources
    sources: dict = {}
    for rid in ordered_ids:
        if rid not in sources_cfg:
            continue
        logger.debug("Loading %s ...", rid)
        _t = time.perf_counter()
        sources[rid] = _load_source(rid, sources_cfg[rid], data_dir=active_data_dir, resolved_files=resolved_files)
        load_timings["sources_ms"][rid] = round((time.perf_counter() - _t) * 1000, 1)

    # Build computed dict
    _t = time.perf_counter()
    computed = _build_computed(sources)
    load_timings["build_computed_ms"] = round((time.perf_counter() - _t) * 1000, 1)

    # Build lean summary
    _t = time.perf_counter()
    summary = _build_summary(sources, computed)
    load_timings["build_summary_ms"] = round((time.perf_counter() - _t) * 1000, 1)

    # Determine overall status
    missing = [rid for rid, s in sources.items() if s.get("status") == "missing"]
    critical = set(load_priority.get("1_critical", []))
    critical_missing = [rid for rid in missing if rid in critical]

    if critical_missing:
        status = "degraded"
        logger.warning("Critical sources missing: %s — platform in degraded mode", critical_missing)
    elif missing:
        status = "partial"
        logger.info("Non-critical sources missing: %s — platform operating normally", missing)
    else:
        status = "ok"

    logger.info(
        "Data load complete. Status=%s. Sources: %d ok / %d missing.",
        status, len(sources) - len(missing), len(missing)
    )

    load_timings["total_ms"] = round((time.perf_counter() - _t_total) * 1000, 1)
    _LAST_LOAD_DURATION_SECONDS = round(load_timings["total_ms"] / 1000.0, 3)
    _LAST_LOAD_TIMINGS = load_timings
    slowest = sorted(load_timings["sources_ms"].items(), key=lambda kv: kv[1], reverse=True)[:5]
    logger.info(
        "SCIP load_all timings: total=%.0fms resolve=%.0fms computed=%.0fms summary=%.0fms slowest=[%s]",
        load_timings["total_ms"],
        load_timings.get("resolve_r_series_ms", 0.0),
        load_timings.get("build_computed_ms", 0.0),
        load_timings.get("build_summary_ms", 0.0),
        ", ".join("%s:%.0fms" % (rid, ms) for rid, ms in slowest) or "none",
    )

    return {
        "dataframes": sources,
        "computed":   computed,
        "summary":    summary,
        "status":     status,
        "missing_sources": missing,
    }


def get_cached_payload(
    data_dir: Optional[Path | str] = None,
    *,
    force_refresh: bool = False,
    ttl_seconds: Optional[int] = None,
    copy_payload: bool = False,
) -> dict:
    """Return the trusted data loader payload using a short in-memory cache.

    This is the performance-critical entry point for FastAPI endpoints. It keeps
    Quickball, command centres, and forecast from reparsing every Excel workbook
    on each click or role change. The payload still contains the original
    lineage, validation status, confidence state and missing-source disclosure.

    Set SCIP_DATA_CACHE_TTL_SECONDS=0 to disable caching. Use force_refresh=True
    only for an explicit source-refresh action, not for role switch or Quickball.
    """
    global _DATA_CACHE_PAYLOAD, _DATA_CACHE_LOADED_AT, _DATA_CACHE_DIR
    global _DATA_CACHE_HITS, _DATA_CACHE_MISSES

    active_dir = _resolve_active_data_dir(data_dir)
    cache_dir = str(active_dir)
    ttl = _cache_ttl_seconds() if ttl_seconds is None else int(ttl_seconds)
    now = time.time()

    if ttl > 0 and not force_refresh:
        with _DATA_CACHE_LOCK:
            is_valid = (
                _DATA_CACHE_PAYLOAD is not None
                and _DATA_CACHE_LOADED_AT is not None
                and _DATA_CACHE_DIR == cache_dir
                and now - _DATA_CACHE_LOADED_AT <= ttl
            )
            if is_valid:
                _DATA_CACHE_HITS += 1
                logger.debug("SCIP data cache hit: dir=%s age=%.2fs", cache_dir, now - _DATA_CACHE_LOADED_AT)
                return copy.deepcopy(_DATA_CACHE_PAYLOAD) if copy_payload else _DATA_CACHE_PAYLOAD

    with _DATA_CACHE_LOCK:
        now = time.time()
        if ttl > 0 and not force_refresh:
            is_valid = (
                _DATA_CACHE_PAYLOAD is not None
                and _DATA_CACHE_LOADED_AT is not None
                and _DATA_CACHE_DIR == cache_dir
                and now - _DATA_CACHE_LOADED_AT <= ttl
            )
            if is_valid:
                _DATA_CACHE_HITS += 1
                logger.debug("SCIP data cache hit after lock: dir=%s", cache_dir)
                return copy.deepcopy(_DATA_CACHE_PAYLOAD) if copy_payload else _DATA_CACHE_PAYLOAD

        _DATA_CACHE_MISSES += 1
        logger.info("SCIP data cache miss; loading source payload: dir=%s force_refresh=%s", cache_dir, force_refresh)
        payload = load_all(data_dir=active_dir)
        if ttl > 0:
            _DATA_CACHE_PAYLOAD = payload
            _DATA_CACHE_LOADED_AT = time.time()
            _DATA_CACHE_DIR = cache_dir
        return copy.deepcopy(payload) if copy_payload else payload


def invalidate_cache() -> None:
    """Clear the in-memory data payload cache."""
    global _DATA_CACHE_PAYLOAD, _DATA_CACHE_LOADED_AT, _DATA_CACHE_DIR
    with _DATA_CACHE_LOCK:
        _DATA_CACHE_PAYLOAD = None
        _DATA_CACHE_LOADED_AT = None
        _DATA_CACHE_DIR = None


def get_cache_status() -> dict:
    now = time.time()
    age = None if _DATA_CACHE_LOADED_AT is None else round(now - _DATA_CACHE_LOADED_AT, 3)
    return {
        "status": "warm" if _DATA_CACHE_PAYLOAD is not None else "cold",
        "data_dir": _DATA_CACHE_DIR,
        "loaded_at_epoch": _DATA_CACHE_LOADED_AT,
        "age_seconds": age,
        "ttl_seconds": _cache_ttl_seconds(),
        "hits": _DATA_CACHE_HITS,
        "misses": _DATA_CACHE_MISSES,
        "last_load_duration_seconds": _LAST_LOAD_DURATION_SECONDS,
        "last_load_timings": _last_load_timings_summary(),
        "lineage_preserved": True,
        "no_silent_fallback": True,
    }


def get_source_rows(payload: dict, source_id: str) -> list[dict]:
    """
    Helper for endpoint functions to extract rows from a loaded source.
    Returns empty list (not crash) if source unavailable.

    Args:
        payload:    Result of load_all()
        source_id:  e.g. "R18"
    """
    return payload.get("dataframes", {}).get(source_id, {}).get("rows", [])


def get_computed(payload: dict, key: str, fallback: Any = None) -> Any:
    """
    Helper for endpoint functions to read a computed constant.
    Returns fallback if key absent.

    Args:
        payload:  Result of load_all()
        key:      e.g. "OD_TODAY"
        fallback: Default if key not found
    """
    return payload.get("computed", {}).get(key, fallback)
